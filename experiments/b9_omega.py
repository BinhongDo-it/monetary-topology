#!/usr/bin/env python3
"""B9-A: the `ω` machinery and the join reconciliation. §10 step 3.

Registered in ``docs/b9_zero_holonomy.md``: §2 for the graph, §14 for the
inclusive definition of each edge, §13 for the `√N` ruling and the SPDR sample.

**This file computes no reading.** ``--selftest`` exercises the machinery on
hand-computed values and touches no data. ``--depth`` opens the cached workbooks
and prints the join: how many dates each side carries, how many survive the
intersection, how many rows fail the date parse and whether that count reconciles
with §14.5's section furniture. **It estimates nothing and prints no `λ`.**

§10's order puts B9-0 and the F1 audit after this, and B9-A-1 after those.
Running ``--depth`` first is `B10 §12.6`'s rule: look at the objects before any
estimator is allowed near them.

Usage::

    python experiments/b9_omega.py --selftest
    python experiments/b9_omega.py --depth --only XLF
    python experiments/b9_omega.py --depth

Inputs are the workbooks already on disk under ``data/raw/b9/<TICKER>/``.
**Nothing here fetches, and nothing here is ever deleted.**
"""

from __future__ import annotations

import argparse
import json
import math
import io
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CACHE = ROOT / "data" / "raw" / "b9"
OUT = ROOT / "results" / "b9_depth.json"

NAVHIST = "constructed_nav_history"
PDHIST = "constructed_premium_discount_history"

#: §13.5's sample. The retired rows live in the availability probe and enter
#: nothing here.
MAIN_ARM = ["XLB", "XLC", "XLE", "XLF", "XLI", "XLK",
            "XLP", "XLRE", "XLU", "XLV", "XLY"]
COMPARISON_ARM = ["SPDW", "SPEM", "SPAB", "JNK"]
NEITHER = ["SPY"]
SAMPLE = MAIN_ARM + COMPARISON_ARM + NEITHER

# ---------------------------------------------------------------------------
# §14.1 and §14.2: the cochain.
#
# The canonical orientation is the only thing defined. Every reversed edge is
# the negation of its canonical twin, computed by the same function from the
# same state, which is what §14.1 means by "antisymmetry is imposed by
# construction and is not an assumption about the world".
#
# **What B9-0 can and cannot catch, stated plainly.** Because antisymmetry is
# definitional, the degenerate loop cannot come out non-zero for an empirical
# reason. It can come out non-zero for a coding reason: a state dict rebuilt
# between the two legs, a factor read from the wrong date, a sum that
# special-cases the repeated edge. **That is exactly the class of failure it is
# registered to catch**, and it is why §6 forbids short-circuiting it.
# ---------------------------------------------------------------------------

POSITIONS = ("cash", "basket", "etf")

#: edge -> multiplicative factor on the numeraire value carried across it
CANONICAL_FACTORS = {
    ("cash", "basket"): lambda st: 1.0,
    ("basket", "etf"): lambda st: 1.0 / (1.0 + st["fee"]),
    ("etf", "cash"): lambda st: st["price"] / st["nav"],
}

LOOP = ("cash", "basket", "etf", "cash")
DEGENERATE = ("cash", "etf", "cash")


def omega(u: str, v: str, st: dict) -> float:
    """`ω` on the directed edge u→v, in log points, per §14.1."""
    if u not in POSITIONS or v not in POSITIONS:
        raise KeyError(f"not a position: {u!r}, {v!r}")
    if (u, v) in CANONICAL_FACTORS:
        return math.log(CANONICAL_FACTORS[(u, v)](st))
    if (v, u) in CANONICAL_FACTORS:
        return -omega(v, u, st)
    raise KeyError(f"no edge {u}->{v}")


PROBE_FEE = 3.7e-4          # §49.1: an arbitrary non-zero fee for the cross-check


def loop_direct(st: dict) -> float:
    """A second expression for the loop sum, **written without `omega`,
    `path_sum` or `CANONICAL_FACTORS`**, so a defect in any of them shows up as
    a disagreement instead of as a shared answer. §49.1: the gate used to assert
    `-x + x == 0`, which no implementation can fail."""
    return math.log(st["price"]) - math.log(st["nav"]) - math.log1p(st["fee"])


def path_sum(path, st: dict) -> float:
    """Sum `ω` along consecutive positions. **No special case for a repeated
    edge**, which is what makes the degenerate loop a test of this function."""
    return sum(omega(path[i], path[i + 1], st) for i in range(len(path) - 1))


def lam(premium: float, fee: float) -> float:
    """§2's closed form, kept beside `path_sum` so the two can be compared
    rather than trusted. `premium` is a fraction here, not a percent."""
    return math.log(1.0 + premium) - math.log(1.0 + fee)


# ---------------------------------------------------------------------------
# Parsing. §14.5.
#
# The date parser is the one §13.1's availability pass exercised on both files.
# It is written out again here rather than imported, because the availability
# probe is a throwaway retrieval tool and this file is the stage.
# ---------------------------------------------------------------------------

DATE_FORMATS = ("%d-%b-%Y", "%b %d %Y", "%b %d, %Y", "%Y-%m-%d", "%m/%d/%Y")


def as_iso_date(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, str):
        s = v.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                pass
    return None


def _cell(v) -> str:
    """A cell as text, for printing and for recording provenance. Dates come
    back as ISO so a record never carries a locale-dependent string."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, float):
        return f"{v:.10g}"
    return str(v)


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("$", "")
        pct = s.endswith("%")
        if pct:
            s = s[:-1]
        try:
            return float(s)
        except ValueError:
            return None
    return None


def read_sheet(ticker: str, name: str) -> list[tuple]:
    from openpyxl import load_workbook
    p = CACHE / ticker / name
    if not p.is_file():
        return []
    wb = load_workbook(io.BytesIO(p.read_bytes()), read_only=True, data_only=True)
    rows = [r for r in wb.worksheets[0].iter_rows(values_only=True)]
    wb.close()
    return rows


def split_rows(rows: list[tuple], value_cols: int) -> dict:
    """Split into dated rows and everything else. **Nothing is discarded**: the
    undated rows are returned so §14.5's reconciliation can be checked rather
    than asserted."""
    dated, undated, blank = [], [], 0
    for r in rows:
        if not r or all(x is None for x in r):
            blank += 1
            continue
        iso = as_iso_date(r[0])
        if iso is None:
            undated.append(" | ".join(str(x) for x in r if x is not None)[:120])
            continue
        vals = [_num(r[i]) if i < len(r) else None for i in range(1, 1 + value_cols)]
        dated.append((iso, *vals))
    # **The identity, not a subset of it.** Real sheets carry trailing blank
    # rows, so `rows − dated == undated` is false for a correct parser. Every
    # row must land in exactly one of the three.
    assert len(dated) + len(undated) + blank == len(rows)
    return {"dated": dated, "undated": undated, "blank": blank}


def quantiles(xs: list[float]) -> dict:
    if not xs:
        return {}
    s = sorted(xs)
    def q(p):
        i = min(len(s) - 1, max(0, int(round(p * (len(s) - 1)))))
        return s[i]
    return {"n": len(s), "min": s[0], "p01": q(0.01), "p25": q(0.25),
            "p50": q(0.50), "p75": q(0.75), "p99": q(0.99), "max": s[-1]}


# ---------------------------------------------------------------------------
# depth
# ---------------------------------------------------------------------------


def depth(tickers: list[str]) -> dict:
    out: dict = {}
    for t in (tickers or SAMPLE):
        nav_rows = read_sheet(t, NAVHIST)
        pd_rows = read_sheet(t, PDHIST)
        if not nav_rows or not pd_rows:
            print(f"{t:6s} missing workbook on disk. Retrieve it first.")
            continue

        nav = split_rows(nav_rows, value_cols=3)     # NAV, shares out, TNA
        pdh = split_rows(pd_rows, value_cols=1)      # premium/discount

        nav_dates = [d[0] for d in nav["dated"]]
        pd_dates = [d[0] for d in pdh["dated"]]
        nav_set, pd_set = set(nav_dates), set(pd_dates)
        both = nav_set & pd_set

        rec = {
            "navhist": {
                "rows": len(nav_rows), "dated": len(nav["dated"]),
                "undated": len(nav["undated"]),
                "duplicate_dates": len(nav_dates) - len(nav_set),
                "span": [min(nav_set), max(nav_set)] if nav_set else [],
            },
            "pdhist": {
                "rows": len(pd_rows), "dated": len(pdh["dated"]),
                "undated": len(pdh["undated"]),
                "duplicate_dates": len(pd_dates) - len(pd_set),
                "span": [min(pd_set), max(pd_set)] if pd_set else [],
                "undated_sample": pdh["undated"][:6],
            },
            "join": {
                "both": len(both),
                "navhist_only": len(nav_set - pd_set),
                "pdhist_only": len(pd_set - nav_set),
                "span": [min(both), max(both)] if both else [],
            },
            # §14.5's reconciliation: rows that failed the date parse must be
            # the quarterly section furniture and nothing else.
            "reconciliation": {
                "pdhist_rows": len(pd_rows),
                "pdhist_dated": len(pdh["dated"]),
                "pdhist_undated": len(pdh["undated"]),
                "pdhist_blank": pdh["blank"],
                "reconciles": (len(pdh["dated"]) + len(pdh["undated"])
                               + pdh["blank"]) == len(pd_rows),
            },
            # **Units are not decided here.** The premium column is printed as it
            # sits so the unit is read off the spread of the numbers rather than
            # assumed. A percent column and a fraction column differ by 100, and
            # guessing wrong moves every λ by two orders of magnitude.
            "premium_column_raw": quantiles(
                [d[1] for d in pdh["dated"] if d[1] is not None]),
            "nav_column_raw": quantiles(
                [d[1] for d in nav["dated"] if d[1] is not None]),
        }
        out[t] = rec

        j, n, p = rec["join"], rec["navhist"], rec["pdhist"]
        print(f"\n{t}")
        print(f"  navhist  rows={n['rows']:>5} dated={n['dated']:>5} "
              f"undated={n['undated']:>3} dup={n['duplicate_dates']:>2} "
              f"{n['span'][0] if n['span'] else '':>10} .. "
              f"{n['span'][1] if n['span'] else '':>10}")
        print(f"  pdhist   rows={p['rows']:>5} dated={p['dated']:>5} "
              f"undated={p['undated']:>3} dup={p['duplicate_dates']:>2} "
              f"{p['span'][0] if p['span'] else '':>10} .. "
              f"{p['span'][1] if p['span'] else '':>10}")
        print(f"  join     both={j['both']:>5} nav_only={j['navhist_only']:>5} "
              f"pd_only={j['pdhist_only']:>3}   window "
              f"{j['span'][0] if j['span'] else ''} .. "
              f"{j['span'][1] if j['span'] else ''}")
        r = rec["reconciliation"]
        print(f"  §14.5    pdhist {r['pdhist_dated']} dated + "
              f"{r['pdhist_undated']} furniture + {r['pdhist_blank']} blank "
              f"= {r['pdhist_rows']} rows: "
              f"{'reconciles' if r['reconciles'] else '**DOES NOT RECONCILE**'}")
        q = rec["premium_column_raw"]
        if q:
            print(f"  premium  n={q['n']} min={q['min']:.6g} p01={q['p01']:.6g} "
                  f"p50={q['p50']:.6g} p99={q['p99']:.6g} max={q['max']:.6g}")
        if p["undated_sample"]:
            print(f"  undated  {p['undated_sample'][0][:110]}")

    if out:
        OUT.parent.mkdir(parents=True, exist_ok=True)
        OUT.write_text(json.dumps(
            {"diagnostic_only": True,
             "diagnostic_reason": "the --depth of step 3 in section 10. prints only the join reconciliation and the raw column distributions; "
                                  "computes no ω, no λ and no π, and reads no prediction.",
             "stage": "B9-A", "funds": out},
            sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        print(f"\nwrote {OUT.relative_to(ROOT)}")

    bad = [t for t, r in out.items() if not r["reconciliation"]["reconciles"]]
    if bad:
        print(f"\n**{len(bad)} fund(s) fail §14.5's reconciliation: {bad}. "
              f"The parser is wrong and nothing downstream may be read.**")
    print("\nunits of the premium column are NOT decided by this run. "
          "Read them off the spread above and rule on it before any λ.")
    return out


# ---------------------------------------------------------------------------
# B9-0, on real state. §6 and §10 step 3.
#
# **The fee is passed as NaN on purpose.** The degenerate loop traverses only
# `cash → etf` and `etf → cash`, neither of which touches `f`. If a future edit
# routes the sum through the fee-bearing edge, the result is NaN rather than
# zero, and the gate fails loudly instead of passing for the wrong reason.
# §14.6 forbids computing `λ` before `f` is read; this gate is not `λ`.
# ---------------------------------------------------------------------------

GATE_OUT = ROOT / "results" / "b9_gate.json"
F1_OUT = ROOT / "results" / "b9_f1.json"


def gate(tickers: list[str]) -> dict:
    per_fund, worst = {}, 0.0
    for t in (tickers or SAMPLE):
        nav_rows, pd_rows = read_sheet(t, NAVHIST), read_sheet(t, PDHIST)
        if not nav_rows or not pd_rows:
            print(f"{t:6s} missing workbook.")
            continue
        nav = {d[0]: d[1] for d in split_rows(nav_rows, 3)["dated"]}
        pdh = {d[0]: d[1] for d in split_rows(pd_rows, 1)["dated"]}
        n_ok = n_bad = n_skipped = 0
        bad_examples = []
        for day in sorted(set(nav) & set(pdh)):
            nav_v, prem_pct = nav[day], pdh[day]
            if nav_v is None or prem_pct is None or nav_v <= 0:
                n_skipped += 1
                continue
            prem = prem_pct / 100.0                      # §15.1
            if abs(prem) >= 0.25:                        # §15.1's guard
                n_skipped += 1
                bad_examples.append({"date": day, "premium_fraction": prem,
                                     "why": "outside §15.1's band"})
                continue
            st = {"nav": nav_v, "price": nav_v * (1.0 + prem),
                  "fee": float("nan")}
            # **§49.1.** `path_sum(DEGENERATE, st)` is `-x + x` by construction
            # and cannot be non-zero. Kept as a construction check, and it is
            # **no longer what the gate asserts.**
            z = path_sum(DEGENERATE, st)
            # The gate proper, failable three ways: the NaN fee must survive the
            # loop rather than be special-cased away, and the loop sum through
            # the cochain must agree with an expression written **without**
            # `omega`, `path_sum` or `CANONICAL_FACTORS`.
            nan_ok = math.isnan(path_sum(LOOP, st))
            probe = dict(st)
            probe["fee"] = PROBE_FEE
            a = path_sum(LOOP, probe)
            b = loop_direct(probe)
            gap = abs(a - b)
            # **Accumulated on every fund-day, not only on failures.** The first
            # version updated `worst` inside the `else`, so a clean run reported
            # `0.0` and that read as "agrees to the last bit" when it meant "no
            # failure was recorded". Same family as everything §49 found.
            worst = max(worst, gap)
            agree = gap <= 1e-12 * max(1.0, abs(b))
            if z == 0.0 and nan_ok and agree:
                n_ok += 1
            else:
                n_bad += 1
                if len(bad_examples) < 5:
                    bad_examples.append(
                        {"date": day, "z": z, "nan_ok": nan_ok,
                         "cochain": a, "direct": b,
                         "why": ("degenerate non-zero" if z != 0.0
                                 else "NaN fee swallowed" if not nan_ok
                                 else "cochain disagrees with the independent "
                                      "expression")})
        per_fund[t] = {"exact_zero": n_ok, "non_zero": n_bad,
                       "skipped": n_skipped, "examples": bad_examples[:5]}
        flag = "" if n_bad == 0 else "  **B9-0 FAILS**"
        print(f"{t:6s} exact zero on {n_ok:>4d} fund-days, non-zero {n_bad:>3d}, "
              f"skipped {n_skipped:>2d}{flag}")

    passed = all(v["non_zero"] == 0 for v in per_fund.values()) and bool(per_fund)
    rec = {"stage": "B9-0", "diagnostic_only": False,
           "passed": passed, "worst_abs_cochain_vs_direct": worst,
           "worst_note": "max over **every** fund-day, not over failures. "
                         "log(p/n) and log(p)-log(n) are not bit-identical; "
                         "a run reporting exactly 0.0 here is reporting a "
                         "branch artefact, not agreement.",
           "probe_fee": PROBE_FEE,
           "gate_asserts": ["degenerate out-and-back is zero (a construction "
                            "check, §49.1, cannot fail)",
                            "a NaN fee survives the loop sum",
                            "the cochain agrees with an independent expression"],
           "funds": per_fund,
           "note": "the degenerate loop cash→etf→cash goes through path_sum with the fee passed as NaN, "
                   "so any accidental traversal of a fee-bearing edge yields NaN rather than zero."}
    GATE_OUT.parent.mkdir(parents=True, exist_ok=True)
    GATE_OUT.write_text(json.dumps(rec, sort_keys=True, ensure_ascii=False,
                                   indent=2) + "\n", encoding="utf-8", newline="\n")
    print(f"\nwrote {GATE_OUT.relative_to(ROOT)}")
    if passed:
        print("**B9-0 passes.** §6: nothing after it was read before this.")
    else:
        print("**B9-0 fails. The `ω` construction is broken and nothing from "
              "this stage is quotable, including numbers already computed** (§8).")
    return rec


# ---------------------------------------------------------------------------
# F1, the traversability audit. §5.
#
# §5's F1 guard: a loop with only one realisable route sums to zero for want of
# a comparison, not because the field is exact. For B9-A the test is whether the
# primary market moved in the fund-day's neighbourhood, read off the
# shares-outstanding path.
#
# **The threshold is scanned, not chosen** (discipline 8: report the gradient,
# not the level). What "no activity" means is exactly the free parameter that
# would let someone tune the sample until the zero appears, so it is reported as
# a curve and the curve is the object.
# ---------------------------------------------------------------------------

#: shares, absolute. 0 means "any change at all counts".
F1_THRESHOLDS = [0, 1, 1_000, 25_000, 50_000, 100_000, 500_000, 1_000_000]
F1_WINDOW = 5   # trading days either side, §5


def f1_audit(tickers: list[str]) -> dict:
    if not GATE_OUT.is_file():
        print("B9-0 has not run. §10 step 3 puts the gate first.")
        return {}
    if not json.loads(GATE_OUT.read_text(encoding="utf-8")).get("passed"):
        print("B9-0 did not pass. §8: nothing after it may be read.")
        return {}

    out = {}
    for t in (tickers or SAMPLE):
        nav_rows, pd_rows = read_sheet(t, NAVHIST), read_sheet(t, PDHIST)
        if not nav_rows or not pd_rows:
            continue
        nav = sorted(split_rows(nav_rows, 3)["dated"], key=lambda r: r[0])
        pd_days = {d[0] for d in split_rows(pd_rows, 1)["dated"]}
        days = [r[0] for r in nav]
        nav_ps = [r[1] for r in nav]
        raw_shares = [r[2] for r in nav]
        tna = [r[3] for r in nav]

        # === share splits, detected before any change is read ===============
        #
        # **Measured, not assumed.** On 2025-12-05 five Select Sector SPDRs
        # (XLB, XLE, XLK, XLU, XLY) did a 2-for-1 split. XLK: NAV 291.0433 →
        # 146.6173, shares 325,805,897 → 650,611,794, and **total net assets
        # ratio 1.0060, that is continuous**. The third column is the
        # confirmation: a split moves NAV and shares in exact opposition and
        # leaves assets alone, while a data error would not.
        #
        # F1 infers primary-market activity from the change in shares. On a
        # split day that change is the split, and reading it as a creation of
        # 325 million shares is the eighth failure mode, membership error: the
        # quantity is not what its name says on that row. So the series is
        # split-adjusted first and the detection is reported.
        splits = []
        factor = [1.0] * len(days)
        cum = 1.0
        for i in range(1, len(days)):
            a, b = nav_ps[i - 1], nav_ps[i]
            if a and b:
                r = a / b
                k = round(r)
                if k >= 2 and abs(r - k) / k < 0.02:
                    ta, tb = tna[i - 1], tna[i]
                    continuous = (ta and tb and abs(tb / ta - 1.0) < 0.10)
                    splits.append({"date": days[i], "factor": k,
                                   "nav_ratio": round(r, 6),
                                   "tna_ratio": round(tb / ta, 6) if ta and tb else None,
                                   "assets_continuous": bool(continuous)})
                    if continuous:
                        cum *= k
                # a reverse split moves the ratio the other way
                elif r < 0.6:
                    k2 = round(1.0 / r)
                    if k2 >= 2 and abs(1.0 / r - k2) / k2 < 0.02:
                        ta, tb = tna[i - 1], tna[i]
                        continuous = (ta and tb and abs(tb / ta - 1.0) < 0.10)
                        splits.append({"date": days[i], "factor": 1.0 / k2,
                                       "nav_ratio": round(r, 6),
                                       "tna_ratio": round(tb / ta, 6) if ta and tb else None,
                                       "assets_continuous": bool(continuous),
                                       "reverse": True})
                        if continuous:
                            cum /= k2
            factor[i] = cum
        # shares expressed in post-split units throughout, so a change is a
        # change in the fund's size and never an artefact of the unit
        shares = [None if s is None else s / factor[i]
                  for i, s in enumerate(raw_shares)]

        # daily change over the whole navhist series, so a fund-day at the
        # start of the join window still has five real neighbours behind it
        delta = [None]
        non_integral = 0
        for i in range(1, len(shares)):
            a, b = shares[i - 1], shares[i]
            if a is None or b is None:
                delta.append(None)
                continue
            d = b - a
            if abs(d - round(d)) > 1e-6:
                non_integral += 1
            delta.append(abs(round(d)))

        nz = sorted(d for d in delta if d not in (None, 0))
        g = 0
        for d in nz:
            g = math.gcd(g, int(d))
        zero_days = sum(1 for d in delta if d == 0)
        known = sum(1 for d in delta if d is not None)

        idx = {d: i for i, d in enumerate(days)}
        candidates = sorted(set(days) & pd_days)
        scan, truncated = {}, 0
        for tau in F1_THRESHOLDS:
            cleared = 0
            for day in candidates:
                i = idx[day]
                lo, hi = max(0, i - F1_WINDOW), min(len(days) - 1, i + F1_WINDOW)
                if any(delta[k] is not None and delta[k] > tau
                       for k in range(lo, hi + 1)):
                    cleared += 1
            scan[str(tau)] = cleared
        for day in candidates:
            i = idx[day]
            if i - F1_WINDOW < 0 or i + F1_WINDOW > len(days) - 1:
                truncated += 1

        # **The same statistics restricted to the window that actually supplies
        # candidates.** The gcd over the full 22-year series is dragged to 1 by a
        # handful of odd changes, and a creation unit inferred from 2004 says
        # nothing about the unit in force in 2025. Whether τ can be earned from
        # behaviour depends on this block, not on the one above.
        cand_set = set(candidates)
        w_pairs = [(d, delta[i]) for i, d in enumerate(days) if d in cand_set]
        w_delta = [x for _, x in w_pairs]
        w_nz = sorted(x for x in w_delta if x not in (None, 0))
        # **Name the exceptions.** Eleven funds show every in-window change as a
        # multiple of the creation unit; five show exactly one change that is
        # not, and that single value is what drags their gcd to 25, 200, 1, 80
        # and 4. One row per fund is small enough to print and too structured to
        # sweep into a residual.
        w_odd = [{"date": d, "delta": int(x), "shares_after": None}
                 for d, x in w_pairs
                 if x not in (None, 0) and int(x) % 50_000 != 0]
        for o in w_odd:
            i = idx_of = days.index(o["date"])
            o["shares_after"] = shares[i]
            o["shares_before"] = shares[i - 1] if i else None
        w_g = 0
        for x in w_nz:
            w_g = math.gcd(w_g, int(x))
        w_known = sum(1 for x in w_delta if x is not None)
        w_zero = sum(1 for x in w_delta if x == 0)
        in_window = {
            "days": len(w_delta), "known": w_known, "zero": w_zero,
            "zero_share": round(w_zero / w_known, 6) if w_known else None,
            "gcd": w_g,
            "smallest_nonzero": [int(x) for x in w_nz[:5]],
            "multiples_of_50k": sum(1 for x in w_nz if int(x) % 50_000 == 0),
            "nonzero_total": len(w_nz),
            "non_multiple_rows": w_odd,
        }

        rec = {
            "navhist_rows": len(nav), "candidates": len(candidates),
            "splits_detected": splits,
            "in_window": in_window,
            "delta_known": known, "delta_zero": zero_days,
            "delta_zero_share": round(zero_days / known, 6) if known else None,
            "delta_non_integral": non_integral,
            "gcd_of_changes": g,
            "smallest_nonzero_changes": [int(x) for x in nz[:5]],
            "delta_quantiles": quantiles([float(x) for x in nz]),
            "cleared_by_threshold": scan,
            "neighbourhood_truncated": truncated,
        }
        out[t] = rec
        print(f"\n{t}  candidates={len(candidates)}  "
              f"zero-change days {zero_days}/{known} "
              f"({rec['delta_zero_share']:.3f})  gcd={g}  "
              f"truncated neighbourhoods={truncated}")
        print("   cleared: " + "  ".join(
            f"τ>{k}:{v}" for k, v in scan.items()))
        w = in_window
        print(f"   in-window: gcd={w['gcd']} zero={w['zero']}/{w['known']} "
              f"({w['zero_share']}) multiples-of-50k "
              f"{w['multiples_of_50k']}/{w['nonzero_total']} "
              f"smallest={w['smallest_nonzero'][:3]}")
        for s in splits:
            print(f"   split {s['date']}: factor {s['factor']}, "
                  f"nav ratio {s['nav_ratio']}, assets ratio {s['tna_ratio']}, "
                  f"assets {'continuous' if s['assets_continuous'] else '**NOT continuous, not adjusted**'}")
        for o in w_odd:
            print(f"   **not a multiple of 50,000 after split adjustment**: "
                  f"{o['date']} Δ={o['delta']:,} shares, "
                  f"{o['shares_before']:,.0f} -> {o['shares_after']:,.0f}")
        if nz:
            print(f"   full-series smallest: {rec['smallest_nonzero_changes']}")

    if out:
        F1_OUT.parent.mkdir(parents=True, exist_ok=True)
        F1_OUT.write_text(json.dumps(
            {"stage": "B9-A F1", "diagnostic_only": True,
             "diagnostic_reason": "section 5's F1 traversability audit. the threshold is swept, not chosen"
                                  " (engineering rule 8). computes no ω and no λ.",
             "window_trading_days": F1_WINDOW,
             "thresholds": F1_THRESHOLDS, "funds": out},
            sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        print(f"\nwrote {F1_OUT.relative_to(ROOT)}")
        print("**No threshold is adopted here.** The curve is the object; "
              "which point on it defines the F1-cleared sample is a ruling, "
              "and §13.3 requires √N to be reported on whichever it is.")
    return out


# ---------------------------------------------------------------------------
# The three inputs B9-A-1 needs that are not `λ` itself. §13, §16.4, §6.
#
# Printed before any reading is computed, because each one hides a different
# unit trap and §15 already cost one of those.
# ---------------------------------------------------------------------------

PRODUCT_DATA = "constructed_product_data_all_funds"

#: §16.2, earned from the gcd of in-window share changes rather than read from
#: the SAI. §16.5 item 1 records that the two have not been cross-checked.
CREATION_UNIT_SHARES = 50_000

#: §16.4. Select Sector SPDR Trust only. The other four funds sit in different
#: trusts whose SAIs have not been opened, so they have no `f` and, per §14.6,
#: no `λ`. **`None` here means "not read", never "zero".**
FIXED_CREATION_FEE_USD = {
    "XLB": 500.0, "XLE": 500.0, "XLF": 500.0, "XLI": 500.0, "XLK": 500.0,
    "XLP": 500.0, "XLRE": 500.0, "XLU": 500.0, "XLV": 500.0, "XLY": 500.0,
    "XLC": 250.0,
    "SPDW": None, "SPEM": None, "SPAB": None, "JNK": None, "SPY": None,
}

STRESS_WINDOW = 60   # §6: the fund's own trailing 60-day realised NAV vol


def _clean_ticker(v) -> str:
    return "".join(ch for ch in str(v or "") if ch.isalnum()).upper()


def read_spread_table() -> dict:
    """`√N` per fund from the cross-fund workbook, located by header text.

    **The column is found by matching its name, never by index.** §13.3 makes
    this figure the denominator of every reading, and the header on that sheet
    is banked across three rows, so an index written today is an index that
    silently points elsewhere after the publisher adds a column."""
    from openpyxl import load_workbook
    for t in SAMPLE:
        p = CACHE / t / PRODUCT_DATA
        if p.is_file():
            break
    else:
        return {}
    wb = load_workbook(io.BytesIO(p.read_bytes()), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()

    tick_col = spread_col = asof_col = None
    for r in rows[:8]:
        for j, c in enumerate(r):
            s = str(c or "").strip().lower()
            if s == "ticker" and tick_col is None:
                tick_col = j
            if "median bid" in s and "spread" in s and spread_col is None:
                spread_col = j
            if s.startswith("as of") and asof_col is None:
                asof_col = j
    if tick_col is None or spread_col is None:
        return {"_error": f"header not found (ticker={tick_col}, spread={spread_col})"}

    out = {"_columns": {"ticker": tick_col, "spread": spread_col, "as_of": asof_col}}
    for r in rows:
        if len(r) <= max(tick_col, spread_col):
            continue
        tk = _clean_ticker(r[tick_col])
        if tk in SAMPLE:
            out[tk] = {"spread_raw": r[spread_col],
                       "spread_num": _num(r[spread_col]),
                       "as_of": _cell(r[asof_col]) if asof_col is not None else ""}
    return out


def realised_vol(navs: list, window: int = STRESS_WINDOW) -> list:
    """Trailing realised volatility of the NAV log return, §6's stress measure.
    Returns one value per input day, `None` until the window is full."""
    rets = [None]
    for i in range(1, len(navs)):
        a, b = navs[i - 1], navs[i]
        rets.append(math.log(b / a) if (a and b and a > 0 and b > 0) else None)
    out = []
    for i in range(len(navs)):
        lo = i - window + 1
        if lo < 1:
            out.append(None)
            continue
        w = [r for r in rets[lo:i + 1] if r is not None]
        if len(w) < window * 0.8:
            out.append(None)
            continue
        m = sum(w) / len(w)
        out.append(math.sqrt(sum((r - m) ** 2 for r in w) / (len(w) - 1)))
    return out


def inputs(tickers: list[str]) -> dict:
    spreads = read_spread_table()
    if "_error" in spreads:
        print(spreads["_error"])
        return {}
    print(f"spread column located at index {spreads['_columns']['spread']}, "
          f"ticker at {spreads['_columns']['ticker']}")

    out = {}
    for t in (tickers or SAMPLE):
        nav_rows, pd_rows = read_sheet(t, NAVHIST), read_sheet(t, PDHIST)
        if not nav_rows or not pd_rows:
            continue
        nav = sorted(split_rows(nav_rows, 3)["dated"], key=lambda r: r[0])
        pd_days = {d[0] for d in split_rows(pd_rows, 1)["dated"]}
        days = [r[0] for r in nav]
        navs = [r[1] for r in nav]
        vol = realised_vol(navs)
        in_win = [i for i, d in enumerate(days) if d in pd_days]
        wv = [vol[i] for i in in_win if vol[i] is not None]
        med = sorted(wv)[len(wv) // 2] if wv else None

        fee = FIXED_CREATION_FEE_USD.get(t)
        wnav = [navs[i] for i in in_win if navs[i]]
        f_max = None
        if fee is not None and wnav:
            fs = sorted(fee / (CREATION_UNIT_SHARES * n) for n in wnav)
            f_max = {"min": fs[0], "median": fs[len(fs) // 2], "max": fs[-1]}

        sp = spreads.get(t, {})
        out[t] = {
            "spread_raw": _cell(sp.get("spread_raw")), "spread_num": sp.get("spread_num"),
            "spread_as_of": sp.get("as_of"),
            "fixed_fee_usd": fee,
            "f_max_fraction": f_max,
            "vol_days": len(wv), "vol_median": med,
            "calm_days": sum(1 for v in wv if v <= med) if med else None,
            "stress_days": sum(1 for v in wv if v > med) if med else None,
        }
        fm = f"{f_max['median']*1e4:.3f}" if f_max else "  none"
        print(f"{t:6s} spread={str(sp.get('spread_raw')):>10s} "
              f"fee={str(fee):>6s} f_max(med)={fm:>7s} bp   "
              f"vol n={len(wv):>3d} median={med:.6f}   "
              f"calm/stress {out[t]['calm_days']}/{out[t]['stress_days']}"
              if med else f"{t:6s} insufficient vol window")

    print("\n**Units of the spread column are not decided here.** Read them off "
          "the raw values above, as §15.1 did for the premium column.")
    print("**f_max is the fee spread over ONE creation unit** (§16.4), that is "
          "the widest the interval can be. It is not an estimate of the fee "
          "actually paid.")
    return out


# ---------------------------------------------------------------------------
# B9-A-1. §6, evaluated on §17's rectangle.
#
# **The aggregation rule is fixed in this file before it is run**, which is the
# operative sense of pre-registration here: the archived copy of this file
# predates the record it writes.
#
#   sample     calm fund-days (§6: the fund's own trailing 60-day realised NAV
#              volatility at or below its own median) that clear F1 at
#              τ = 50,000 shares (§16.2)
#   statistic  the median of |λ| / √N over that sample, at each corner of §17's
#              rectangle, plus the share of days below 1
#   verdict    read at the worst corner, three-way per §16.4 and §17.3
#
# F3 is cleared by construction: the reading is a ratio to the floor and §4
# forbids reporting λ alone. **F2 is vacuous for this statistic and that is a
# claim, not an omission**: F2 guards against state coarseness driving π to
# zero, and |λ|/√N has no state grid. B9-A-4 is where F2 binds.
# ---------------------------------------------------------------------------

A1_OUT = ROOT / "results" / "b9_a1.json"
FM_OUT = ROOT / "results" / "b9_floor_measurement.json"

#: §24.4. The closing price is the NBBO midpoint on a half-cent grid (§24.1,
#: 2,020 of 2,020 reconstructed prices land on it). So the measurement quantum is
#: half a cent, exact and per fund-day, with **no publication rounding of its
#: own**. The cost floor is the published spread and answers a different
#: question.
MIDPOINT_GRID_USD = 0.005


def floor_measurement(nav: float, mode: str = "quantum") -> float:
    """`F_m`, §24.4. `quantum` is the grid step itself, the conservative
    reading; `sd` is the standard deviation of a uniform rounding error on that
    grid, which is the floor if the only error is the rounding."""
    q = MIDPOINT_GRID_USD / nav
    return q if mode == "quantum" else q / math.sqrt(12.0)


def midpoint_grid_check(tickers: list[str]) -> dict:
    """§24.1's test, kept in the file that depends on it. **Every reconstructed
    `P` must land on the half-cent grid.** A fund that fails it is a fund whose
    premium is not computed from these two published numbers, and its `λ` is not
    what this stage thinks it is."""
    spreads = read_spread_table()
    out = {}
    for t in (tickers or SAMPLE):
        nav_rows, pd_rows = read_sheet(t, NAVHIST), read_sheet(t, PDHIST)
        if not nav_rows or not pd_rows:
            continue
        nav = {d[0]: d[1] for d in split_rows(nav_rows, 3)["dated"]}
        pdh = {d[0]: d[1] for d in split_rows(pd_rows, 1)["dated"]}
        on_half = on_whole = off = 0
        navs, lams = [], []
        for d in sorted(set(nav) & set(pdh)):
            nv, pr = nav[d], pdh[d]
            if nv is None or pr is None or nv <= 0:
                continue
            p = nv * (1.0 + pr / 100.0)
            steps = p / MIDPOINT_GRID_USD
            if abs(steps - round(steps)) > 2e-3:
                off += 1
            elif round(steps) % 2 == 0:
                on_whole += 1
            else:
                on_half += 1
            navs.append(nv)
            lams.append(abs(math.log1p(pr / 100.0)))
        n = on_half + on_whole + off
        if not n:
            continue
        med = lambda xs: sorted(xs)[len(xs) // 2]
        nav_med = med(navs)
        tick_bp = 1e4 * 0.01 / nav_med
        lam_bp = 1e4 * med(lams)
        out[t] = {
            "n": n, "on_half_cent": on_half / n, "on_whole_cent": on_whole / n,
            "off_grid": off / n,
            "nav_median": nav_med, "tick_bp": tick_bp,
            "median_abs_lambda_bp": lam_bp,
            "lambda_over_half_tick": lam_bp / (tick_bp / 2),
            "published_spread": _cell(spreads.get(t, {}).get("spread_raw")),
        }
        flag = "" if off == 0 else "  **OFF GRID, λ is not what this stage thinks**"
        print(f"{t:6s} half={out[t]['on_half_cent']:.3f} "
              f"whole={out[t]['on_whole_cent']:.3f} off={out[t]['off_grid']:.3f}  "
              f"tick={tick_bp:6.3f}bp  |λ|={lam_bp:6.3f}bp  "
              f"ratio={out[t]['lambda_over_half_tick']:5.2f}{flag}")

    # **§24.3's comparison is within one underlying, and pooling breaks it.**
    # The first version spanned all sixteen and printed "tick 25.97x, |λ| 33.00x",
    # which reads as quantisation confirmed. It was the comparison arm: SPDW,
    # SPEM and JNK carry 11 to 39 basis points of stale-NAV premium (§6.1), a
    # confound registered long before this. The test needs funds whose λ is not
    # already known to be dominated by something else.
    US_EQUITY = set(MAIN_ARM) | {"SPY"}
    grp = {"US equity (main arm + SPY)": [t for t in out if t in US_EQUITY],
           "comparison arm, stale NAV (§6.1)": [t for t in out
                                                if t not in US_EQUITY]}
    print()
    for name, ts in grp.items():
        if len(ts) < 3:
            continue
        tk = [out[t]["tick_bp"] for t in ts]
        lm = [out[t]["median_abs_lambda_bp"] for t in ts]
        print(f"§24.3  {name}: tick spans {max(tk)/min(tk):.2f}x, "
              f"|λ| spans {max(lm)/min(lm):.2f}x   (n={len(ts)})")
    print("       **quantisation must scale with the quantum. On the US equity "
          "group the second number is far below the first, so it does not.**")
    print("       **The comparison arm is not a test of this**: its λ is "
          "dominated by stale NAV, which §6.1 registered as a confound.")
    # --- the record, and why it is not a diagnostic ------------------------
    # **Flipped 2026-08-18, the author ruled.** This record carried
    # `diagnostic_only: True` with the reason "reads no prediction", which is
    # true and is **not this field's test**. the project's engineering rule 8 fixes the field
    # as "the numbers in this record are not this stage's licensed readings",
    # and these two are: §24.1 is what identifies the disclosed price as the
    # closing NBBO midpoint, and §24.3 is what rejects the quantisation account.
    # **Scoring no registered prediction and not being a licensed reading are
    # different things**, and §24 is the case that separates them: it exists
    # because §4 was found to be wrong, so there was no prediction of it to
    # score, which the project's engineering rule 8 allows for exactly this shape.
    #
    # **The criteria below are written after the data and say so.** Both could
    # have failed and the second one does fail on the wrong sample, which is
    # what the third entry is for.
    us = [t for t in out if t in US_EQUITY]
    tk = [out[t]["tick_bp"] for t in us]
    lm = [out[t]["median_abs_lambda_bp"] for t in us]
    tick_span, lam_span = max(tk) / min(tk), max(lm) / min(lm)
    all_tk = [out[t]["tick_bp"] for t in out]
    all_lm = [out[t]["median_abs_lambda_bp"] for t in out]
    pooled_tick, pooled_lam = max(all_tk) / min(all_tk), max(all_lm) / min(all_lm)
    n_prices = sum(v["n"] for v in out.values())
    worst_off = max(v["off_grid"] for v in out.values())
    criteria = [
        {"name": "B9-24-1  the disclosed price is a half-cent midpoint",
         "passed": worst_off == 0.0,
         "detail": (f"off-grid share = {worst_off:.3f} across {len(out)} funds "
                    f"and {n_prices} reconstructed closes. Half a cent is not a "
                    f"price a trade can print at, so P is the closing NBBO "
                    f"midpoint and the reconstruction is exact rather than "
                    f"approximate. **Written after the data** (§24 exists "
                    f"because §4 was wrong); it could have failed, and a "
                    f"non-zero here would mean λ is not what this stage thinks")},
        {"name": "B9-24-2  |λ| does not scale with the tick, so it is not quantisation",
         "passed": lam_span < tick_span,
         "detail": (f"on the {len(us)} US-listed equity funds the tick spans "
                    f"{tick_span:.2f}x while |λ| spans {lam_span:.2f}x. "
                    f"Quantisation noise must scale with the quantum. SPY "
                    f"carries the point alone: its tick is the finest in the "
                    f"group and its |λ| is the same 1.7 bp as the rest")},
        {"name": "B9-24-3  the same comparison inverts when the stale-NAV arm is pooled in",
         "passed": pooled_lam > lam_span,
         "diagnostic": True,
         "detail": (f"all {len(out)} funds: tick spans {pooled_tick:.2f}x, "
                    f"|λ| spans {pooled_lam:.2f}x, which reads as quantisation "
                    f"confirmed. It is SPDW, SPEM and JNK carrying 11 to 39 bp "
                    f"of stale-NAV premium, a confound §6.1 registered before "
                    f"any of this ran. **This entry is here to show B9-24-2 "
                    f"could have failed**, and the first version of this run "
                    f"printed exactly this pooled line")},
    ]
    FM_OUT.parent.mkdir(parents=True, exist_ok=True)
    FM_OUT.write_text(json.dumps(
        {"stage": "B9-A §24", "diagnostic_only": False,
         "criteria": criteria,
         "post_hoc_note": ("§24 was written after §18 to §23, as the diagnosis "
                           "of a defect in §4, so neither criterion was "
                           "pre-registered. the project's engineering rule 8 covers this "
                           "shape: a pre-registration that could not have been "
                           "written without seeing the data is written after "
                           "it. Both are stated with the alternative they "
                           "could have returned."),
         "midpoint_grid_usd": MIDPOINT_GRID_USD, "funds": out},
        sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print(f"wrote {FM_OUT.relative_to(ROOT)}")
    print(f"§24     criteria: {sum(1 for c in criteria if not c.get('diagnostic') and c['passed'])}"
          f"/{sum(1 for c in criteria if not c.get('diagnostic'))} live passed, "
          f"1 diagnostic. **This record is no longer diagnostic_only** (the author "
          f"2026-08-18): it carries the grid identification and the "
          f"quantisation rejection, and both are licensed readings.")
    return out
SPREAD_QUANTUM = 0.0001        # one basis point, §17.1
F1_TAU = 50_000                # §16.2


def spread_to_fraction(raw) -> tuple:
    """Published spread as a fraction, with the unit taken from the string.

    Returns (fraction, note). **A cell without a percent sign is not silently
    assumed to be a percent**: it is returned with a note so the caller reports
    it rather than dividing by a hundred on faith."""
    if raw is None:
        return None, "missing"
    s = str(raw).strip()
    if s.endswith("%"):
        v = _num(s)
        return (None, "unparseable") if v is None else (v / 100.0, "percent_string")
    v = _num(s)
    if v is None:
        return None, "unparseable"
    return v, "bare_number_assumed_fraction"


def a1(tickers: list[str]) -> dict:
    if not GATE_OUT.is_file() or not json.loads(
            GATE_OUT.read_text(encoding="utf-8")).get("passed"):
        print("B9-0 has not passed. §8: nothing after it may be read.")
        return {}

    spreads = read_spread_table()
    if "_error" in spreads:
        print(spreads["_error"])
        return {}

    out = {}
    for t in (tickers or SAMPLE):
        nav_rows, pd_rows = read_sheet(t, NAVHIST), read_sheet(t, PDHIST)
        if not nav_rows or not pd_rows:
            continue
        nav = sorted(split_rows(nav_rows, 3)["dated"], key=lambda r: r[0])
        pdh = {d[0]: d[1] for d in split_rows(pd_rows, 1)["dated"]}
        days = [r[0] for r in nav]
        navs = [r[1] for r in nav]

        # split adjustment, §16.3, same rule as the F1 audit
        cum, factor = 1.0, [1.0] * len(days)
        for i in range(1, len(days)):
            a, b = navs[i - 1], navs[i]
            if a and b:
                r = a / b
                k = round(r)
                if k >= 2 and abs(r - k) / k < 0.02:
                    cum *= k
            factor[i] = cum
        shares = [None if r[2] is None else r[2] / factor[i]
                  for i, r in enumerate(nav)]
        delta = [None] + [None if (shares[i] is None or shares[i - 1] is None)
                          else abs(round(shares[i] - shares[i - 1]))
                          for i in range(1, len(shares))]

        vol = realised_vol(navs)
        idx = {d: i for i, d in enumerate(days)}
        cand = [d for d in days if d in pdh]
        wv = [vol[idx[d]] for d in cand if vol[idx[d]] is not None]
        vmed = sorted(wv)[len(wv) // 2] if wv else None

        sn, sn_note = spread_to_fraction(spreads.get(t, {}).get("spread_raw"))
        fee = FIXED_CREATION_FEE_USD.get(t)

        rows, dropped = [], {"no_premium": 0, "band": 0, "no_vol": 0, "no_f1": 0}
        for d in cand:
            i = idx[d]
            prem_pct, nv = pdh[d], navs[i]
            if prem_pct is None or nv is None or nv <= 0:
                dropped["no_premium"] += 1
                continue
            prem = prem_pct / 100.0
            if abs(prem) >= 0.25:                      # §15.1
                dropped["band"] += 1
                continue
            if vol[i] is None:
                dropped["no_vol"] += 1
                continue
            lo, hi = max(0, i - F1_WINDOW), min(len(days) - 1, i + F1_WINDOW)
            if not any(delta[k] is not None and delta[k] > F1_TAU
                       for k in range(lo, hi + 1)):
                dropped["no_f1"] += 1
                continue
            lam_hi = math.log1p(prem)
            f_max = None if fee is None else fee / (CREATION_UNIT_SHARES * nv)
            lam_lo = None if f_max is None else lam_hi - math.log1p(f_max)
            worst_num = abs(lam_hi) if lam_lo is None else max(abs(lam_hi), abs(lam_lo))
            best_num = abs(lam_hi) if lam_lo is None else min(abs(lam_hi), abs(lam_lo))
            rows.append({"date": d, "calm": vol[i] <= vmed,
                         "lam_hi": lam_hi, "lam_lo": lam_lo,
                         "worst": worst_num, "best": best_num})

        calm = [r for r in rows if r["calm"]]
        rec = {"spread_raw": _cell(spreads.get(t, {}).get("spread_raw")),
               "spread_fraction": sn, "spread_note": sn_note,
               "fixed_fee_usd": fee, "vol_median": vmed,
               "days_in_window": len(cand), "days_used": len(rows),
               "calm_days": len(calm), "dropped": dropped}

        if sn is None or sn <= 0.0 or not calm:
            rec["verdict"] = ("no_floor_published" if (sn is not None and sn <= 0)
                              else "insufficient")
            rec["note"] = ("§17.3: a fund published at 0.00% has a floor below "
                           "the reporting resolution and gets no ratio.")
            out[t] = rec
            print(f"{t:6s} {rec['verdict']}  (spread {rec['spread_raw']}, "
                  f"calm days {len(calm)})")
            continue

        sn_lo, sn_hi = max(sn - SPREAD_QUANTUM / 2, 0.0), sn + SPREAD_QUANTUM / 2
        worst = sorted(r["worst"] / sn_lo for r in calm) if sn_lo > 0 else []
        best = sorted(r["best"] / sn_hi for r in calm)
        med = lambda xs: xs[len(xs) // 2] if xs else None
        rec.update({
            "sqrtN_lo": sn_lo, "sqrtN_hi": sn_hi,
            "ratio_worst_median": med(worst), "ratio_best_median": med(best),
            "share_below_1_worst": (sum(1 for x in worst if x < 1) / len(worst)
                                    if worst else None),
            "share_below_1_best": sum(1 for x in best if x < 1) / len(best),
            "f_max_median_bp": (None if fee is None else
                                1e4 * med(sorted(fee / (CREATION_UNIT_SHARES * navs[idx[r['date']]])
                                                 for r in calm))),
        })
        wm, bm = rec["ratio_worst_median"], rec["ratio_best_median"]
        if wm is not None and wm < 1.0:
            v = "pass"
        elif bm >= 1.0:
            v = "fail"
        else:
            # which indeterminacy flips it: refloat the worst corner with the
            # fee removed, and separately with the quantum removed
            # **Attribution by necessity and sufficiency, not by order of
            # writing.** The first version tested `no_quantum < 1` second and
            # labelled XLB (`no_quantum = 1.0203`, so the fee alone already
            # carries it past 1 at the published floor) as fee-and-floor, while
            # XLU (both removals fix it, so neither alone suffices) came back as
            # fee. Five funds carried the wrong cause.
            #
            #   no_fee     < 1  ⇒ removing the fee fixes it     ⇒ fee necessary
            #   no_quantum < 1  ⇒ removing the widening fixes it ⇒ quantum necessary
            #
            # not necessary ⇒ the other one is sufficient on its own.
            no_fee = med(sorted(abs(r["lam_hi"]) / sn_lo for r in calm))
            no_quantum = med(sorted(r["worst"] / sn for r in calm))
            fee_needed = no_fee is not None and no_fee < 1.0
            quantum_needed = no_quantum is not None and no_quantum < 1.0
            if not fee_needed:
                v = "floor_indeterminate"          # survives removing the fee
            elif not quantum_needed:
                v = "fee_indeterminate"            # survives removing the widening
            else:
                v = "fee_and_floor_indeterminate"  # neither alone does it
            rec["flip_diagnostics"] = {
                "worst_without_fee": no_fee, "worst_without_quantum": no_quantum,
                "fee_necessary": fee_needed, "quantum_necessary": quantum_needed}
        rec["verdict"] = v
        out[t] = rec
        print(f"{t:6s} {v:26s} worst med={wm:.4f}  best med={bm:.4f}  "
              f"share<1 worst={rec['share_below_1_worst']:.3f} "
              f"best={rec['share_below_1_best']:.3f}  "
              f"calm={len(calm)}/{len(rows)}  √N={rec['spread_raw']}")

    if out:
        A1_OUT.parent.mkdir(parents=True, exist_ok=True)
        A1_OUT.write_text(json.dumps(
            {"stage": "B9-A-1", "diagnostic_only": False,
             "tau_shares": F1_TAU, "spread_quantum": SPREAD_QUANTUM,
             "creation_unit_shares": CREATION_UNIT_SHARES,
             "aggregation": "median of |λ|/√N over calm, F1-cleared fund-days; "
                            "verdict read at §17.3's worst corner",
             "funds": out}, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        print(f"\nwrote {A1_OUT.relative_to(ROOT)}")
    print("F3 cleared by construction (§4: the ratio, never λ alone). "
          "**F2 is vacuous for this statistic and that is a claim**: it guards "
          "state coarseness in π, and |λ|/√N has no state grid. B9-A-4 is where "
          "F2 binds.")
    return out


# ---------------------------------------------------------------------------
# B9-A-2. §6's gradient, §18.6's rules.
#
# **Immune to §17's quantum by construction.** `√N` is a per-fund constant
# (§13.2), so it sits on both sides of a within-fund calm-against-stress
# comparison and cancels from the direction entirely. §18.1's indeterminacy was
# a property of a level test and does not carry over.
#
# **The trap it is not immune to**: `f_max = fee / (unit × NAV)` and NAV falls in
# stress, so the fee end rises mechanically exactly when the prediction says the
# ratio should rise. §18.6 therefore counts the gradient only at the `f = 0` end
# and reports the fee end beside it, labelled.
# ---------------------------------------------------------------------------

A2_OUT = ROOT / "results" / "b9_a2.json"


def a2(tickers: list[str]) -> dict:
    if not GATE_OUT.is_file() or not json.loads(
            GATE_OUT.read_text(encoding="utf-8")).get("passed"):
        print("B9-0 has not passed. §8: nothing after it may be read.")
        return {}
    spreads = read_spread_table()
    if "_error" in spreads:
        print(spreads["_error"])
        return {}

    out = {}
    for t in (tickers or SAMPLE):
        nav_rows, pd_rows = read_sheet(t, NAVHIST), read_sheet(t, PDHIST)
        if not nav_rows or not pd_rows:
            continue
        nav = sorted(split_rows(nav_rows, 3)["dated"], key=lambda r: r[0])
        pdh = {d[0]: d[1] for d in split_rows(pd_rows, 1)["dated"]}
        days = [r[0] for r in nav]
        navs = [r[1] for r in nav]

        cum, factor = 1.0, [1.0] * len(days)
        for i in range(1, len(days)):
            a, b = navs[i - 1], navs[i]
            if a and b:
                r = a / b
                k = round(r)
                if k >= 2 and abs(r - k) / k < 0.02:
                    cum *= k
            factor[i] = cum
        shares = [None if r[2] is None else r[2] / factor[i]
                  for i, r in enumerate(nav)]
        delta = [None] + [None if (shares[i] is None or shares[i - 1] is None)
                          else abs(round(shares[i] - shares[i - 1]))
                          for i in range(1, len(shares))]

        vol = realised_vol(navs)
        idx = {d: i for i, d in enumerate(days)}
        cand = [d for d in days if d in pdh]
        wv = [vol[idx[d]] for d in cand if vol[idx[d]] is not None]
        vmed = sorted(wv)[len(wv) // 2] if wv else None
        sn, _ = spread_to_fraction(spreads.get(t, {}).get("spread_raw"))
        fee = FIXED_CREATION_FEE_USD.get(t)
        if not sn or sn <= 0 or vmed is None:
            print(f"{t:6s} no usable floor or volatility window")
            continue

        calm_hi, calm_lo, str_hi, str_lo, nav_calm, nav_str = [], [], [], [], [], []
        for d in cand:
            i = idx[d]
            prem_pct, nv = pdh[d], navs[i]
            if prem_pct is None or nv is None or nv <= 0 or vol[i] is None:
                continue
            prem = prem_pct / 100.0
            if abs(prem) >= 0.25:
                continue
            lo, hi = max(0, i - F1_WINDOW), min(len(days) - 1, i + F1_WINDOW)
            if not any(delta[k] is not None and delta[k] > F1_TAU
                       for k in range(lo, hi + 1)):
                continue
            lam_hi = math.log1p(prem)
            f_max = None if fee is None else fee / (CREATION_UNIT_SHARES * nv)
            lam_lo = None if f_max is None else lam_hi - math.log1p(f_max)
            r_hi = abs(lam_hi) / sn
            r_lo = None if lam_lo is None else abs(lam_lo) / sn
            if vol[i] <= vmed:
                calm_hi.append(r_hi); nav_calm.append(nv)
                if r_lo is not None:
                    calm_lo.append(r_lo)
            else:
                str_hi.append(r_hi); nav_str.append(nv)
                if r_lo is not None:
                    str_lo.append(r_lo)

        med = lambda xs: sorted(xs)[len(xs) // 2] if xs else None
        ch, sh = med(calm_hi), med(str_hi)
        cl, sl = med(calm_lo), med(str_lo)
        rec = {
            "spread_fraction": sn, "fixed_fee_usd": fee,
            "n_calm": len(calm_hi), "n_stress": len(str_hi),
            "median_calm_f0": ch, "median_stress_f0": sh,
            "ratio_f0": (sh / ch) if (ch and sh) else None,
            "rises_f0": (sh > ch) if (ch and sh) else None,
            "median_calm_fmax": cl, "median_stress_fmax": sl,
            "ratio_fmax": (sl / cl) if (cl and sl) else None,
            "rises_fmax": (sl > cl) if (cl and sl) else None,
            # the mechanical part of the fee end, measured rather than argued
            "median_nav_calm": med(nav_calm), "median_nav_stress": med(nav_str),
        }
        if rec["median_nav_calm"] and rec["median_nav_stress"]:
            rec["fmax_stress_over_calm"] = (
                rec["median_nav_calm"] / rec["median_nav_stress"])
        out[t] = rec
        print(f"{t:6s} f=0: calm {ch:.4f} stress {sh:.4f} "
              f"ratio {rec['ratio_f0']:.3f} {'UP' if rec['rises_f0'] else 'down':>4s}"
              + (f"   f=max: ratio {rec['ratio_fmax']:.3f} "
                 f"{'UP' if rec['rises_fmax'] else 'down':>4s}"
                 f"   fee mech {rec.get('fmax_stress_over_calm', float('nan')):.3f}"
                 if cl else "   (no fee read)"))

    main = [t for t in out if t in MAIN_ARM]
    up_f0 = [t for t in main if out[t]["rises_f0"]]
    up_fmax = [t for t in main if out[t].get("rises_fmax")]
    print(f"\nmain arm, f=0 end: **{len(up_f0)} of {len(main)} funds rise** "
          f"({', '.join(sorted(up_f0)) or 'none'})")
    print(f"main arm, f=max end: {len(up_fmax)} of {len(main)} rise "
          "(§18.6: this end carries the fee's mechanical response to a falling "
          "NAV and does not count on its own)")
    if out:
        A2_OUT.parent.mkdir(parents=True, exist_ok=True)
        A2_OUT.write_text(json.dumps(
            {"stage": "B9-A-2", "diagnostic_only": False,
             "rule": "§18.6: the gradient counts at the f=0 end only; the f=max "
                     "end is reported beside it and labelled mechanical.",
             "main_arm_rising_f0": sorted(up_f0),
             "main_arm_rising_fmax": sorted(up_fmax),
             "funds": out}, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        print(f"wrote {A2_OUT.relative_to(ROOT)}")
    return out


# ---------------------------------------------------------------------------
# §19.5's discriminators, and §19.4's independence check.
#
# §19.3 recorded that B9-A-2's rise is equally consistent with impaired
# arbitrage and with a mean-zero λ whose noise scale tracks volatility. D1 and
# D2 are registered because **each has a definite prediction under that noise
# null**, which the gradient does not.
#
# Nothing here computes a new λ: the series is the `f = 0` end, signed, on the
# same F1-cleared in-window fund-days A-2 used.
# ---------------------------------------------------------------------------

DISC_OUT = ROOT / "results" / "b9_disc.json"
PI_OUT = ROOT / "results" / "b9_pi_check.json"

#: median `f_max` per fund in the window, from `--inputs`. §16.4's interval uses
#: the per-day value; §22.4 only needs a representative one, and using the median
#: keeps this check from re-reading navhist for a number it does not test.
FEE_MED = {
    "XLB": 1.133e-4, "XLC": 0.435e-4, "XLE": 1.135e-4, "XLF": 1.873e-4,
    "XLI": 0.655e-4, "XLK": 0.389e-4, "XLP": 1.254e-4, "XLRE": 2.400e-4,
    "XLU": 1.214e-4, "XLV": 0.678e-4, "XLY": 0.445e-4,
}


def pearson(xs, ys) -> float | None:
    n = len(xs)
    if n < 3:
        return None
    mx, my = sum(xs) / n, sum(ys) / n
    sxy = sum((a - mx) * (b - my) for a, b in zip(xs, ys))
    sxx = sum((a - mx) ** 2 for a in xs)
    syy = sum((b - my) ** 2 for b in ys)
    if sxx <= 0 or syy <= 0:
        return None
    return sxy / math.sqrt(sxx * syy)


# ---------------------------------------------------------------------------
# §40.3 (B9-A-7): run the stage twice on the same fund-days, once on the
# disclosed premium and once on the premium reconstructed from a venue feed.
#
# **`control` exists because comparing a restricted run against an unrestricted
# one would confound a change of sample with a change of price.** Both modes use
# exactly the fund-days the reconstruction covers.
# ---------------------------------------------------------------------------

RECON_PATH = ROOT / "results" / "b9_recon_premium.json"
RECON_MODE = "off"          # off | recon | control
RECON_DATA: dict = {}


def load_recon(mode: str) -> None:
    global RECON_MODE, RECON_DATA
    RECON_MODE = mode
    if mode == "off":
        return
    if not RECON_PATH.is_file():
        raise SystemExit(
            f"{RECON_PATH.relative_to(ROOT)} is absent. Run "
            f"`b9a_availability.py --nbbo-export --dataset ARCX.PILLAR` first.")
    payload = json.loads(RECON_PATH.read_text(encoding="utf-8"))
    RECON_DATA = payload.get("premium", {})
    n = sum(len(v) for v in RECON_DATA.values())
    print(f"§40.3  mode={mode}  reconstruction from "
          f"{payload.get('dataset')} covering {n} fund-days"
          + ("  (prices reconstructed)" if mode == "recon"
             else "  (**disclosed prices, reconstruction's day set**)"))


def out_path(p: Path) -> Path:
    """Results from a §40.3 run never overwrite the stage's own."""
    if RECON_MODE == "off":
        return p
    return p.with_name(f"{p.stem}_{RECON_MODE}{p.suffix}")


#: §49.5(4). `lambda_series` used four bare `continue`s and counted none of
#: them, while `--a1` counts its own drops in four named buckets. Every stage
#: that consumes the series inherits its sample, so the sample needs a
#: denominator. Filled per ticker on each call.
LAMBDA_DROPS: dict = {}


def lambda_series(t: str, spreads: dict) -> dict:
    """The signed `λ` at the `f = 0` end on A-2's sample, with each day's calm
    flag. One place, so D1, D2 and the demeaning cannot disagree about which
    days they are talking about. **Drops are counted into `LAMBDA_DROPS[t]`.**"""
    drops = {"candidates": 0, "no_premium_or_nav": 0, "band": 0, "no_vol": 0,
             "no_f1": 0, "recon_absent": 0, "kept": 0}
    LAMBDA_DROPS[t] = drops
    nav_rows, pd_rows = read_sheet(t, NAVHIST), read_sheet(t, PDHIST)
    if not nav_rows or not pd_rows:
        return {}
    nav = sorted(split_rows(nav_rows, 3)["dated"], key=lambda r: r[0])
    pdh = {d[0]: d[1] for d in split_rows(pd_rows, 1)["dated"]}
    days = [r[0] for r in nav]
    navs = [r[1] for r in nav]

    cum, factor = 1.0, [1.0] * len(days)
    for i in range(1, len(days)):
        a, b = navs[i - 1], navs[i]
        if a and b:
            r = a / b
            k = round(r)
            if k >= 2 and abs(r - k) / k < 0.02:
                cum *= k
        factor[i] = cum
    shares = [None if r[2] is None else r[2] / factor[i] for i, r in enumerate(nav)]
    delta = [None] + [None if (shares[i] is None or shares[i - 1] is None)
                      else abs(round(shares[i] - shares[i - 1]))
                      for i in range(1, len(shares))]
    vol = realised_vol(navs)
    idx = {d: i for i, d in enumerate(days)}
    cand = [d for d in days if d in pdh]
    wv = [vol[idx[d]] for d in cand if vol[idx[d]] is not None]
    if not wv:
        return {}
    vmed = sorted(wv)[len(wv) // 2]

    out = {}
    drops["candidates"] = len(cand)
    for d in cand:
        i = idx[d]
        prem_pct, nv = pdh[d], navs[i]
        if prem_pct is None or nv is None or nv <= 0:
            drops["no_premium_or_nav"] += 1
            continue
        if vol[i] is None:
            drops["no_vol"] += 1
            continue
        prem = prem_pct / 100.0
        if RECON_MODE != "off":
            # Both modes are restricted to the reconstruction's coverage, so
            # `recon` and `control` differ in the price and in nothing else.
            r = RECON_DATA.get(t, {})
            if d not in r:
                drops["recon_absent"] += 1
                continue
            if RECON_MODE == "recon":
                prem = r[d]
        if abs(prem) >= 0.25:
            drops["band"] += 1
            continue
        lo, hi = max(0, i - F1_WINDOW), min(len(days) - 1, i + F1_WINDOW)
        if not any(delta[k] is not None and delta[k] > F1_TAU
                   for k in range(lo, hi + 1)):
            drops["no_f1"] += 1
            continue
        drops["kept"] += 1
        # `nav` is carried so that §36's comparison can rebuild the disclosed
        # price without parsing the workbooks a second time. **One data model,
        # one filter**: whatever this function excludes is excluded there too.
        out[d] = {"lam": math.log1p(prem), "prem": prem, "nav": nv,
                  "calm": vol[i] <= vmed, "order": i}
    return out


def discriminators(tickers: list[str]) -> dict:
    spreads = read_spread_table()
    series = {t: lambda_series(t, spreads) for t in (tickers or SAMPLE)}
    series = {t: s for t, s in series.items() if s}
    out = {}

    for t, s in series.items():
        days = sorted(s)
        # D1: lag-1 pairs on consecutive rows of the underlying series, classed
        # by the state of the later day. **Non-consecutive pairs are excluded
        # and counted**, since a gap is not a lag of one.
        pairs = {"calm": ([], []), "stress": ([], [])}
        skipped = 0
        for a, b in zip(days, days[1:]):
            if s[b]["order"] - s[a]["order"] != 1:
                skipped += 1
                continue
            k = "calm" if s[b]["calm"] else "stress"
            pairs[k][0].append(s[a]["lam"])
            pairs[k][1].append(s[b]["lam"])
        ac_calm = pearson(*pairs["calm"])
        ac_stress = pearson(*pairs["stress"])

        # D2: share of days at a discount
        calm_d = [d for d in days if s[d]["calm"]]
        str_d = [d for d in days if not s[d]["calm"]]
        disc_calm = (sum(1 for d in calm_d if s[d]["prem"] < 0) / len(calm_d)
                     if calm_d else None)
        disc_str = (sum(1 for d in str_d if s[d]["prem"] < 0) / len(str_d)
                    if str_d else None)

        out[t] = {
            "d1_autocorr_calm": ac_calm, "d1_autocorr_stress": ac_stress,
            "d1_diff": (None if (ac_calm is None or ac_stress is None)
                        else ac_stress - ac_calm),
            "d1_pairs_calm": len(pairs["calm"][0]),
            "d1_pairs_stress": len(pairs["stress"][0]),
            "d1_pairs_skipped_gap": skipped,
            "d2_discount_share_calm": disc_calm,
            "d2_discount_share_stress": disc_str,
            "d2_shift": (None if (disc_calm is None or disc_str is None)
                         else disc_str - disc_calm),
            "n_calm": len(calm_d), "n_stress": len(str_d),
        }
        print(f"{t:6s} D1 calm={_f(ac_calm)} stress={_f(ac_stress)} "
              f"diff={_f(out[t]['d1_diff'])}   "
              f"D2 calm={_f(disc_calm)} stress={_f(disc_str)} "
              f"shift={_f(out[t]['d2_shift'])}")

    # --- §19.4: are the eleven independent -------------------------------
    main = [t for t in series if t in MAIN_ARM]
    stress_sets = {t: {d for d, v in series[t].items() if not v["calm"]}
                   for t in main}
    jac = []
    for i, a in enumerate(main):
        for b in main[i + 1:]:
            u = stress_sets[a] | stress_sets[b]
            jac.append(len(stress_sets[a] & stress_sets[b]) / len(u) if u else 0.0)
    all_days = sorted(set().union(*(set(series[t]) for t in main))) if main else []
    per_day = {d: [series[t][d]["lam"] for t in main if d in series[t]]
               for d in all_days}
    common = {d: sum(v) / len(v) for d, v in per_day.items() if v}

    # variance share taken by the day effect, pooled over the main arm
    pooled = [(t, d, series[t][d]["lam"]) for t in main for d in series[t]]
    if pooled:
        mu = sum(x for _, _, x in pooled) / len(pooled)
        tot = sum((x - mu) ** 2 for _, _, x in pooled)
        res = sum((x - common[d]) ** 2 for _, d, x in pooled if d in common)
        day_share = 1 - res / tot if tot > 0 else None
    else:
        day_share = None

    # rerun A-2's direction on the within-day residual
    demeaned_up = []
    for t in main:
        c = [abs(series[t][d]["lam"] - common[d]) for d in series[t]
             if series[t][d]["calm"] and d in common]
        s_ = [abs(series[t][d]["lam"] - common[d]) for d in series[t]
              if not series[t][d]["calm"] and d in common]
        mc = sorted(c)[len(c) // 2] if c else None
        ms = sorted(s_)[len(s_) // 2] if s_ else None
        out[t]["demeaned_calm"], out[t]["demeaned_stress"] = mc, ms
        out[t]["demeaned_ratio"] = (ms / mc) if (mc and ms) else None
        if mc and ms and ms > mc:
            demeaned_up.append(t)

    # D1b, §20.5: the same question on the within-day residual, which asks
    # whether a fund carries persistence of its own beyond the market-wide one.
    # **D2 is deliberately not demeaned** (§20.5): if every fund moves to a
    # discount together that is the market-wide capacity constraint, and
    # subtracting the daily mean would remove the object being measured.
    for t in main:
        s = series[t]
        pairs_b = {"calm": ([], []), "stress": ([], [])}
        days_t = sorted(s)
        for a, b in zip(days_t, days_t[1:]):
            if s[b]["order"] - s[a]["order"] != 1:
                continue
            if a not in common or b not in common:
                continue
            k = "calm" if s[b]["calm"] else "stress"
            pairs_b[k][0].append(s[a]["lam"] - common[a])
            pairs_b[k][1].append(s[b]["lam"] - common[b])
        bc, bs = pearson(*pairs_b["calm"]), pearson(*pairs_b["stress"])
        out[t]["d1b_autocorr_calm"] = bc
        out[t]["d1b_autocorr_stress"] = bs
        out[t]["d1b_diff"] = None if (bc is None or bs is None) else bs - bc

    d1_up = [t for t in main if (out[t]["d1_diff"] or 0) > 0]
    d1b_up = [t for t in main if (out[t]["d1b_diff"] or 0) > 0]
    d1b_pos_calm = [t for t in main if (out[t]["d1b_autocorr_calm"] or 0) > 0]
    d2_up = [t for t in main if (out[t]["d2_shift"] or 0) > 0]
    print(f"\n§19.4  mean pairwise Jaccard of stress-day sets = "
          f"{sum(jac)/len(jac):.3f}" if jac else "")
    print(f"§19.4  variance share taken by the day effect = {_f(day_share)}")
    print(f"§19.4  funds still rising after within-day demeaning: "
          f"{len(demeaned_up)}/{len(main)}  ({', '.join(sorted(demeaned_up)) or 'none'})")
    print(f"D1     funds with stress autocorr above calm: {len(d1_up)}/{len(main)}")
    print(f"D1b    idiosyncratic (within-day residual): "
          f"{len(d1b_up)}/{len(main)} rise, "
          f"{len(d1b_pos_calm)}/{len(main)} positive already on calm days")
    print(f"D2     funds shifting toward discount in stress: {len(d2_up)}/{len(main)}")
    print("       (§20.5: D2 is not demeaned on purpose. A common move to a "
          "discount is the market-wide constraint, not a nuisance.)")

    rec = {"stage": "B9-A-2 discriminators", "diagnostic_only": False,
           "rule": "§19.5. D1 and D2 each have a definite prediction under the "
                   "noise null; the gradient does not.",
           "mean_pairwise_jaccard_stress_days": (sum(jac) / len(jac)) if jac else None,
           "day_effect_variance_share": day_share,
           "main_arm_demeaned_rising": sorted(demeaned_up),
           "main_arm_d1_positive": sorted(d1_up),
           "main_arm_d1b_positive": sorted(d1b_up),
           "main_arm_d1b_calm_positive": sorted(d1b_pos_calm),
           "main_arm_d2_positive": sorted(d2_up),
           "funds": out}
    DISC_OUT_P = out_path(DISC_OUT)
    DISC_OUT_P.parent.mkdir(parents=True, exist_ok=True)
    DISC_OUT_P.write_text(json.dumps(rec, sort_keys=True, ensure_ascii=False,
                                   indent=2) + "\n", encoding="utf-8", newline="\n")
    print(f"wrote {DISC_OUT_P.relative_to(ROOT)}")
    print("**§19.5's falsification stands**: D1 and D2 both null with the rise "
          "intact files the rise as dispersion scaling, and B9-A-2 is reported "
          "as not discriminating on this carrier.")
    return rec


def pi_check(tickers: list[str]) -> dict:
    """§22.4. **This computes `π` to check §22.2's arithmetic, not to test
    holonomy.** A reading near the closed form confirms the derivation and says
    nothing about the world; a reading far from it means the derivation is wrong
    and §22.3's disposition is withdrawn."""
    spreads = read_spread_table()
    out = {}
    for t in (tickers or SAMPLE):
        s = lambda_series(t, spreads)
        if not s:
            continue
        fee = FIXED_CREATION_FEE_USD.get(t)
        for end in ("f0", "fmax"):
            if end == "fmax" and fee is None:
                continue
            cells = []
            for d, v in s.items():
                a1 = -math.log1p(v["prem"])
                if end == "f0":
                    a2 = 0.0
                else:
                    # NAV is not carried in the series; recover f from λ's own
                    # scale is not possible, so the fee end uses the fund's
                    # median f_max, which is what §16.4 bounds the interval by.
                    a2 = -math.log1p(FEE_MED.get(t, 0.0))
                cells.append((a1, a2))
            if not cells:
                continue
            allv = [x for c in cells for x in c]
            n = len(allv)
            mu = sum(allv) / n
            tot = sum((x - mu) ** 2 for x in allv) / n
            within = sum(((c[0] - c[1]) / 2) ** 2 for c in cells) * 2 / n
            pi = within / tot if tot > 0 else None
            # closed form of §22.2
            prem = [v["prem"] for v in s.values()]
            V = (sum((p - sum(prem) / len(prem)) ** 2 for p in prem) / len(prem))
            f_used = 0.0 if end == "f0" else FEE_MED.get(t, 0.0)
            dd = (sum(prem) / len(prem) - f_used) ** 2
            closed = (V + dd) / (2 * V + dd) if (2 * V + dd) > 0 else None
            out.setdefault(t, {})[end] = {
                "pi": pi, "closed_form": closed,
                "abs_gap": None if (pi is None or closed is None) else abs(pi - closed),
                "n_cells": len(cells)}
        if t in out:
            r = out[t]
            line = f"{t:6s}"
            for end in ("f0", "fmax"):
                if end in r:
                    line += (f"  {end}: pi={r[end]['pi']:.4f} "
                             f"closed={r[end]['closed_form']:.4f} "
                             f"gap={r[end]['abs_gap']:.4f}")
            print(line)
    gaps = [v[e]["abs_gap"] for v in out.values() for e in v
            if v[e]["abs_gap"] is not None]
    if gaps:
        print(f"\nlargest gap between π and §22.2's closed form: {max(gaps):.4f}")
        print("**§22.4: a small gap confirms the arithmetic and says nothing "
              "about holonomy. A large gap means the derivation is wrong and "
              "§22.3's disposition is withdrawn.**")
    PI_OUT_P = out_path(PI_OUT)
    PI_OUT_P.parent.mkdir(parents=True, exist_ok=True)
    PI_OUT_P.write_text(json.dumps(
        {"stage": "B9-A-4 derivation check", "diagnostic_only": True,
         "diagnostic_reason": "section 22.4: this run tests the derivation in section 22.2, not the world.",
         "funds": out}, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print(f"wrote {PI_OUT_P.relative_to(ROOT)}")
    return out


MEAS_OUT = ROOT / "results" / "b9_measured.json"
GATESPEED_OUT = ROOT / "results" / "b9_gate_speed.json"


def gate_speed(tickers: list[str]) -> dict:
    """§26. A3 §6.4d's wall, mapped to the primary market.

    **Two margins, because A3's claim is about their contrast**: a wall moves the
    extensive margin (does the primary market move at all) and leaves the
    intensive one (how big when it does); an auction does the reverse.

    **This is about `H⁰`. It may not be added to B9-A-2's reading** (§26.1).
    """
    spreads = read_spread_table()
    out = {}
    for t in (tickers or SAMPLE):
        s = lambda_series(t, spreads)
        nav_rows = read_sheet(t, NAVHIST)
        if not s or not nav_rows:
            continue
        nav = sorted(split_rows(nav_rows, 3)["dated"], key=lambda r: r[0])
        days = [r[0] for r in nav]
        navs = [r[1] for r in nav]
        cum, factor = 1.0, [1.0] * len(days)
        for i in range(1, len(days)):
            a, b = navs[i - 1], navs[i]
            if a and b:
                r = a / b
                k = round(r)
                if k >= 2 and abs(r - k) / k < 0.02:
                    cum *= k
            factor[i] = cum
        shares = [None if r[2] is None else r[2] / factor[i]
                  for i, r in enumerate(nav)]
        idx = {d: i for i, d in enumerate(days)}

        ext = {"calm": [0, 0], "stress": [0, 0]}      # [zero days, total]
        inten = {"calm": [], "stress": []}
        for d, v in s.items():
            i = idx.get(d)
            if i is None or i == 0 or shares[i] is None or shares[i - 1] is None:
                continue
            delta = abs(shares[i] - shares[i - 1])
            k = "calm" if v["calm"] else "stress"
            ext[k][1] += 1
            if delta < 0.5:
                ext[k][0] += 1
            elif shares[i]:
                inten[k].append(delta / shares[i])
        if not ext["calm"][1] or not ext["stress"][1]:
            continue
        med = lambda xs: sorted(xs)[len(xs) // 2] if xs else None
        e_c = ext["calm"][0] / ext["calm"][1]
        e_s = ext["stress"][0] / ext["stress"][1]
        i_c, i_s = med(inten["calm"]), med(inten["stress"])
        rec = {
            "extensive_zero_share_calm": e_c, "extensive_zero_share_stress": e_s,
            "extensive_shift": e_s - e_c,
            "intensive_median_calm": i_c, "intensive_median_stress": i_s,
            "intensive_ratio": (i_s / i_c) if (i_c and i_s) else None,
            "n_calm": ext["calm"][1], "n_stress": ext["stress"][1],
            "moving_days_calm": len(inten["calm"]),
            "moving_days_stress": len(inten["stress"]),
        }
        out[t] = rec
        print(f"{t:6s} extensive zero-share {e_c:.3f} -> {e_s:.3f} "
              f"({rec['extensive_shift']:+.3f})   "
              f"intensive {i_c:.5f} -> {i_s:.5f} "
              f"(x{rec['intensive_ratio']:.3f})" if i_c and i_s else
              f"{t:6s} extensive {e_c:.3f} -> {e_s:.3f}   intensive n/a")

    main = [t for t in out if t in MAIN_ARM]
    ext_up = [t for t in main if out[t]["extensive_shift"] > 0]
    int_dn = [t for t in main if (out[t]["intensive_ratio"] or 1) < 1]
    print(f"\n§26.3  extensive margin RISES in {len(ext_up)}/{len(main)} "
          f"(A3's gate closing) — falls in {len(main)-len(ext_up)} "
          f"(the flow account)")
    print(f"§26.4  intensive margin falls in {len(int_dn)}/{len(main)}")
    print("§26.5  **one market-wide pattern seen across eleven sectors, never "
          "eleven independent draws.** SPDW and SPEM sit near 0.9 already and "
          "have little room to rise.")
    print("§26.1  **this is H⁰. It may not be added to B9-A-2's reading.**")
    GATESPEED_OUT_P = out_path(GATESPEED_OUT)
    GATESPEED_OUT_P.parent.mkdir(parents=True, exist_ok=True)
    GATESPEED_OUT_P.write_text(json.dumps(
        {"stage": "B9 §26 gate speed", "diagnostic_only": False,
         "maps": "a3_asset_channel.md §6.4d, the wall rather than the auction",
         "substitution_declared": "volatility proxies A-track pressure (§26.2)",
         "main_arm_extensive_rising": sorted(ext_up),
         "main_arm_intensive_falling": sorted(int_dn),
         "funds": out}, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print(f"wrote {GATESPEED_OUT_P.relative_to(ROOT)}")
    return out


def measured(tickers: list[str]) -> dict:
    """B9-A-1 and B9-A-2 against §24.4's measurement floor `F_m`.

    **The conservative corner flips with the claim.** §17 asked "is it zero" and
    took the largest `|λ|` over the smallest floor. This asks "is it resolvable",
    so the conservative corner is the **smallest** `|λ|` over the **largest**
    floor, and a claim of resolvability must survive there.

    **The mechanical direction also flips.** `F_m = 0.005 / NAV` rises when NAV
    falls, so a stress-period drawdown **lowers** the ratio. §18.6's fee trap
    pushed the other way. A rise that survives here survives against the
    mechanism rather than because of it, and the version with `F_m` frozen at its
    calm median is reported beside it so the two can be told apart.
    """
    spreads = read_spread_table()
    out = {}
    for t in (tickers or SAMPLE):
        s = lambda_series(t, spreads)
        if not s:
            continue
        nav_rows = read_sheet(t, NAVHIST)
        navs = {d[0]: d[1] for d in split_rows(nav_rows, 3)["dated"]}
        fee = FIXED_CREATION_FEE_USD.get(t)

        rows = []
        for d, v in s.items():
            nv = navs.get(d)
            if not nv or nv <= 0:
                continue
            lam_hi = v["lam"]
            f_max = None if fee is None else fee / (CREATION_UNIT_SHARES * nv)
            lam_lo = None if f_max is None else lam_hi - math.log1p(f_max)
            best = abs(lam_hi) if lam_lo is None else min(abs(lam_hi), abs(lam_lo))
            worst = abs(lam_hi) if lam_lo is None else max(abs(lam_hi), abs(lam_lo))
            rows.append({"d": d, "calm": v["calm"], "nav": nv,
                         "hi": abs(lam_hi), "best": best, "worst": worst,
                         "fq": floor_measurement(nv),
                         "fsd": floor_measurement(nv, "sd")})
        if not rows:
            continue
        med = lambda xs: sorted(xs)[len(xs) // 2] if xs else None
        calm = [r for r in rows if r["calm"]]
        strs = [r for r in rows if not r["calm"]]
        if not calm or not strs:
            continue

        cons = med([r["best"] / r["fq"] for r in calm])       # conservative
        gene = med([r["worst"] / r["fsd"] for r in calm])      # generous
        plain = med([r["hi"] / r["fq"] for r in calm])         # f = 0, quantum

        fq_calm_med = med([r["fq"] for r in calm])
        a2_var_c = med([r["hi"] / r["fq"] for r in calm])
        a2_var_s = med([r["hi"] / r["fq"] for r in strs])
        a2_fix_c = med([r["hi"] / fq_calm_med for r in calm])
        a2_fix_s = med([r["hi"] / fq_calm_med for r in strs])
        nav_mech = med([r["nav"] for r in calm]) / med([r["nav"] for r in strs])

        rec = {
            "n_calm": len(calm), "n_stress": len(strs),
            "a1_conservative": cons, "a1_generous": gene, "a1_f0_quantum": plain,
            "a1_verdict": ("resolvable" if cons and cons > 1.0
                           else "not_resolvable_at_the_conservative_corner"),
            "a2_ratio_varying_floor": (a2_var_s / a2_var_c) if a2_var_c else None,
            "a2_ratio_frozen_floor": (a2_fix_s / a2_fix_c) if a2_fix_c else None,
            "a2_rises_varying": (a2_var_s > a2_var_c) if a2_var_c else None,
            "a2_rises_frozen": (a2_fix_s > a2_fix_c) if a2_fix_c else None,
            "nav_calm_over_stress": nav_mech,
            "floor_quantum_bp_median": 1e4 * fq_calm_med,
        }
        out[t] = rec
        print(f"{t:6s} A1 cons={cons:6.2f} f0={plain:6.2f} gen={gene:7.2f}  "
              f"{rec['a1_verdict'][:11]:11s} | A2 var={rec['a2_ratio_varying_floor']:.3f}"
              f"{'UP' if rec['a2_rises_varying'] else 'dn':>3s} "
              f"frozen={rec['a2_ratio_frozen_floor']:.3f}"
              f"{'UP' if rec['a2_rises_frozen'] else 'dn':>3s}  "
              f"navmech={nav_mech:.3f}")

    main = [t for t in out if t in MAIN_ARM]
    res = [t for t in main if out[t]["a1_verdict"] == "resolvable"]
    up_v = [t for t in main if out[t]["a2_rises_varying"]]
    up_f = [t for t in main if out[t]["a2_rises_frozen"]]
    print(f"\nA1  main arm resolvable at the conservative corner: "
          f"{len(res)}/{len(main)}")
    print(f"A2  rising with the floor varying: {len(up_v)}/{len(main)}   "
          f"with the floor frozen at its calm median: {len(up_f)}/{len(main)}")
    print("    (§24.7: a falling NAV raises F_m and **lowers** the ratio, so the "
          "varying column understates a real rise rather than manufacturing one.)")
    MEAS_OUT_P = out_path(MEAS_OUT)
    MEAS_OUT_P.parent.mkdir(parents=True, exist_ok=True)
    MEAS_OUT_P.write_text(json.dumps(
        {"stage": "B9-A-1 and A-2 against the measurement floor (§24)",
         "diagnostic_only": False,
         "floor": "F_m = 0.005 / NAV, the half-cent midpoint grid (§24.4)",
         "conservative_corner": "min |λ| over the quantum floor; the corner a "
                                "resolvability claim must survive",
         "main_arm_resolvable": sorted(res),
         "main_arm_a2_rising_varying": sorted(up_v),
         "main_arm_a2_rising_frozen": sorted(up_f),
         "funds": out}, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print(f"wrote {MEAS_OUT_P.relative_to(ROOT)}")
    return out


def _f(x) -> str:
    return "   None" if x is None else f"{x:7.4f}"


# ---------------------------------------------------------------------------
# selftest
# ---------------------------------------------------------------------------


def selftest() -> int:
    fails: list[str] = []

    def check(name: str, cond: bool, detail: str = "") -> None:
        print(f"  {'ok  ' if cond else 'FAIL'} {name}{(' :: ' + detail) if detail else ''}")
        if not cond:
            fails.append(name)

    st = {"nav": 100.0, "price": 100.05, "fee": 0.0003}

    print("the cochain, §14.1")
    check("cash->basket is zero", omega("cash", "basket", st) == 0.0)
    check("basket->etf is -log(1+f)",
          abs(omega("basket", "etf", st) + math.log(1 + st["fee"])) < 1e-15)
    check("etf->cash is log(P/NAV)",
          abs(omega("etf", "cash", st) - math.log(st["price"] / st["nav"])) < 1e-15)
    check("a reversed edge is the exact negation",
          omega("cash", "etf", st) == -omega("etf", "cash", st))
    # All six ordered pairs among three positions are legal: three canonical
    # and three reversals. So there is no "absent edge" to test, and a check
    # written for one would be an inert cell that always passes.
    check("basket->cash is the legal reversal of a zero edge, and is zero",
          omega("basket", "cash", st) == 0.0)
    check("all six ordered pairs resolve",
          len([1 for u in POSITIONS for v in POSITIONS
               if u != v and isinstance(omega(u, v, st), float)]) == 6)
    check("a non-position raises", _raises(lambda: omega("cash", "gold", st)))
    check("a self-edge raises", _raises(lambda: omega("cash", "cash", st)))

    print("B9-0, §6, through the same summation")
    z = path_sum(DEGENERATE, st)
    check("the degenerate loop is exactly zero", z == 0.0, repr(z))
    check("it is zero for a discount too",
          path_sum(DEGENERATE, {"nav": 100.0, "price": 99.4, "fee": 0.01}) == 0.0)
    check("it is zero when the fee is zero",
          path_sum(DEGENERATE, {"nav": 50.0, "price": 51.0, "fee": 0.0}) == 0.0)

    print("the cycle, §14.2")
    premium = st["price"] / st["nav"] - 1.0
    check("path_sum equals the closed form",
          abs(path_sum(LOOP, st) - lam(premium, st["fee"])) < 1e-15,
          f"{path_sum(LOOP, st):.12g} vs {lam(premium, st['fee']):.12g}")
    check("a zero premium and a zero fee give exactly zero",
          path_sum(LOOP, {"nav": 10.0, "price": 10.0, "fee": 0.0}) == 0.0)

    print("§14.6, the sign of the fee, fixed before any reading")
    base = path_sum(LOOP, {"nav": 100.0, "price": 100.0, "fee": 0.0})
    fee_up = path_sum(LOOP, {"nav": 100.0, "price": 100.0, "fee": 0.002})
    check("a larger fee pushes lambda more negative", fee_up < base,
          f"{fee_up:.6g} < {base:.6g}")
    check("a fee cannot manufacture a zero from a premium",
          path_sum(LOOP, {"nav": 100.0, "price": 100.5, "fee": 0.002}) != 0.0)

    print("§14.5, the join and its reconciliation")
    rows = [
        ("Fund Name:", "State Street XLF"),
        ("Ticker Symbol:", "XLF"),
        ("Date", "Premium/Discount"),
        ("14-Aug-2026", -0.007977),
        ("13-Aug-2026", 0.010658),
        ("Q2 2026", None),
        (None, None),
    ]
    sp = split_rows(rows, value_cols=1)
    check("two dated rows found", len(sp["dated"]) == 2, str(len(sp["dated"])))
    check("section furniture lands in undated", len(sp["undated"]) == 4,
          str(sp["undated"]))
    check("a fully blank row is counted as blank, not as furniture",
          sp["blank"] == 1 and all("None" not in u for u in sp["undated"]))
    check("every row lands in exactly one bucket",
          len(sp["dated"]) + len(sp["undated"]) + sp["blank"] == len(rows))
    check("dates are parsed to iso", sp["dated"][0][0] == "2026-08-14")
    check("the value comes through as a number", sp["dated"][0][1] == -0.007977)

    print("numbers")
    check("a percent string parses to its numeral", _num("0.02%") == 0.02)
    check("a thousands separator parses", _num("1,160,850,000.00") == 1160850000.0)
    check("a dash is not a number", _num("-") is None)
    check("None stays None", _num(None) is None)

    print("quantiles")
    q = quantiles([0.0, 1.0, 2.0, 3.0, 4.0])
    check("median of five", q["p50"] == 2.0, str(q["p50"]))
    check("empty gives empty", quantiles([]) == {})

    print("B9-0's NaN guard, §10 step 3")
    nanst = {"nav": 100.0, "price": 100.2, "fee": float("nan")}
    check("the degenerate loop is zero even with an unusable fee",
          path_sum(DEGENERATE, nanst) == 0.0)
    check("a path that does touch the fee edge goes NaN, so a wrong route "
          "cannot pass as a zero",
          math.isnan(path_sum(LOOP, nanst)))

    print("§15.1's unit conversion and band")
    check("the percent column converts once", (-0.007977) / 100.0 == -7.977e-05)
    check("SPDW's largest raw value is inside the band after conversion",
          abs(6.24634 / 100.0) < 0.25)
    check("the band would catch a column that arrived as a fraction",
          abs(6.24634) >= 0.25)

    print("split detection, on XLK's measured numbers")
    nav_a, nav_b = 291.0433, 146.6173
    sh_a, sh_b = 325_805_897, 650_611_794
    tna_a, tna_b = 94_823_623_836, 95_390_938_128
    r = nav_a / nav_b
    check("the nav ratio rounds to two", round(r) == 2, f"{r:.6f}")
    check("it is within 2% of exactly two", abs(r - 2) / 2 < 0.02)
    check("assets are continuous across it, which is the confirmation",
          abs(tna_b / tna_a - 1.0) < 0.10, f"{tna_b / tna_a:.6f}")
    check("shares adjusted by the factor leave a creation-unit multiple",
          (sh_b / 2 - sh_a) % 50_000 == 0, f"{sh_b / 2 - sh_a:,.0f}")
    check("XLF on the same day is not a split",
          abs(52.8481 / 53.5217 - 1.0) < 0.02)

    print("§16.4's fee, and the interval it forces")
    check("the fee table records the four comparison funds as unread, not zero",
          all(FIXED_CREATION_FEE_USD[k] is None
              for k in ("SPDW", "SPEM", "SPAB", "JNK")))
    check("XLC carries the SAI's lower fee", FIXED_CREATION_FEE_USD["XLC"] == 250.0)
    check("the other ten sector funds carry 500",
          {FIXED_CREATION_FEE_USD[k] for k in
           ("XLB", "XLE", "XLF", "XLI", "XLK", "XLP", "XLRE", "XLU", "XLV", "XLY")} == {500.0})
    f_max_xlf = 500.0 / (CREATION_UNIT_SHARES * 53.69)
    check("one unit of XLF gives about 1.9 basis points",
          1.5e-4 < f_max_xlf < 2.5e-4, f"{f_max_xlf * 1e4:.3f} bp")
    check("the interval is ordered: lambda at f=0 is the upper end",
          lam(1e-4, 0.0) > lam(1e-4, f_max_xlf))

    print("§6's stress measure")
    import random as _r
    _r.seed(11)
    # A constant *log* return is the zero-vol case. An arithmetic ramp is not:
    # its log returns drift down, and it reads 1.7e-07 rather than 0. The first
    # version asserted < 1e-9 on the ramp and failed, correctly.
    geo = [100.0 * (1.0001 ** i) for i in range(200)]
    v = realised_vol(geo)
    check("a constant log return has zero realised vol",
          v[-1] is not None and v[-1] < 1e-12, str(v[-1]))
    ramp = [100.0 * (1.0 + 0.0001 * i) for i in range(200)]
    vr = realised_vol(ramp)
    check("an arithmetic ramp is not the zero case, and is tiny not zero",
          1e-9 < vr[-1] < 1e-5, f"{vr[-1]:.3e}")
    check("the window is not filled before it is full",
          v[STRESS_WINDOW - 2] is None)
    noisy = [100.0]
    for i in range(199):
        noisy.append(noisy[-1] * (1.0 + _r.gauss(0, 0.01)))
    vn = realised_vol(noisy)
    check("a 1% daily series reads about 1%",
          vn[-1] is not None and 0.005 < vn[-1] < 0.02, f"{vn[-1]:.5f}")
    check("vol is higher for the noisy series than the smooth one",
          vn[-1] > v[-1])

    print("§17's rectangle")
    check("a percent string becomes a fraction",
          spread_to_fraction("0.02%") == (0.0002, "percent_string"))
    check("a bare number is flagged, not divided by a hundred on faith",
          spread_to_fraction("0.0002")[1] == "bare_number_assumed_fraction")
    check("a zero floor parses to zero, not to None",
          spread_to_fraction("0.00%")[0] == 0.0)
    check("missing is missing", spread_to_fraction(None) == (None, "missing"))
    check("the quantum is one basis point", SPREAD_QUANTUM == 0.0001)
    # worst corner: bigger |λ| over smaller √N
    sn = 0.0002
    sn_lo, sn_hi = sn - SPREAD_QUANTUM / 2, sn + SPREAD_QUANTUM / 2
    # XLRE's measured f_max (2.389 bp) against its published floor (0.02%),
    # which is the case §17.4 predicts will come back indeterminate. The first
    # version used 1.9 bp here and the straddle check failed, correctly: at
    # 1.9 bp the worst corner is 0.933 and nothing straddles 1.
    lam_hi_, f_ = math.log1p(0.00005), 0.0002389
    lam_lo_ = lam_hi_ - math.log1p(f_)
    check("the fee end is the larger magnitude here", abs(lam_lo_) > abs(lam_hi_))
    check("the worst corner exceeds the best corner",
          max(abs(lam_hi_), abs(lam_lo_)) / sn_lo
          > min(abs(lam_hi_), abs(lam_lo_)) / sn_hi)
    check("a case that passes at the best corner and fails at the worst exists, "
          "which is what the three-way verdict is for",
          min(abs(lam_hi_), abs(lam_lo_)) / sn_hi < 1.0
          <= max(abs(lam_hi_), abs(lam_lo_)) / sn_lo)
    check("tau matches the creation unit measured in §16.2",
          F1_TAU == CREATION_UNIT_SHARES)

    print("§24: the two floors, and how the tick gave the cost floor away")
    check("the measurement quantum is the half-cent grid over NAV",
          abs(floor_measurement(50.0) - 1e-4) < 1e-12,
          f"{floor_measurement(50.0)*1e4:.3f} bp at NAV 50")
    check("the rounding standard deviation is the quantum over root twelve",
          abs(floor_measurement(50.0, "sd") * math.sqrt(12) -
              floor_measurement(50.0)) < 1e-15)
    check("the measurement floor falls when the price rises, as a floor must",
          floor_measurement(200.0) < floor_measurement(50.0))
    # §24.3's logic, made into an assertion rather than a paragraph
    ticks = {"XLK": 0.469, "XLV": 0.685, "XLP": 1.228, "XLF": 1.919, "XLRE": 2.389}
    lams = {"XLK": 1.268, "XLV": 1.655, "XLP": 1.285, "XLF": 1.492, "XLRE": 1.398}
    tspan = max(ticks.values()) / min(ticks.values())
    lspan = max(lams.values()) / min(lams.values())
    check("the tick spans about five-fold", 4.5 < tspan < 5.5, f"{tspan:.2f}")
    check("**lambda does not, so it is not quantisation**", lspan < 2.0,
          f"{lspan:.2f}")
    # A synthetic λ that really is quantisation: a fixed multiple of the fund's
    # own half tick. Its span must equal the tick's, which is the comparison the
    # real numbers fail. (The first version compared tspan with itself and could
    # not fail; an inert check is worse than no check.)
    fake = {k: 0.6 * (v / 2) for k, v in ticks.items()}
    fspan = max(fake.values()) / min(fake.values())
    check("a synthetic quantisation lambda spans exactly as the tick does",
          abs(fspan - tspan) < 1e-9, f"{fspan:.2f} vs {tspan:.2f}")
    check("and the measured lambda does not, which is the whole argument",
          lspan < tspan / 2, f"{lspan:.2f} vs {tspan:.2f}")
    check("against the measurement floor every fund exceeds one",
          all(lams[k] / (ticks[k] / 2) > 1.0 for k in ticks),
          f"min {min(lams[k]/(ticks[k]/2) for k in ticks):.2f}")
    # **Pooling breaks the test**, and the first --grid run printed the pooled
    # figure. A fund whose λ is dominated by a registered confound must not sit
    # in a comparison about quantisation.
    stale = {"SPDW": (2.315, 27.608), "SPEM": (2.145, 39.387), "JNK": (1.037, 11.465)}
    pooled_t = list(ticks.values()) + [v[0] for v in stale.values()]
    pooled_l = list(lams.values()) + [v[1] for v in stale.values()]
    p_tspan = max(pooled_t) / min(pooled_t)
    p_lspan = max(pooled_l) / min(pooled_l)
    check("pooling the stale-NAV funds inverts the comparison",
          p_lspan > p_tspan, f"tick {p_tspan:.2f}x vs |λ| {p_lspan:.2f}x")
    check("and the US equity group alone does not", lspan < tspan,
          f"tick {tspan:.2f}x vs |λ| {lspan:.2f}x")

    print("§26: the wall and the auction, told apart on synthetic series")

    def margins(deltas, shares=1.0e9):
        zero = sum(1 for d in deltas if d < 0.5) / len(deltas)
        mv = [d / shares for d in deltas if d >= 0.5]
        return zero, (sorted(mv)[len(mv) // 2] if mv else None)

    wall_calm = [50_000.0 * k for k in (1, 2, 3, 4, 0, 1, 2, 3, 0, 1)]
    wall_stress = [0.0] * 7 + [50_000.0 * k for k in (1, 2, 3)]
    auc_calm = list(wall_calm)
    auc_stress = [50_000.0 * k for k in (1, 1, 1, 1, 0, 1, 1, 1, 0, 1)]
    wz_c, wi_c = margins(wall_calm)
    wz_s, wi_s = margins(wall_stress)
    az_c, ai_c = margins(auc_calm)
    az_s, ai_s = margins(auc_stress)
    check("a wall raises the extensive margin", wz_s > wz_c,
          f"{wz_c:.2f} -> {wz_s:.2f}")
    check("a wall leaves the intensive margin roughly alone",
          abs(wi_s / wi_c - 1.0) < 0.35, f"x{wi_s/wi_c:.2f}")
    check("an auction leaves the extensive margin alone",
          abs(az_s - az_c) < 0.05, f"{az_c:.2f} -> {az_s:.2f}")
    check("an auction lowers the intensive margin", ai_s < ai_c,
          f"x{ai_s/ai_c:.2f}")
    check("**the two are distinguishable by the pair, not by either alone**",
          (wz_s > wz_c) != (az_s > az_c) or (wi_s < wi_c) != (ai_s < ai_c))

    print("§24.7: the conservative corner and the mechanical direction both flip")
    check("for a zero claim the conservative corner is max|λ| over min floor",
          (0.002 / 0.001) > (0.001 / 0.002))
    check("for a resolvability claim it is min|λ| over max floor, which is "
          "the smaller ratio", (0.001 / 0.002) < (0.002 / 0.001))
    fq_calm, fq_stress = floor_measurement(60.0), floor_measurement(50.0)
    check("a falling NAV raises the measurement floor", fq_stress > fq_calm)
    flat_lam = 1.4e-4
    check("**so with lambda held flat the ratio falls in stress**, the opposite "
          "of §18.6's fee trap",
          flat_lam / fq_stress < flat_lam / fq_calm,
          f"{flat_lam/fq_calm:.3f} -> {flat_lam/fq_stress:.3f}")
    # The mechanical part, as an identity rather than as a restatement. The
    # first version wrote `abs(x - x) == 0`, which cannot fail.
    varying_stress = flat_lam / fq_stress
    frozen_stress = flat_lam / fq_calm
    check("varying over frozen is exactly the NAV ratio, so the mechanical part "
          "is fully accounted",
          abs(varying_stress / frozen_stress - 50.0 / 60.0) < 1e-12,
          f"{varying_stress/frozen_stress:.6f} vs {50.0/60.0:.6f}")

    print("§22.2's degeneracy, on synthetic premiums of two different sizes")

    def pi_two_route(prems, fee=0.0):
        cells = [(-math.log1p(p), -math.log1p(fee)) for p in prems]
        allv = [x for c in cells for x in c]
        n = len(allv)
        mu = sum(allv) / n
        tot = sum((x - mu) ** 2 for x in allv) / n
        within = sum(((a - b) / 2) ** 2 for a, b in cells) * 2 / n
        # Exact null: every traversal carries the same accumulated ω, so there
        # is no variance to apportion. 0/0 is 0 here by definition of the
        # statistic, and it is written out rather than left to raise.
        return 0.0 if tot == 0.0 else within / tot

    _r.seed(3)
    tiny = [_r.gauss(0, 0.0001) for _ in range(2000)]
    huge = [p * 100 for p in tiny]
    p_tiny, p_huge = pi_two_route(tiny), pi_two_route(huge)
    check("a one-basis-point premium gives pi about one half",
          abs(p_tiny - 0.5) < 0.02, f"{p_tiny:.4f}")
    check("a hundred-basis-point premium gives the same pi",
          abs(p_huge - 0.5) < 0.02, f"{p_huge:.4f}")
    check("**pi does not move with the size of lambda on two routes**",
          abs(p_tiny - p_huge) < 0.02, f"{abs(p_tiny - p_huge):.5f}")
    check("a non-zero fee pushes it above one half",
          pi_two_route(tiny, fee=0.0005) > p_tiny)
    check("an exactly zero premium and zero fee give pi = 0, the hard null",
          pi_two_route([0.0] * 100, fee=0.0) == 0.0)

    print("§19.5's discriminators, on series whose answer is known")
    _r.seed(7)
    white = [_r.gauss(0, 1) for _ in range(400)]
    ac_white = pearson(white[:-1], white[1:])
    check("white noise has autocorrelation near zero",
          abs(ac_white) < 0.12, f"{ac_white:.4f}")
    ar = [0.0]
    for _ in range(399):
        ar.append(0.7 * ar[-1] + _r.gauss(0, 1))
    ac_ar = pearson(ar[:-1], ar[1:])
    check("an AR(1) with phi=0.7 reads clearly positive",
          ac_ar > 0.5, f"{ac_ar:.4f}")
    check("D1 separates them", ac_ar - ac_white > 0.4)
    # a folded series with a bigger scale but no persistence: the A-2 rise
    # without the D1 signal, which is exactly the null §19.3 named
    loud = [3.0 * x for x in white]
    check("scaling the noise does not create persistence",
          abs(pearson(loud[:-1], loud[1:]) - ac_white) < 1e-12)
    check("but it does raise the folded median, which is why A-2 alone cannot "
          "tell them apart",
          sorted(abs(x) for x in loud)[200] > sorted(abs(x) for x in white)[200])
    print("D1b, §20.5: removing a common component leaves the idiosyncratic one")
    common_ = [_r.gauss(0, 1) for _ in range(400)]
    idio = [0.0]
    for _ in range(399):
        idio.append(0.6 * idio[-1] + _r.gauss(0, 0.5))
    raw = [c + i for c, i in zip(common_, idio)]
    resid = [x - c for x, c in zip(raw, common_)]
    ac_raw = pearson(raw[:-1], raw[1:])
    ac_res = pearson(resid[:-1], resid[1:])
    check("a white common component dilutes the raw autocorrelation",
          ac_res > ac_raw, f"raw {ac_raw:.4f} -> resid {ac_res:.4f}")
    check("and the residual recovers the idiosyncratic persistence",
          ac_res > 0.4, f"{ac_res:.4f}")
    flat_idio = [_r.gauss(0, 0.5) for _ in range(400)]
    raw2 = [c + i for c, i in zip(common_, flat_idio)]
    res2 = [x - c for x, c in zip(raw2, common_)]
    check("with no idiosyncratic persistence the residual reads near zero",
          abs(pearson(res2[:-1], res2[1:])) < 0.12)

    check("pearson refuses a constant series",
          pearson([1.0, 1.0, 1.0], [1.0, 2.0, 3.0]) is None)
    check("pearson refuses too few points", pearson([1.0], [2.0]) is None)
    sym = [0.001, -0.001, 0.002, -0.002]
    check("a symmetric premium set is half at a discount",
          sum(1 for x in sym if x < 0) / len(sym) == 0.5)

    print("§18.6: why the gradient is immune to the quantum and the fee is not")
    sn_a, sn_b = 0.0001, 0.0002        # two funds, one quantum apart
    lam_calm, lam_stress = 0.00008, 0.00016
    check("the floor cancels from the direction",
          (lam_stress / sn_a > lam_calm / sn_a)
          == (lam_stress / sn_b > lam_calm / sn_b))
    check("and from the ratio of medians, exactly",
          abs((lam_stress / sn_a) / (lam_calm / sn_a)
              - (lam_stress / sn_b) / (lam_calm / sn_b)) < 1e-12)
    # the fee end: NAV falls in stress, so f_max rises with no change in lambda
    fee_, unit_ = 500.0, float(CREATION_UNIT_SHARES)
    nav_calm_, nav_stress_ = 60.0, 50.0
    f_calm = fee_ / (unit_ * nav_calm_)
    f_stress = fee_ / (unit_ * nav_stress_)
    check("a falling NAV raises f_max on its own", f_stress > f_calm)
    flat = 0.00002
    r_calm = abs(math.log1p(flat) - math.log1p(f_calm)) / sn_b
    r_stress = abs(math.log1p(flat) - math.log1p(f_stress)) / sn_b
    check("with lambda held flat the fee end still rises, which is the trap",
          r_stress > r_calm, f"{r_calm:.4f} -> {r_stress:.4f}")
    # §49.5(7): the first version compared an expression with itself. What is
    # worth asserting is that the f=0 end is a function of the premium alone.
    check("the f=0 end depends on the premium and not on NAV",
          abs(path_sum(LOOP, {"nav": 40.0, "price": 40.0 * (1 + flat), "fee": 0.0})
              - path_sum(LOOP, {"nav": 400.0, "price": 400.0 * (1 + flat),
                                "fee": 0.0})) < 1e-15)

    print("attribution truth table, on the first run's own numbers")

    def attribute(no_fee, no_quantum):
        fee_needed = no_fee < 1.0
        quantum_needed = no_quantum < 1.0
        if not fee_needed:
            return "floor_indeterminate"
        if not quantum_needed:
            return "fee_indeterminate"
        return "fee_and_floor_indeterminate"

    check("XLC: the failure survives removing the fee, so the floor does it",
          attribute(2.5874, 1.5205) == "floor_indeterminate")
    check("XLB: it survives removing the widening, so the fee does it",
          attribute(0.8740, 1.0203) == "fee_indeterminate")
    check("XLU: both removals fix it, so neither alone does it",
          attribute(0.7499, 0.9086) == "fee_and_floor_indeterminate")
    check("XLRE: fee necessary, quantum not, so fee",
          attribute(0.8477, 1.2215) == "fee_indeterminate")
    check("the three labels are exhaustive over the quadrants",
          {attribute(a, b) for a in (0.5, 1.5) for b in (0.5, 1.5)}
          == {"floor_indeterminate", "fee_indeterminate",
              "fee_and_floor_indeterminate"})

    print("cell formatting, the helper the first --inputs run tripped over")
    check("None gives an empty string", _cell(None) == "")
    check("a datetime gives an iso date",
          _cell(datetime(2026, 8, 13)) == "2026-08-13")
    check("a float keeps enough digits", _cell(0.0002) == "0.0002")
    check("a string passes through", _cell("0.02%") == "0.02%")

    print("ticker cleaning, because the sheet writes XLF with a registered mark")
    check("a registered mark is stripped", _clean_ticker("XLF®") == "XLF")
    check("whitespace is stripped", _clean_ticker(" xly ") == "XLY")
    check("None gives empty", _clean_ticker(None) == "")

    print("F1 scanning, §5 and discipline 8")
    check("thresholds are a scan, not a single number", len(F1_THRESHOLDS) >= 5)
    check("zero is on the grid, so 'any change at all' is one of the points",
          0 in F1_THRESHOLDS)
    check("the window is stated in trading days", F1_WINDOW == 5)
    check("a creation-unit-sized change is on the grid",
          50_000 in F1_THRESHOLDS)

    print("§31: the identity that ties D1 to D1b, and its four cells")
    # **A regression guard on a revert, and it is registered as that rather
    # than as a test of the world.** The shipped divisor was `se_stress` alone
    # for four sections. These two lines fire on any version that goes back to
    # it: on the real pair counts the two divisors differ by `39%`, which is
    # the whole of the correction. **They say nothing about the data**, and
    # §49's lesson is that a check whose only property is being true is not a
    # check, so what they are is written here rather than implied.
    _SE_C, _SE_S, _D = 1.0 / math.sqrt(209), 1.0 / math.sqrt(194), 0.069132
    check("the difference's se is the root sum of squares, not either end's",
          abs(_se_of_difference(_SE_C, _SE_S)
              - math.sqrt(_SE_C ** 2 + _SE_S ** 2)) < 1e-15
          and _se_of_difference(_SE_C, _SE_S) > 1.35 * max(_SE_C, _SE_S),
          f"se(diff)={_se_of_difference(_SE_C, _SE_S):.4f} against "
          f"se_stress={_SE_S:.4f}")
    check("the corrected divisor moves d(rho_c) off 0.96 and onto 0.69",
          abs(_D / _se_of_difference(_SE_C, _SE_S) - 0.693) < 0.002
          and abs(_D / _SE_S - 0.963) < 0.002,
          "the wrong divisor is kept here on purpose so a revert fails")
    check("a missing end gives no difference se rather than a wrong one",
          _se_of_difference(None, _SE_S) is None
          and _se_of_difference(_SE_C, None) is None)
    # **This one bites on the divergence, and it fired on the first print.**
    # `8 / d^2` on the `recon` arm's `d = 9.7e-05` returns `859,039,612` pairs.
    # The flag beside it has to be false there and true nowhere near it.
    _SD = _se_of_difference(_SE_C, _SE_S)
    check("the window figure diverges as the difference goes to zero",
          8.0 / (9.7e-05 ** 2) > 8.0e8 and 8.0 / (_D ** 2) < 2.0e3,
          f"recon-shaped {8.0 / (9.7e-05 ** 2):.3e} against "
          f"main-shaped {8.0 / (_D ** 2):.0f}")
    check("the flag reads the estimate against its own se, not a constant",
          (_D / _SD >= 1.0) is False and (0.15 / _SD >= 1.0) is True,
          f"0.0691 is {_D / _SD:.2f} se so the flag is false; "
          f"0.15 is {0.15 / _SD:.2f} se so it is true")
    check("§31.5's table has exactly the four sign cells",
          len(DECOMP_CELLS) == 4 and set(DECOMP_CELLS) ==
          {(a, b) for a in ("rises", "falls")
           for b in ("rises", "flat_or_falls")})
    check("V_e falls with rho_c rising files as pinning",
          decomp_verdict(0.8, +0.05)[0] == "pinning")
    check("V_e falls with rho_c flat files as composition, not pinning",
          decomp_verdict(0.8, -0.05)[0] == "composition")
    check("V_e rises with rho_c rising files as both, leaving §21.2 open",
          decomp_verdict(1.2, +0.05)[0] == "both")
    check("V_e rises with rho_c flat files as a defect, not as a reading",
          decomp_verdict(1.2, -0.05)[0] == "inconsistent_with_d1")
    check("an unestimable component is indeterminate and never a cell",
          decomp_verdict(None, 0.05)[0] == "indeterminate"
          and decomp_verdict(1.2, None)[0] == "indeterminate")
    check("the boundary ratio 1.0 counts as a fall, stated not implied",
          decomp_verdict(1.0, 0.05)[0] == "pinning")
    check("six of eleven is a majority and five is not",
          MARKET_STRESS_MAJORITY * 2 > len(MAIN_ARM)
          and (MARKET_STRESS_MAJORITY - 1) * 2 <= len(MAIN_ARM))

    # The first --decomp run refused here, correctly: `order` is a row number in
    # each fund's own navhist and XLC's history is 1646 rows where an older
    # fund's is 5310. **The absolute values were never comparable.**
    _s1 = {"a": {"order": 100}, "b": {"order": 101}, "c": {"order": 102}}
    _s2 = {"a": {"order": 5}, "b": {"order": 6}, "c": {"order": 7}}
    _do, _cl, _an = market_day_order({"X": _s1, "Y": _s2}, ["X", "Y"])
    check("a different inception date is a constant offset, not a clash",
          not _cl and _do == {"a": 0, "b": 1, "c": 2} and _an == "a")
    _s3 = {"a": {"order": 5}, "b": {"order": 7}, "c": {"order": 8}}
    _do2, _cl2, _ = market_day_order({"X": _s1, "Y": _s3}, ["X", "Y"])
    check("a genuinely different calendar still clashes, so the guard keeps "
          "its teeth after the fix", len(_cl2) == 2)
    _do3, _cl3, _an3 = market_day_order(
        {"X": {"a": {"order": 1}}, "Y": {"b": {"order": 1}}}, ["X", "Y"])
    check("no shared day is a refusal, not an empty run",
          _an3 is None and len(_cl3) == 1 and not _do3)

    _sq = [("d0", 0), ("d1", 1), ("d2", 3), ("d3", 4)]
    _sv = {"d0": 1.0, "d1": 2.0, "d2": 3.0, "d3": 4.0}
    _px, _py = lag1_pairs(_sq, lambda d: _sv[d], lambda d: True)
    check("a gap in order is not a lag of one",
          _px == [1.0, 3.0] and _py == [2.0, 4.0])
    _qx, _qy = lag1_pairs(_sq, lambda d: _sv[d], lambda d: d == "d1")
    check("a pair is classed by the later day, D1's convention",
          _qx == [1.0] and _qy == [2.0])
    check("a constant series gives no scale, so nothing closes by accident",
          _geo_scale([1.0] * 10, [1.0] * 10) is None)

    # §31.1 on two periodic series whose periods are coprime, so both lagged
    # cross-covariances are **exactly** zero over a whole number of common
    # cycles. 201 points give 200 pairs, which is ten cycles of twenty.
    _cs, _es = [1.0, 0.0, -1.0, 0.0], [1.0, 1.0, -1.0, -1.0, 0.0]
    _n = 201
    _cc = [_cs[i % 4] for i in range(_n)]
    _ee = [_es[i % 5] for i in range(_n)]
    _lam = [_cc[i] + _ee[i] for i in range(_n)]
    _ord = [(i, i) for i in range(_n)]

    def _closure(evals, keep_e=None):
        lx, ly = lag1_pairs(_ord, lambda d: _lam[d], lambda d: True)
        cx, cy = lag1_pairs(_ord, lambda d: _cc[d], lambda d: True)
        ex, ey = lag1_pairs(_ord, lambda d: evals[d], keep_e or (lambda d: True))
        gc, ge = _geo_scale(cx, cy), _geo_scale(ex, ey)
        return abs(pearson(lx, ly) * _geo_scale(lx, ly)
                   - pearson(cx, cy) * gc
                   - pearson(ex, ey) * ge) / (gc + ge)

    check("§31.1 closes exactly when the components are the real ones",
          _closure(_ee) < 1e-9, f"residual {_closure(_ee):.2e}")

    # The four-term split. **This identity cannot fail unless the pairing is
    # wrong**, which is why it replaces §31.6(1) as the pair check, and why
    # §31.6(1)'s residual turned out to be a measurement of `cross`.
    def _ft(cvals, evals):
        lam = [cvals[i] + evals[i] for i in range(_n)]
        lx, ly = lag1_pairs(_ord, lambda d: lam[d], lambda d: True)
        cx, cy = lag1_pairs(_ord, lambda d: cvals[d], lambda d: True)
        ex, ey = lag1_pairs(_ord, lambda d: evals[d], lambda d: True)
        return four_terms(lx, ly, cx, cy, ex, ey)

    _t0 = _ft(_cc, _ee)
    check("the four-term identity holds to machine precision",
          abs(_t0["identity_residual"]) < 1e-15,
          f"residual {_t0['identity_residual']:.1e}")
    check("on coprime periods the cross term is exactly zero, which is why "
          "§31.1 closed on the synthetic and not on the data",
          abs(_t0["cross"]) < 1e-15)
    # A residual that genuinely lags the common component. The three-term form
    # must break and the four-term form must not. **The period-4 series is not
    # used here**: on it the two halves of `cross` cancel exactly, which is a
    # property of that series and not of the estimator, and a control that
    # cancels is a control that never bites.
    _lead = [0.5 * _ee[i - 1] for i in range(_n)]
    _t1 = _ft(_ee, _lead)
    check("a residual that lags the common component gives a non-zero cross",
          abs(_t1["cross"]) > 1e-6, f"cross {_t1['cross']:.4f}")
    check("and the four-term identity still holds, so the cross is the whole "
          "of what §31.6(1) was measuring",
          abs(_t1["identity_residual"]) < 1e-14)
    check("while the three-term closure breaks on the same series",
          abs(_t1["cov_lambda"] - _t1["cov_cc"] - _t1["cov_ee"]) > 1e-6)
    check("§32.6: the cross share is normalised by a sum of scales, which "
          "cannot vanish, and the old one is kept under its own name",
          _t1["cross_share"] is not None
          and "cross_over_cov_lambda" in _t1)

    print("§32.5's three rows, thresholds fixed before the run")
    check("a collapsed cross term with clean cells is the artefact row",
          loo_verdict(0.05, 0.01, 20, 11)[0] == "artefact_confirmed")
    check("a surviving cross term that signs consistently is propagation",
          loo_verdict(0.05, 0.048, 3, 10)[0] == "propagation")
    check("a surviving cross term with no consistent sign is neither",
          loo_verdict(0.05, 0.048, 3, 6)[0] == "neither")
    check("collapse is evaluated first, so a collapsed term is not read "
          "for its sign", loo_verdict(0.05, 0.001, 22, 11)[0]
          == "artefact_confirmed")
    check("halving alone is not enough, the cells must clear too",
          loo_verdict(0.05, 0.02, 3, 6)[0] == "neither")
    check("an unestimable cross term is indeterminate",
          loo_verdict(None, 0.01, 22, 11)[0] == "indeterminate")
    check("the sign majority is a majority of eleven",
          LOO_SIGN_MAJORITY * 2 > len(MAIN_ARM)
          and LOO_CLEAR_CELLS < 2 * len(MAIN_ARM))

    print("§33: the cross term against its own shift distribution")
    check("a count near chance is row 1, not distinguishable",
          shift_verdict(1, 22)[0] == "not_distinguishable")
    check("a clear majority in the tail is row 2, real",
          shift_verdict(15, 22)[0] == "real")
    check("a few cells only is row 3, not established either way",
          shift_verdict(7, 22)[0] == "not_established")
    check("the row 1 boundary is stated: three yes, four no",
          shift_verdict(3, 22)[0] == "not_distinguishable"
          and shift_verdict(4, 22)[0] == "not_established")
    check("the row 2 boundary is stated: twelve yes, eleven no",
          shift_verdict(12, 22)[0] == "real"
          and shift_verdict(11, 22)[0] == "not_established")
    check("no cells is indeterminate", shift_verdict(0, 0)[0] == "indeterminate")
    check("row 1's cut sits near the binomial mean of 22 x 0.05",
          SHIFT_ROW1_MAX < 2 * 22 * 0.05 + 2)

    _r1, _r2 = _lcg(400, 7), _lcg(400, 99)
    _pi = [(i, i + 1) for i in range(399)]
    _n0 = shift_null(_pi, _r1, _r2, 400)
    _o0 = abs((_cov([_r1[i] for i, _ in _pi], [_r2[j] for _, j in _pi])
               + _cov([_r2[i] for i, _ in _pi], [_r1[j] for _, j in _pi]))
              / (_geo_scale([_r1[i] for i, _ in _pi], [_r1[j] for _, j in _pi])
                 + _geo_scale([_r2[i] for i, _ in _pi],
                              [_r2[j] for _, j in _pi])))
    _p0 = sum(1 for x in _n0 if x < _o0) / len(_n0)
    check("two independent series sit inside their own shift distribution",
          _p0 < SHIFT_NULL_QUANTILE, f"percentile {_p0:.3f}, {len(_n0)} shifts")

    _lag = [0.0] + [0.8 * _r1[i - 1] + 0.2 * _r2[i] for i in range(1, 400)]
    _n1 = shift_null(_pi, _r1, _lag, 400)
    _o1 = abs((_cov([_r1[i] for i, _ in _pi], [_lag[j] for _, j in _pi])
               + _cov([_lag[i] for i, _ in _pi], [_r1[j] for _, j in _pi]))
              / (_geo_scale([_r1[i] for i, _ in _pi], [_r1[j] for _, j in _pi])
                 + _geo_scale([_lag[i] for i, _ in _pi],
                              [_lag[j] for _, j in _pi])))
    _p1 = sum(1 for x in _n1 if x < _o1) / len(_n1)
    check("a series that genuinely lags the other lands in the tail, so the "
          "null can detect what it is for",
          _p1 >= SHIFT_NULL_QUANTILE, f"percentile {_p1:.3f}")
    # Two negative controls, because a tolerance that nothing can breach is
    # not a tolerance. **A rotation of `e` is deliberately not one of them**:
    # a shift leaves a periodic series' own autocovariance unchanged, so it
    # closes, and using it would have been an inert cell.
    _scaled = [1.5 * x for x in _ee]
    check("a component with the wrong scale breaks closure",
          _closure(_scaled) > 0.01, f"residual {_closure(_scaled):.3f}")
    # **Closure does not subsume the pair-count check.** On a series periodic
    # enough, halving the pair set leaves the covariance per pair unchanged and
    # the residual is still zero. That is why §31.6(2) counts pairs separately
    # instead of trusting §31.6(1) to notice.
    _full = lag1_pairs(_ord, lambda d: _ee[d], lambda d: True)[0]
    _half = lag1_pairs(_ord, lambda d: _ee[d], lambda d: d % 2 == 0)[0]
    check("a different day set is caught by the pair count and NOT by closure, "
          "which is why §31.6(2) is a separate check",
          len(_half) != len(_full)
          and _closure(_ee, lambda d: d % 2 == 0) < 1e-9,
          f"{len(_half)} vs {len(_full)} pairs, residual still "
          f"{_closure(_ee, lambda d: d % 2 == 0):.1e}")

    print(f"\n{len(fails)} failed" if fails else "\nall passed")
    return 1 if fails else 0


def _lcg(n: int, seed: int) -> list:
    """A deterministic sequence for the selftest only. **No stage uses random
    numbers**; §33's null is every shift rather than a sample of them."""
    x, out = seed, []
    for _ in range(n):
        x = (1103515245 * x + 12345) % 2147483648
        out.append(x / 2147483648.0 - 0.5)
    return out


def _raises(fn) -> bool:
    try:
        fn()
        return False
    except Exception:
        return True


# ---------------------------------------------------------------------------
# §31: `ρ_λ = w·ρ_c + (1 − w)·ρ_e` ties D1 to D1b. D1 measured `ρ_λ`, D1b
# measured `ρ_e`, and `ρ_c` with `w` have never been measured. §31 was registered
# before any of this ran, including the four-cell table below.
# ---------------------------------------------------------------------------

DECOMP_OUT = ROOT / "results" / "b9_decomp.json"
MARKET_STRESS_MAJORITY = 6      # §31.4: six of the eleven main-arm funds

DECOMP_CELLS = {
    ("falls", "rises"): (
        "pinning",
        "§31.5 row 1. A common object binds and becomes more persistent while "
        "the funds are pushed together, and the residual is what is left after "
        "the pinning. §21.2 resolves toward a commonly binding constraint."),
    ("falls", "flat_or_falls"): (
        "composition",
        "§31.5 row 2. `w` rises and weight shifts onto a component that did not "
        "itself change. Nothing became more persistent, and D1's rise loses its "
        "economic content."),
    ("rises", "rises"): (
        "both",
        "§31.5 row 3. A common object binds and idiosyncratic variance grows "
        "with it. §21.2 stays open, and D1b's fall is then partly a variance "
        "effect rather than a whitening."),
    ("rises", "flat_or_falls"): (
        "inconsistent_with_d1",
        "§31.5 row 4. With `w` falling and `ρ_e` falling the identity cannot "
        "produce D1's measured rise. That is a defect in the pair construction "
        "and belongs in §8's ledger, not in a reading."),
}


def decomp_verdict(ve_ratio, d_rho_c):
    """§31.5's table, **by sign, with no threshold beyond zero**. The standard
    error is printed beside the verdict so a sign inside the noise is visible as
    a sign inside the noise."""
    if ve_ratio is None or d_rho_c is None:
        return "indeterminate", "a component was not estimable"
    ve = "rises" if ve_ratio > 1.0 else "falls"
    rc = "rises" if d_rho_c > 0.0 else "flat_or_falls"
    return DECOMP_CELLS[(ve, rc)]


def lag1_pairs(ordered, value_of, keep):
    """Lag-one pairs on consecutive rows, classed by the **later** day, which is
    D1's convention (§19.5). A gap in `order` is not a lag of one and is
    dropped, so the three components are always paired on the same days."""
    prev, nxt = [], []
    for (da, oa), (db, ob) in zip(ordered, ordered[1:]):
        if ob - oa != 1 or not keep(db):
            continue
        va, vb = value_of(da), value_of(db)
        if va is None or vb is None:
            continue
        prev.append(va)
        nxt.append(vb)
    return prev, nxt


def _geo_scale(xs, ys):
    """`sqrt(Var(prev)·Var(next))`, which is the denominator Pearson actually
    divides by. Using it rather than a plain variance keeps §31.6's closure
    check honest about Pearson's own separate demeaning."""
    n = len(xs)
    if n < 2:
        return None
    mx, my = sum(xs) / n, sum(ys) / n
    sxx = sum((a - mx) ** 2 for a in xs)
    syy = sum((b - my) ** 2 for b in ys)
    if sxx <= 0 or syy <= 0:
        return None
    return math.sqrt(sxx * syy) / n


def _cov(xs, ys):
    """Sample covariance with each list demeaned by its own mean, which is what
    Pearson does. **Means are linear, so `λ = c + e` gives the four-term
    identity exactly under this convention**, separate demeaning and all."""
    n = len(xs)
    if n < 2:
        return None
    mx, my = sum(xs) / n, sum(ys) / n
    return sum((a - mx) * (b - my) for a, b in zip(xs, ys)) / n


def four_terms(lx, ly, cx, cy, ex, ey):
    """`Cov(λ_prev, λ_next)` split into the four terms it is identically equal
    to. **`cross` is not an error term.** It is `Cov(c_{t−1}, e_t)` plus
    `Cov(e_{t−1}, c_t)`, and it vanishes only when the funds' residuals are
    cross-sectionally independent at lag one, which §31.1 assumed and the first
    run falsified. Reported signed, and split, because the two halves are a
    lead-lag direction and not a magnitude."""
    out = {"cov_lambda": _cov(lx, ly), "cov_cc": _cov(cx, cy),
           "cov_ee": _cov(ex, ey), "cov_c_leads_e": _cov(cx, ey),
           "cov_e_leads_c": _cov(ex, cy)}
    if any(v is None for v in out.values()):
        return {k: None for k in list(out) + ["cross", "cross_share",
                                              "identity_residual"]}
    out["cross"] = out["cov_c_leads_e"] + out["cov_e_leads_c"]
    out["identity_residual"] = (out["cov_lambda"] - out["cov_cc"]
                                - out["cov_ee"] - out["cross"])
    # **§32.6.** The first print normalised by `cov_lambda`, whose sign changes
    # (XLF reads `ρ_λ = −0.005` on calm days), so the ratio blew up on the
    # denominator. `G_c + G_e` is a sum of scales and cannot vanish. The bad
    # figure is kept under its own name rather than removed, so that its range
    # cannot be rediscovered later as a finding.
    gc, ge = _geo_scale(cx, cy), _geo_scale(ex, ey)
    out["cross_share"] = (out["cross"] / (gc + ge)
                          if (gc is not None and ge is not None and gc + ge)
                          else None)
    out["cross_over_cov_lambda"] = (out["cross"] / out["cov_lambda"]
                                    if out["cov_lambda"] else None)
    return out


# §32.5's three rows, with the thresholds fixed here **before the run**.
LOO_COLLAPSE_FACTOR = 0.5      # median |cross| must at least halve
LOO_CLEAR_CELLS = 17           # of 22, under §31.6(1)'s own 0.01
LOO_SIGN_MAJORITY = 9          # of 11, for a sign to count as consistent


def loo_verdict(med_all, med_loo, cells_clear, sign_max):
    """§32.5. Evaluated collapse first: a cross term that has collapsed has
    nothing left to sign, so the sign test only speaks for a surviving one."""
    if med_all is None or med_loo is None:
        return "indeterminate", "the cross term was not estimable"
    if med_loo < LOO_COLLAPSE_FACTOR * med_all and cells_clear >= LOO_CLEAR_CELLS:
        return "artefact_confirmed", (
            "§32.5 row 1. The cross term was the equal-weight mean including "
            "the fund itself. `c_t` is a poor common component on an arm this "
            "heterogeneous, and §31.1's identity is recovered on the corrected "
            "residual.")
    if sign_max >= LOO_SIGN_MAJORITY:
        return "propagation", (
            "§32.5 row 3. The cross term survives leave-one-out **and signs "
            "consistently**, which is a lead-lag between the market-wide wedge "
            "and individual funds. Registered as the interesting row and as "
            "the one expected not to happen.")
    return "neither", (
        "§32.5 row 2. The residuals carry a cross-sectional structure that no "
        "mean-based decomposition removes, and `ρ_c` against `ρ_e` is the wrong "
        "pair of coordinates on this arm.")


def _var(xs):
    n = len(xs)
    if n < 2:
        return None
    m = sum(xs) / n
    return sum((x - m) ** 2 for x in xs) / n


def _se_of_difference(se_a, se_b):
    """The standard error of `a - b` when `a` and `b` are estimated on disjoint
    samples.

    **One place, because the two callers of this quantity drifted apart into
    one caller using the wrong divisor.** `rho_c` is estimated twice, once on
    the calm pairs and once on the stress pairs, and those two sets partition
    the consecutive pairs of the window. Neither estimate is a known constant
    to the other, so the difference carries both sampling errors and its
    standard error is the root sum of squares, about `1.41x` either end's.

    Returns `None` if either end is missing, which is the same convention every
    other quantity in this module uses.
    """
    if se_a is None or se_b is None:
        return None
    return math.sqrt(se_a ** 2 + se_b ** 2)


def market_day_order(series, main):
    """§15.2 says the main arm shares a trading calendar over the connect
    window. **It does not say the funds share an index.** `order` is a row
    number in each fund's own navhist and the funds have different inception
    dates, so XLC's 2025-01-02 is row 1646 where an older fund's is row 5310.
    **Only the differences are comparable.**

    So each fund is normalised to a common anchor day and the agreement §15.2
    asserts is then checked on the normalised values. Returns
    `(day_order, clashes, anchor)`."""
    common_days = None
    for t in main:
        s = set(series[t])
        common_days = s if common_days is None else (common_days & s)
    if not common_days:
        return {}, [("no day is present in every main-arm fund", None, None, None)], None
    anchor = min(common_days)
    off = {t: series[t][anchor]["order"] for t in main}
    day_order, clash = {}, []
    for t in main:
        for d, v in series[t].items():
            o = v["order"] - off[t]
            if d in day_order and day_order[d] != o:
                clash.append((d, t, day_order[d], o))
            day_order[d] = o
    # §49.5(6): a map built from disagreeing calendars is returned empty rather
    # than last-writer-wins, so no later caller can read it as authoritative.
    return ({} if clash else day_order), clash, anchor


# §33: the cross term's own noise floor, measured rather than assumed.
SHIFT_MIN = 10               # §33.3: stated, not tuned, and not moved after
SHIFT_NULL_QUANTILE = 0.95
SHIFT_ROW1_MAX = 3           # Binomial(22, 0.05) has mean 1.1, sd 1.02
SHIFT_ROW2_MIN = 12          # a clear majority of the twenty-two cells


def shift_null(pairs_idx, cvals, evals, m):
    """§33.3. **Every** circular shift of `e` against `c` from `SHIFT_MIN` to
    `m − SHIFT_MIN`, with no random numbers, so the figure is reproducible to
    the digit. A shift preserves both series' own autocorrelation and destroys
    only the relation between them, **which is the null §31.1 asserted**. A
    permutation would destroy `ρ_e` as well and would therefore test a different
    null, one that calls the cross term significant whenever `e` is persistent
    at all."""
    cx = [cvals[i] for i, _ in pairs_idx]
    cy = [cvals[j] for _, j in pairs_idx]
    gc = _geo_scale(cx, cy)
    out = []
    if gc is None or m <= 2 * SHIFT_MIN + 1:
        return out
    for L in range(SHIFT_MIN, m - SHIFT_MIN):
        ex = [evals[(i + L) % m] for i, _ in pairs_idx]
        ey = [evals[(j + L) % m] for _, j in pairs_idx]
        ge = _geo_scale(ex, ey)
        if ge is None or not (gc + ge):
            continue
        out.append(abs((_cov(cx, ey) + _cov(ex, cy)) / (gc + ge)))
    return sorted(out)


def shift_verdict(above, cells):
    """§33.4's three rows. The count is of cells whose observed cross term sits
    at or above the 95th percentile of its own shift distribution."""
    if not cells:
        return "indeterminate", "no cell produced a shift distribution"
    if above <= SHIFT_ROW1_MAX:
        return "not_distinguishable", (
            "§33.4 row 1. The cross term is not distinguishable from zero. "
            "§31.1's premise stands, §31.6(1)'s tolerance was below the noise "
            "floor, and §32.2 and §32.4 are both withdrawn.")
    if above >= SHIFT_ROW2_MIN:
        return "real", (
            "§33.4 row 2. The cross term is real and §32.5's `neither` stands "
            "with a measured significance behind it. The one-factor "
            "decomposition is the wrong instrument on this arm, demonstrated "
            "rather than asserted.")
    return "not_established", (
        "§33.4 row 3. Neither way. Report the count, change nothing, stop: "
        "eleven funds sharing 0.703 of their variance is too small an arm.")


def decomp(tickers: list[str]) -> dict:
    """§31. Measures `ρ_c`, `V_c` and `V_e` by regime, checks §31.1's identity
    fund by fund, and files the reading into §31.5's four cells."""
    spreads = read_spread_table()
    series = {t: lambda_series(t, spreads) for t in (tickers or SAMPLE)}
    series = {t: s for t, s in series.items() if s}
    main = [t for t in MAIN_ARM if t in series]
    if len(main) < MARKET_STRESS_MAJORITY:
        print(f"§31 refuses to run: {len(main)} main-arm funds present and "
              f"§31.4's majority needs {MARKET_STRESS_MAJORITY} of "
              f"{len(MAIN_ARM)}.")
        return {}

    # --- §15.2 cross-checked rather than assumed --------------------------
    day_order, clash, anchor = market_day_order(series, main)
    if clash:
        print(f"§15.2 violated: {len(clash)} calendar clashes across the main "
              f"arm after anchoring, first {clash[0]}. §31 refuses to run.")
        return {}
    print(f"§15.2  calendar agrees across {len(main)} main-arm funds on "
          f"{len(day_order)} days, anchored at {anchor}")

    all_days = sorted(day_order)
    ordered = [(d, day_order[d]) for d in all_days]
    common = {d: (lambda v: sum(v) / len(v))(
                 [series[t][d]["lam"] for t in main if d in series[t]])
              for d in all_days}

    # --- §31.4's market-wide regime ---------------------------------------
    depth_of = {d: sum(1 for t in main
                       if d in series[t] and not series[t][d]["calm"])
                for d in all_days}
    mstress = {d: depth_of[d] >= MARKET_STRESS_MAJORITY for d in all_days}
    n_ms = sum(1 for d in all_days if mstress[d])

    head = {}
    for lab, want in (("calm", False), ("stress", True)):
        xs, ys = lag1_pairs(ordered, lambda d: common[d],
                            lambda d: mstress[d] is want)
        res = [series[t][d]["lam"] - common[d]
               for t in main for d in series[t] if mstress[d] is want]
        cvals = [common[d] for d in all_days if mstress[d] is want]
        head[lab] = {"rho_c": pearson(xs, ys), "pairs_c": len(xs),
                     "v_c": _var(cvals), "v_e": _var(res),
                     "n_days": len(cvals), "n_fund_days": len(res),
                     "se_rho_c": (1.0 / math.sqrt(len(xs))) if xs else None}
        vc, ve = head[lab]["v_c"], head[lab]["v_e"]
        head[lab]["w"] = (vc / (vc + ve)) if (vc is not None and ve) else None

    d_rho = (None if None in (head["calm"]["rho_c"], head["stress"]["rho_c"])
             else head["stress"]["rho_c"] - head["calm"]["rho_c"])
    # **The divisor, corrected 2026-08-18.** `Delta rho_c` is a difference of two
    # lag-one correlations estimated on **disjoint** pair sets: calm and stress
    # partition the consecutive pairs, `209 + 194 = 403 = 404 - 1`. Its standard
    # error is therefore `sqrt(se_calm^2 + se_stress^2)`, about `1.41x` either
    # end's. The earlier line divided by `se_stress` alone, which treats the
    # calm estimate as a known constant. It is not one, and the stage carried
    # `0.96 se` through four sections on that arithmetic; the difference's own
    # standard error gives `0.69`.
    #
    # **The consequence is not the ratio, it is the window.** Two standard
    # errors on the difference needs `se_diff <= |d| / 2`, and with equal `n`
    # per regime `se_diff = sqrt(2/n)`, hence `n = 8 / d^2` pairs per regime and
    # twice that in trading days. On `d = 0.0691` that is about `1,674` and
    # `3,348`, against the `837` and `1,600` the single-end divisor implied.
    se_d_rho = _se_of_difference(head["calm"]["se_rho_c"],
                                 head["stress"]["se_rho_c"])
    d_rho_in_se = (None if (d_rho is None or not se_d_rho)
                   else abs(d_rho) / se_d_rho)
    pairs_for_2se = (None if not d_rho else 8.0 / (d_rho ** 2))
    # **The divergence, caught on the first print of the corrected field.**
    # `8 / d^2` is a function of the **point estimate** of the difference, and
    # that point estimate has its own error. On the `recon` arm `d_rho` is
    # `9.7e-05` and the formula returns `859,039,612` pairs, `1.7` billion
    # trading days. **That number is arithmetic about a quantity
    # indistinguishable from zero, and it reads like a window requirement.**
    #
    # This is §32.6's family with a different denominator: a figure whose size
    # is a property of how close its divisor sits to zero. §32.6 kept the badly
    # normalised value under its own name rather than deleting it, and that is
    # what happens here: the number is stored, and the flag beside it says
    # whether the estimate it extrapolates from is itself bigger than its own
    # standard error. **The threshold is the quantity's own se, so it is not a
    # calibrated constant**, which is the same rule §24 and §33 cost this stage.
    d_rho_over_1se = (d_rho_in_se is not None and d_rho_in_se >= 1.0)
    ve_ratio = (None if not (head["calm"]["v_e"] and head["stress"]["v_e"])
                else head["stress"]["v_e"] / head["calm"]["v_e"])
    vc_ratio = (None if not (head["calm"]["v_c"] and head["stress"]["v_c"])
                else head["stress"]["v_c"] / head["calm"]["v_c"])
    cell, why = decomp_verdict(ve_ratio, d_rho)

    # §32.5: the common component with the fund itself left out, so that the
    # self-term of §31.1's cancellation is gone and only the residuals' genuine
    # cross-structure can survive.
    common_loo = {}
    for t in main:
        m = {}
        for d in all_days:
            vals = [series[j][d]["lam"] for j in main if j != t and d in series[j]]
            if vals:
                m[d] = sum(vals) / len(vals)
        common_loo[t] = m

    # --- §31.6(1): the identity, fund by fund, on the fund's own regime ----
    per_fund, breaches = {}, []
    for t in main:
        s = series[t]
        od = [(d, s[d]["order"]) for d in sorted(s)]
        pos = {d: i for i, (d, _) in enumerate(od)}
        m_t = len(od)
        cv_t = [common[d] for d, _ in od]
        ev_t = [s[d]["lam"] - common[d] for d, _ in od]
        rec = {}
        for lab, want in (("calm", True), ("stress", False)):
            def keep(d, w=want, ss=s):
                return ss[d]["calm"] is w
            lx, ly = lag1_pairs(od, lambda d, ss=s: ss[d]["lam"], keep)
            cx, cy = lag1_pairs(od, lambda d: common[d], keep)
            ex, ey = lag1_pairs(od, lambda d, ss=s: ss[d]["lam"] - common[d], keep)
            r_l, r_c, r_e = pearson(lx, ly), pearson(cx, cy), pearson(ex, ey)
            g_l, g_c, g_e = _geo_scale(lx, ly), _geo_scale(cx, cy), _geo_scale(ex, ey)
            resid = None
            if None not in (r_l, r_c, r_e, g_l, g_c, g_e) and (g_c + g_e) > 0:
                resid = abs(r_l * g_l - r_c * g_c - r_e * g_e) / (g_c + g_e)
            ev = _var([s[d]["lam"] - common[d] for d in s if s[d]["calm"] is want])
            ft = four_terms(lx, ly, cx, cy, ex, ey)
            rec[lab] = {"rho_lambda": r_l, "rho_c": r_c, "rho_e": r_e,
                        "w": (g_c / (g_c + g_e)) if (g_c is not None and g_e)
                             else None,
                        "v_e": ev, "pairs": len(lx),
                        "closure_residual": resid,
                        # §49.3: comparing the three lengths is a tautology,
                        # they come from one call pattern. Rebuilt from `od`.
                        "pairs_match": len(lx) == sum(
                            1 for (a_, oa_), (b_, ob_) in zip(od, od[1:])
                            if ob_ - oa_ == 1 and s[b_]["calm"] is want),
                        **ft}
            if resid is None or resid >= 0.01:
                breaches.append((t, lab, resid))
            if not rec[lab]["pairs_match"]:
                breaches.append((t, lab, "pair sets differ"))
            # **This one cannot fail unless the pairing is wrong**, because
            # `λ = c + e` is an identity and demeaning is linear. It is the
            # pair check with teeth, where §31.6(1) turned out to be a
            # measurement of the cross term rather than a check.
            ir, cl = ft["identity_residual"], ft["cov_lambda"]
            if ir is None or (cl and abs(ir / cl) > 1e-9):
                breaches.append((t, lab, f"four-term identity {ir}"))

            # §32.5's leave-one-out. The pair list is built once and every
            # component is read off it, so the three cannot land on different
            # days, which is §31.6(2)'s failure mode by construction.
            clo = common_loo[t]
            pl = [(a, b) for (a, oa), (b, ob) in zip(od, od[1:])
                  if ob - oa == 1 and s[b]["calm"] is want
                  and a in clo and b in clo]
            if len(pl) >= 2:
                ft_lo = four_terms(
                    [s[a]["lam"] for a, _ in pl], [s[b]["lam"] for _, b in pl],
                    [clo[a] for a, _ in pl], [clo[b] for _, b in pl],
                    [s[a]["lam"] - clo[a] for a, _ in pl],
                    [s[b]["lam"] - clo[b] for _, b in pl])
                ft_lo["pairs"] = len(pl)
                ilo = ft_lo["identity_residual"]
                if ilo is None or (ft_lo["cov_lambda"]
                                   and abs(ilo / ft_lo["cov_lambda"]) > 1e-9):
                    breaches.append((t, lab, f"loo four-term identity {ilo}"))
            else:
                ft_lo = {"pairs": len(pl), "cross": None, "cross_share": None}
            rec[lab]["loo"] = ft_lo

            # §33.3: the cross term against its own shift distribution.
            pidx = [(pos[a], pos[b]) for (a, oa), (b, ob) in zip(od, od[1:])
                    if ob - oa == 1 and s[b]["calm"] is want]
            null = shift_null(pidx, cv_t, ev_t, m_t)
            obs = (abs(ft["cross_share"])
                   if ft.get("cross_share") is not None else None)
            if null and obs is not None:
                rec[lab]["shift_percentile"] = (
                    sum(1 for x in null if x < obs) / len(null))
                rec[lab]["shift_null_median"] = null[len(null) // 2]
                rec[lab]["shift_null_p95"] = null[
                    min(len(null) - 1, int(SHIFT_NULL_QUANTILE * len(null)))]
                rec[lab]["shift_null_n"] = len(null)
            else:
                rec[lab].update({"shift_percentile": None,
                                 "shift_null_median": None,
                                 "shift_null_p95": None, "shift_null_n": 0})
        rec["v_e_ratio"] = ((rec["stress"]["v_e"] / rec["calm"]["v_e"])
                            if (rec["calm"]["v_e"] and rec["stress"]["v_e"])
                            else None)
        per_fund[t] = rec
        for lab, tag in (("calm", "calm"), ("stress", "strs")):
            q = rec[lab]
            print(f"{t if lab == 'calm' else '':6s} {tag}  "
                  f"rho_l={_f(q['rho_lambda'])} rho_c={_f(q['rho_c'])} "
                  f"rho_e={_f(q['rho_e'])} w={_f(q['w'])}  "
                  f"cross={_f(q['cross_share'])} "
                  f"loo={_f(q['loo'].get('cross_share'))}")
        print(f"{'':6s}       Ve s/c={_f(rec['v_e_ratio'])}")

    ve_down = [t for t in main if (per_fund[t]["v_e_ratio"] or 1.0) < 1.0]
    rc_up = [t for t in main
             if per_fund[t]["stress"]["rho_c"] is not None
             and per_fund[t]["calm"]["rho_c"] is not None
             and per_fund[t]["stress"]["rho_c"] > per_fund[t]["calm"]["rho_c"]]

    print(f"\n§31.4  market-stress days (>= {MARKET_STRESS_MAJORITY} of "
          f"{len(main)} in their own stress): {n_ms}/{len(all_days)}")
    for lab in ("calm", "stress"):
        h = head[lab]
        e3 = lambda x: ("%.3e" % x) if isinstance(x, float) else "None"
        print(f"§31.4  {lab:6s} rho_c={_f(h['rho_c'])} (se {_f(h['se_rho_c'])}, "
              f"{h['pairs_c']} pairs)  V_c={e3(h['v_c'])}  V_e={e3(h['v_e'])}  "
              f"w={_f(h['w'])}")
    print(f"§31.5  V_e stress/calm = {_f(ve_ratio)}   "
          f"V_c stress/calm = {_f(vc_ratio)}   "
          f"d(rho_c) = {_f(d_rho)}"
          + (f" = {d_rho_in_se:.2f} se(diff), se(diff)={se_d_rho:.4f}"
             if d_rho_in_se is not None else ""))
    if pairs_for_2se:
        print(f"§34.6  two se on the **difference** needs about "
              f"{pairs_for_2se:.0f} pairs per regime, that is about "
              f"{2 * pairs_for_2se:.0f} trading days. The window is "
              f"{len(all_days)}.")
        if not d_rho_over_1se:
            print(f"       **and that figure extrapolates from a point "
                  f"estimate of {d_rho:+.5f}, which is {d_rho_in_se:.2f} of "
                  f"its own standard error.** It inherits that uncertainty. "
                  f"Read it as an order of magnitude, never as a target, and "
                  f"**not at all** when the difference is near zero: the "
                  f"formula diverges there and its size is a statement about "
                  f"the divisor.")
    tot_c = sum(v["candidates"] for k, v in LAMBDA_DROPS.items() if k in main)
    tot_k = sum(v["kept"] for k, v in LAMBDA_DROPS.items() if k in main)
    agg = {}
    for k, v in LAMBDA_DROPS.items():
        if k in main:
            for kk, vv in v.items():
                if kk not in ("candidates", "kept"):
                    agg[kk] = agg.get(kk, 0) + vv
    print(f"§49.5  main-arm candidates {tot_c}, kept {tot_k}, dropped "
          f"{tot_c - tot_k}: "
          + " ".join(f"{k}={v}" for k, v in sorted(agg.items()) if v))
    print(f"§31.5  VERDICT: {cell}")
    print(f"        {why}")
    cross, all_abs, loo_abs, sign_max = {}, [], [], 0
    for lab in ("calm", "stress"):
        sh = [per_fund[t][lab]["cross_share"] for t in main
              if per_fund[t][lab]["cross_share"] is not None]
        lsh = [per_fund[t][lab]["loo"]["cross_share"] for t in main
               if per_fund[t][lab]["loo"].get("cross_share") is not None]
        lx_ = [per_fund[t][lab]["loo"]["cross"] for t in main
               if per_fund[t][lab]["loo"].get("cross") is not None]
        pos = sum(1 for x in lx_ if x > 0)
        sign_max = max(sign_max, pos, len(lx_) - pos)
        all_abs += [abs(x) for x in sh]
        loo_abs += [abs(x) for x in lsh]
        srt = sorted(abs(x) for x in sh)
        lsrt = sorted(abs(x) for x in lsh)
        cross[lab] = {
            "median_abs_cross_share": srt[len(srt) // 2] if srt else None,
            "max_abs_cross_share": srt[-1] if srt else None,
            "loo_median_abs_cross_share": lsrt[len(lsrt) // 2] if lsrt else None,
            "loo_max_abs_cross_share": lsrt[-1] if lsrt else None,
            "cross_positive": sum(1 for t in main
                                  if (per_fund[t][lab]["cross"] or 0) > 0),
            "loo_cross_positive": pos, "loo_n": len(lx_),
            "c_leads_e_positive": sum(1 for t in main
                                      if (per_fund[t][lab]["cov_c_leads_e"] or 0) > 0),
            "e_leads_c_positive": sum(1 for t in main
                                      if (per_fund[t][lab]["cov_e_leads_c"] or 0) > 0)}
        z = cross[lab]
        print(f"§32.3  {lab:6s} |cross|/(G_c+G_e) median={_f(z['median_abs_cross_share'])} "
              f"max={_f(z['max_abs_cross_share'])}   sign +{z['cross_positive']}"
              f"/-{len(main) - z['cross_positive']}   "
              f"(c->e +{z['c_leads_e_positive']}, e->c +{z['e_leads_c_positive']})")
        print(f"§32.5  {lab:6s} leave-one-out     median={_f(z['loo_median_abs_cross_share'])} "
              f"max={_f(z['loo_max_abs_cross_share'])}   sign +{z['loo_cross_positive']}"
              f"/-{z['loo_n'] - z['loo_cross_positive']}")

    med_all = (sorted(all_abs)[len(all_abs) // 2]) if all_abs else None
    med_loo = (sorted(loo_abs)[len(loo_abs) // 2]) if loo_abs else None
    cells_clear = sum(1 for x in loo_abs if x < 0.01)
    loo_cell, loo_why = loo_verdict(med_all, med_loo, cells_clear, sign_max)
    print(f"§32.5  pooled median |cross| {_f(med_all)} -> {_f(med_loo)} "
          f"leave-one-out, cells under 0.01: {cells_clear}/{len(loo_abs)}, "
          f"largest same-sign run {sign_max}/{len(main)}")
    print(f"§32.5  VERDICT: {loo_cell}")
    print(f"        {loo_why}")
    # --- §33: the cross term against its own noise floor -------------------
    cells = [(t, lab) for t in main for lab in ("calm", "stress")
             if per_fund[t][lab]["shift_percentile"] is not None]
    above = [(t, lab) for t, lab in cells
             if per_fund[t][lab]["shift_percentile"] >= SHIFT_NULL_QUANTILE]
    p95s = sorted(per_fund[t][lab]["shift_null_p95"] for t, lab in cells)
    nmeds = sorted(per_fund[t][lab]["shift_null_median"] for t, lab in cells)
    med_p95 = p95s[len(p95s) // 2] if p95s else None
    med_null = nmeds[len(nmeds) // 2] if nmeds else None
    s_cell, s_why = shift_verdict(len(above), len(cells))
    print(f"§33.3  shift null: {per_fund[main[0]]['calm']['shift_null_n']} "
          f"shifts per cell, deterministic")
    print(f"§33.3  the floor itself: median null |cross| {_f(med_null)}, "
          f"median null 95th percentile {_f(med_p95)}")
    print(f"§33.5  **§31.6(1) used 0.01. The measured floor is "
          f"{_f(med_p95)}.**")
    print(f"§33.4  cells at or above their own 95th percentile: "
          f"{len(above)}/{len(cells)}  (chance predicts about "
          f"{0.05 * len(cells):.1f})"
          + (f"  {sorted(t + ':' + l for t, l in above)}" if above else ""))
    # **§49.2.** A verdict table whose rows cannot fire on this data is not a
    # verdict table. Reachability is reported next to the verdict, not left for
    # an audit to discover afterwards.
    unreach = []
    if med_p95 is not None and 0.01 < med_p95 / 2:
        unreach.append(f"§32.5 `artefact_confirmed` needs {LOO_CLEAR_CELLS}/"
                       f"{len(loo_abs)} cells under 0.01 while the measured "
                       f"floor is {med_p95:.4f}; only {cells_clear} are")
    if LOO_SIGN_MAJORITY > len(main):
        unreach.append(f"§32.5 `propagation` needs {LOO_SIGN_MAJORITY} of "
                       f"{len(main)} funds")
    if SHIFT_ROW2_MIN > len(cells):
        unreach.append(f"§33.4 `real` needs {SHIFT_ROW2_MIN} of "
                       f"{len(cells)} cells")
    print(f"§33.4  VERDICT: {s_cell}")
    if unreach:
        print("§49.2  **rows that could not have fired on this data**:")
        for u in unreach:
            print(f"        {u}")
    else:
        print("§49.2  every registered verdict row was reachable on this data")
    print(f"        {s_why}")
    print("       (the cross term is measured, not tolerated. §31.1 assumed it "
          "was zero and the first run falsified that.)")
    print(f"§31.6  closure breaches: {len(breaches)}"
          + (f"  first {breaches[0]}" if breaches else "  (all funds close)"))
    print(f"        per-fund agreement: V_e falls in {len(ve_down)}/{len(main)}, "
          f"rho_c rises in {len(rc_up)}/{len(main)}")
    print("        (§20.2: the count is close to one observation on this arm, "
          "day-effect share 0.703. The level is what carries.)")

    rec = {"stage": "B9 §31 variance decomposition", "diagnostic_only": False,
           "rule": "§31.5's four cells, by sign, registered before the run.",
           "market_stress_majority": MARKET_STRESS_MAJORITY,
           "market_stress_days": n_ms, "days": len(all_days),
           "headline": head, "v_e_ratio": ve_ratio, "v_c_ratio": vc_ratio,
           "d_rho_c": d_rho, "se_d_rho_c": se_d_rho,
           "d_rho_c_in_se": d_rho_in_se,
           "d_rho_c_se_note": ("se of the difference, sqrt(se_calm^2 + "
                               "se_stress^2), because the two regimes "
                               "partition the pairs. Dividing by one end's se "
                               "understates the spread by about sqrt(2)."),
           "pairs_per_regime_for_2se": pairs_for_2se,
           "d_rho_c_over_own_se": d_rho_over_1se,
           "pairs_for_2se_note": ("8 / d^2, a function of the point estimate "
                                  "of the difference. It inherits that "
                                  "estimate's error and diverges as the "
                                  "difference goes to zero, so it is only "
                                  "readable when d_rho_c_over_own_se is true, "
                                  "and then only as an order of magnitude."),
           "trading_days_for_2se": (None if not pairs_for_2se
                                    else 2.0 * pairs_for_2se),
           "verdict": cell, "verdict_note": why,
           "closure_breaches": [list(b) for b in breaches],
           "cross_term": cross,
           "loo_verdict": loo_cell, "loo_verdict_note": loo_why,
           "loo_median_abs_cross_share": med_loo,
           "median_abs_cross_share": med_all,
           "loo_cells_under_tolerance": cells_clear,
           "loo_largest_same_sign_run": sign_max,
           "shift_verdict": s_cell, "shift_verdict_note": s_why,
           "shift_cells_above_p95": [list(x) for x in above],
           "shift_cells": len(cells),
           "shift_median_null_p95": med_p95,
           "shift_median_null_median": med_null,
           "unreachable_rows": unreach,
           "lambda_drops": {k: v for k, v in LAMBDA_DROPS.items() if k in main},
           "per_fund_v_e_falls": sorted(ve_down),
           "per_fund_rho_c_rises": sorted(rc_up),
           "funds": per_fund}
    DECOMP_OUT_P = out_path(DECOMP_OUT)
    DECOMP_OUT_P.parent.mkdir(parents=True, exist_ok=True)
    DECOMP_OUT_P.write_text(json.dumps(rec, sort_keys=True, ensure_ascii=False,
                                     indent=2) + "\n", encoding="utf-8",
                          newline="\n")
    print(f"wrote {DECOMP_OUT_P.relative_to(ROOT)}")
    print("§31.7: this does not close §7's fourth link. The ordinary account "
          "predicts the same on every quantity above.")
    return rec


DAYS_OUT = ROOT / "results" / "b9_days.json"


def dump_days(tickers: list[str]) -> dict:
    """§36 step 0. Writes the connect window, the NAV and the disclosed premium
    per fund-day, and §6's calm flag, so that the retrieval tool needs no
    workbook parser of its own.

    **Exported through `lambda_series`, not rebuilt**, so the sample §36 compares
    on is the sample §25 and §31 measured on, filter for filter."""
    spreads = read_spread_table()
    series = {t: lambda_series(t, spreads) for t in (tickers or SAMPLE)}
    series = {t: s for t, s in series.items() if s}
    days = sorted({d for s in series.values() for d in s})
    funds = {t: {d: {"nav": v["nav"], "prem": v["prem"], "calm": v["calm"],
                     "disclosed_price": v["nav"] * (1.0 + v["prem"])}
                 for d, v in s.items()}
             for t, s in series.items()}
    rec = {"stage": "B9 §36 step 0, the day list and its ground truth",
           "diagnostic_only": True,
           "rule": "§24.1: the disclosed price is NAV x (1 + premium) and lands "
                   "on the half-cent grid, so it is known to the digit.",
           "days": days, "n_days": len(days),
           "first": days[0] if days else "", "last": days[-1] if days else "",
           "funds": funds,
           "fund_day_counts": {t: len(s) for t, s in funds.items()}}
    DAYS_OUT_P = out_path(DAYS_OUT)
    DAYS_OUT_P.parent.mkdir(parents=True, exist_ok=True)
    DAYS_OUT_P.write_text(json.dumps(rec, sort_keys=True, ensure_ascii=False,
                                   indent=1) + "\n", encoding="utf-8",
                        newline="\n")
    print(f"{len(days)} days, {days[0] if days else '-'} to "
          f"{days[-1] if days else '-'}, {len(funds)} funds")
    for t in sorted(funds):
        print(f"  {t:6s} {len(funds[t]):4d} fund-days")
    print(f"wrote {DAYS_OUT_P.relative_to(ROOT)}")
    return rec


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--selftest", action="store_true", help="no data touched")
    ap.add_argument("--depth", action="store_true",
                    help="print the join and the column spreads; estimate nothing")
    ap.add_argument("--gate", action="store_true",
                    help="B9-0 on real fund-days, §6. Fee passed as NaN.")
    ap.add_argument("--a1", action="store_true",
                    help="B9-A-1 on §17's rectangle; needs B9-0 to have passed")
    ap.add_argument("--a2", action="store_true",
                    help="B9-A-2's gradient, §18.6. Immune to §17's quantum; "
                         "counts at the f=0 end only")
    ap.add_argument("--gate-speed", action="store_true",
                    help="§26: A3 §6.4d's wall mapped to the primary market. "
                         "The two accounts differ in sign.")
    ap.add_argument("--measured", action="store_true",
                    help="§24: A-1 and A-2 against the measurement floor F_m")
    ap.add_argument("--grid", action="store_true",
                    help="§24.1: does every reconstructed close land on the "
                         "half-cent midpoint grid, and §24.3's tick comparison")
    ap.add_argument("--pi-check", action="store_true",
                    help="§22.4: compute π once to check §22.2's derivation. "
                         "Not a test of holonomy.")
    ap.add_argument("--disc", action="store_true",
                    help="§19.5's D1 and D2, plus §19.4's independence check")
    ap.add_argument("--dump-days", action="store_true",
                    help="§36 step 0: export the connect window, NAV, disclosed "
                         "premium and calm flag for the retrieval tool")
    ap.add_argument("--decomp", action="store_true",
                    help="§31: measures rho_c and w, which D1 and D1b never "
                         "did, and files §21.2 into §31.5's four cells")
    ap.add_argument("--inputs", action="store_true",
                    help="print √N, f_max and the stress split; compute no λ")
    ap.add_argument("--f1", action="store_true",
                    help="§5's traversability audit; scans the threshold, "
                         "adopts none. Refuses to run unless B9-0 passed.")
    ap.add_argument("--recon", choices=["off", "recon", "control"],
                    default="off",
                    help="§40.3: `recon` substitutes the reconstructed premium, "
                         "`control` keeps the disclosed one on the same "
                         "fund-days. Compare those two, never against a full run.")
    ap.add_argument("--only", action="append", default=[], metavar="TICKER")
    args = ap.parse_args()

    only = [t.upper() for t in args.only]
    # **§49.5(5).** `--recon` only reaches stages that read `lambda_series`.
    # The first version printed the mode banner for every stage and then ran
    # `--a1` on the full disclosed sample, writing the ordinary results file
    # with no marker. It now refuses instead of misleading.
    RECON_AWARE = ("disc", "decomp", "measured", "pi_check", "gate_speed",
                   "dump_days")
    if args.recon != "off":
        asked = [k for k in ("depth", "gate", "a1", "a2", "f1", "inputs",
                             "grid") if getattr(args, k, False)]
        if asked:
            print(f"**Refused.** --recon does not reach {', '.join(asked)}: "
                  f"those stages rebuild the premium themselves and never call "
                  f"lambda_series. §49.5(5). Recon-aware stages: "
                  f"{', '.join(RECON_AWARE)}.")
            return 2
    load_recon(args.recon)
    if args.selftest:
        return selftest()
    if args.depth:
        depth(only)
        return 0
    if args.gate:
        return 0 if gate(only).get("passed") else 1
    if args.f1:
        f1_audit(only)
        return 0
    if args.inputs:
        inputs(only)
        return 0
    if args.a1:
        a1(only)
        return 0
    if args.a2:
        a2(only)
        return 0
    if args.disc:
        discriminators(only)
        return 0
    if args.dump_days:
        dump_days(only)
        return 0
    if args.decomp:
        decomp(only)
        return 0
    if args.gate_speed:
        gate_speed(only)
        return 0
    if args.measured:
        measured(only)
        return 0
    if args.grid:
        midpoint_grid_check(only)
        return 0
    if args.pi_check:
        pi_check(only)
        return 0
    ap.print_help()
    print("\nOrder is §10 step 3: --depth, then --gate, then --f1. "
          "B9-A-1 comes after all three and after `f` is read (§14.6).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
