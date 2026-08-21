#!/usr/bin/env python3
"""B9-A step one: what is retrievable, per field, per fund.

Registered in ``docs/b9_zero_holonomy.md`` §10, whose first step reads in full:
confirm for a small sample of funds that daily NAV, daily close, shares
outstanding, median bid-ask spread and a determinate creation fee are all
retrievable without a terminal; report per field, per fund; **write the result
down whichever way it comes out.**

This file is that step and nothing else. It estimates nothing, it constructs no
``ω``, it computes no ``λ`` and no ``π``, and it reads no prediction of B9 or of
any other stage. It does not depend on B8 or B10: see ``PROJECT_PLAN.md`` §25.

**The load-bearing question here is granularity, not presence.** Rule 6c-11
requires the median bid-ask spread on the fund's website, and a website figure is
normally *the most recent 30 calendar days*, that is a snapshot. §12.1 of the
pre-registration requires ``√N`` computed on **exactly the fund-days that enter
the zero**, after the F1 filter. A snapshot cannot do that. So the probe's
headline output is a per-field ``granularity`` of ``history`` / ``snapshot`` /
``none``, and the verdict states whether §12.1 is satisfiable on free data. If it
is not, that is the finding of this step and it is reported, not worked around.

Three things this probe is built **not** to do, each one a rule it would
otherwise break:

* **It never infers absence from a failed fetch.** A 403 is recorded as a 403.
  "The issuer does not publish it" and "we did not get it" are different objects
  and are kept in different status codes.
* **It never selects a download link.** ``--discover`` enumerates every candidate
  on the fund's own page and prints them with the keyword that matched. A human
  picks. Guessing an endpoint would produce a probe that reports "unavailable"
  for a field that is published, which is the worst outcome this step can have.
* **It never promotes a truncated payload.** the project's engineering rule 6. Bytes land in
  ``<name>.partial``, the declared length is checked, and only then is the final
  name written. A ``.partial`` is kept, never deleted, and never read as data.

Nothing is ever deleted (the project's engineering rule 5). Re-running is free: a payload
whose sidecar hash still matches is not fetched again.

Usage::

    python experiments/b9a_availability.py --selftest
    python experiments/b9a_availability.py --discover --only IVV --only EFA
    python experiments/b9a_availability.py --discover
    python experiments/b9a_availability.py --report

``--probe`` is deliberately not implemented in this version. It cannot be, until
``--discover`` has returned the real link set and a human has filled
``data/raw/b9/endpoints.json``.
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import io
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from urllib.parse import urljoin, urlsplit

ROOT = Path(__file__).resolve().parents[1]
CACHE = ROOT / "data" / "raw" / "b9"
OUT_JSON = ROOT / "results" / "b9a_availability.json"
OUT_MD = ROOT / "results" / "b9a_availability.md"

#: A human fills this after reading --discover output. Absent is the normal
#: state on the first run and is not an error.
ENDPOINTS = CACHE / "endpoints.json"
#: A human corrects a wrong landing page here rather than in this file, so that
#: the registered fund table below stays as it was written before retrieval.
LANDING_OVERRIDES = CACHE / "landing_overrides.json"

USER_AGENT = "b9a-availability-probe/1 (academic replication; contact via repo)"
TIMEOUT_S = 30
SLEEP_S = 1.5

# ---------------------------------------------------------------------------
# The five fields. Fixed here, before retrieval, with what each one is for.
# ---------------------------------------------------------------------------

FIELDS = [
    {
        "id": "nav_history",
        "label": "daily NAV",
        "needed_by": "§2's λ = log(1+premium) − log(1+f); premium = (P−NAV)/NAV",
        "granularity_required": "history",
    },
    {
        "id": "close_history",
        "label": "daily closing price P",
        "needed_by": "§2, the same expression",
        "granularity_required": "history",
    },
    {
        "id": "shares_outstanding_history",
        "label": "shares outstanding, daily",
        "needed_by": "§5 F1: primary-market activity is inferred from its path, "
                     "and fund-days with none within ±5 days do not enter the zero",
        "granularity_required": "history",
    },
    {
        "id": "median_bid_ask_spread",
        "label": "median bid-ask spread",
        "needed_by": "§4's √N, and §12.1 requires it on exactly the F1-cleared "
                     "fund-days, so a 30-day website snapshot is not enough",
        "granularity_required": "history",
    },
    {
        "id": "creation_fee",
        "label": "creation transaction fee, a determinate number",
        "needed_by": "§2's f, and §7 drops a fund whose fee is not determinate, "
                     "counted separately from a missing price",
        "granularity_required": "snapshot",
    },
]

# ---------------------------------------------------------------------------
# The fund sample. Enumerated before retrieval (discipline 12), with the arm
# each fund belongs to fixed here so that it is not assigned after seeing which
# ones were easy to get.
#
# §7's filters: US-listed, primary listing only. Leveraged, inverse and
# single-stock funds are excluded, declared here rather than after seeing them.
#
# §6.1 splits the sample: the main arm is funds whose underlying trades
# contemporaneously with the fund, that is US-listed funds on US-listed
# equities. International and fixed-income funds are **reported beside** it and
# are not in it.
#
# `structure` is unverified on purpose. §7 says "Rule 6c-11 funds", and 6c-11 covers
# ETFs organised as open-end funds. SPY, QQQ and DIA are unit investment trusts
# operating under exemptive relief that predates the rule, so they are neither
# arm here and are carried only as a labelled comparison. **Which regime a fund
# is under is earned from the fund's own documents, not from this comment**,
# which is C0b's rule applied to a legal fact instead of a column layout.
#
# `landing` is a candidate. Several of these were written from memory and may
# 404. That is why --discover reports the HTTP status per fund and why
# landing_overrides.json exists: a wrong URL here is a correctable input, not a
# finding about the issuer.
# ---------------------------------------------------------------------------

FUNDS = [
    # === main arm, per §13.5: the Select Sector SPDR family ==============
    # Open-end, US-listed, on US-listed equities, so inside §6.1's main arm.
    # Eleven sectors, one issuer, one publishing practice, which is what makes
    # the cross-fund dispersion of √N (§13.3 item 2) a comparison between funds
    # rather than between issuers' disclosure habits.
    {"ticker": "XLB", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US materials equity", "landing": ""},
    {"ticker": "XLC", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US communication services equity", "landing": ""},
    {"ticker": "XLE", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US energy equity", "landing": ""},
    {"ticker": "XLI", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US industrials equity", "landing": ""},
    {"ticker": "XLK", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US technology equity", "landing": ""},
    {"ticker": "XLP", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US consumer staples equity", "landing": ""},
    {"ticker": "XLRE", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US real estate equity", "landing": ""},
    {"ticker": "XLU", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US utilities equity", "landing": ""},
    {"ticker": "XLV", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US health care equity", "landing": ""},
    {"ticker": "XLY", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US consumer discretionary equity", "landing": ""},

    # === comparison arm, §6.1: stale NAV expected, reported beside =======
    {"ticker": "SPDW", "issuer": "SSGA", "arm": "comparison",
     "structure": "open-end, unverified", "underlying": "developed ex-US equity", "landing": ""},
    {"ticker": "SPEM", "issuer": "SSGA", "arm": "comparison",
     "structure": "open-end, unverified", "underlying": "emerging markets equity", "landing": ""},
    {"ticker": "SPAB", "issuer": "SSGA", "arm": "comparison",
     "structure": "open-end, unverified", "underlying": "aggregate bonds", "landing": ""},
    {"ticker": "JNK", "issuer": "SSGA", "arm": "comparison",
     "structure": "open-end, unverified", "underlying": "high yield corporate bonds", "landing": ""},

    # === retired by availability, kept because nothing is deleted ========
    # §13.5: retrieval works at SSGA and does not at iShares. These rows stay
    # declared so that the sample this stage started from is visible, and so a
    # later run can retry them with one flag. **They are in no arm and enter no
    # reading.** Their soft-404 evidence is in
    # the B9 iShares availability measurement, outside this repository.
    {"ticker": "IVV", "issuer": "iShares", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "US large-cap equity",
     "landing": "https://www.ishares.com/us/products/239726/ishares-core-sp-500-etf"},
    {"ticker": "IJR", "issuer": "iShares", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "US small-cap equity",
     "landing": "https://www.ishares.com/us/products/239774/ishares-core-sp-smallcap-etf"},
    {"ticker": "IWM", "issuer": "iShares", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "US small-cap equity",
     "landing": "https://www.ishares.com/us/products/239710/ishares-russell-2000-etf"},
    {"ticker": "VOO", "issuer": "Vanguard", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "US large-cap equity",
     "landing": "https://investor.vanguard.com/investment-products/etfs/profile/voo"},
    {"ticker": "SCHX", "issuer": "Schwab", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "US large-cap equity",
     "landing": "https://www.schwabassetmanagement.com/products/schx"},
    {"ticker": "RSP", "issuer": "Invesco", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "US large-cap equity, equal weight",
     "landing": "https://www.invesco.com/us/financial-products/etfs/product-detail"
                "?audienceType=Investor&ticker=RSP"},
    {"ticker": "XLF", "issuer": "SSGA", "arm": "contemporaneous",
     "structure": "open-end, unverified", "underlying": "US financials equity",
     "landing": "https://www.sectorspdrs.com/mainfund/xlf"},

    # --- comparison arm: §6.1, stale NAV expected, reported beside -------
    {"ticker": "EFA", "issuer": "iShares", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "developed ex-US equity",
     "landing": "https://www.ishares.com/us/products/239623/ishares-msci-eafe-etf"},
    {"ticker": "EWJ", "issuer": "iShares", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "Japan equity, non-overlapping session",
     "landing": "https://www.ishares.com/us/products/239665/ishares-msci-japan-etf"},
    {"ticker": "VEA", "issuer": "Vanguard", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "developed ex-US equity",
     "landing": "https://investor.vanguard.com/investment-products/etfs/profile/vea"},
    {"ticker": "LQD", "issuer": "iShares", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "IG corporate bonds",
     "landing": "https://www.ishares.com/us/products/239566/"
                "ishares-iboxx-investment-grade-corporate-bond-etf"},
    {"ticker": "VCIT", "issuer": "Vanguard", "arm": "retired_availability",
     "structure": "open-end, unverified", "underlying": "IG corporate bonds, intermediate",
     "landing": "https://investor.vanguard.com/investment-products/etfs/profile/vcit"},

    # --- neither arm: carried only to record the regime question ---------
    {"ticker": "SPY", "issuer": "SSGA", "arm": "neither",
     "structure": "UIT, unverified; predates the 6c-11 exemptive order",
     "underlying": "US large-cap equity",
     "landing": "https://www.ssga.com/us/en/intermediary/etfs/spdr-sp-500-etf-trust-spy"},
]

EXCLUDED_BY_SECTION_7 = [
    "leveraged funds", "inverse funds", "single-stock funds",
    "secondary listings of a fund whose primary listing is elsewhere",
]

#: Declared third-party fallback for the closing price only. the project's engineering rule 9
#: permits third-party data that is real and usable; provenance is kept separate
#: from issuer-published data so the two never merge silently.
THIRD_PARTY_CLOSE = "https://stooq.com/q/d/l/?s={t}.us&i=d"

# ---------------------------------------------------------------------------
# Link classification. Keyword to field. Enumerates, never selects.
# ---------------------------------------------------------------------------

LINK_KEYWORDS = {
    "nav_history": ["nav", "net-asset-value", "netassetvalue"],
    "close_history": ["market-price", "marketprice", "closing", "price-history"],
    "shares_outstanding_history": ["shares-outstanding", "sharesoutstanding",
                                   "shares_out", "fund-size"],
    "median_bid_ask_spread": ["bid-ask", "bidask", "bid_ask", "spread"],
    "creation_fee": ["sai", "statement-of-additional", "statementofadditional",
                     "prospectus", "creation", "transaction-fee"],
    "premium_discount": ["premium", "discount", "prem-disc", "premdisc"],
    "holdings": ["holdings", "basket", "constituent"],
}

DOWNLOADABLE = (".csv", ".xls", ".xlsx", ".txt", ".json", ".pdf")

#: Static assets, dropped before classification and **counted**. Without this,
#: `navigation-CX4YlZ4h.css` reaches `nav_history` and `holdings-v3-*.css`
#: reaches `holdings`. Both did, on the first IVV run.
ASSET_EXT = (".css", ".js", ".mjs", ".map", ".woff", ".woff2", ".ttf", ".eot",
             ".svg", ".png", ".jpg", ".jpeg", ".gif", ".ico", ".webp")

HREF_RE = re.compile(
    r"""(?:href|src|data-link|data-url|data-file)\s*=\s*["']([^"']+)["']""",
    re.IGNORECASE)

#: **The data an issuer serves is usually behind a URL that never appears in an
#: href.** iShares serves NAV and premium/discount history from
#: `<product>/1467271812596.ajax?fileType=csv&fileName=<TICKER>_NAV&dataType=fund`,
#: emitted inside script or JSON. Enumerating anchors only made the first IVV
#: run report nineteen candidates of which not one was a data endpoint. This
#: scans the whole page text.
ENDPOINT_RE = re.compile(
    r"""((?:https?://[^\s"'<>()\\]+|/[^\s"'<>()\\]*)"""
    r"""(?:\.ajax\?|\.csv\b|\.xlsx?\b|\.json\b)[^\s"'<>()\\]*)""",
    re.IGNORECASE)


def _token_hit(low: str, kw: str) -> bool:
    """Keyword at a token boundary. `nav` hits `IVV_NAV` and `nav.csv` and does
    **not** hit `navigation`, which is the whole point of this function."""
    return re.search(rf"(?<![a-z0-9]){re.escape(kw)}(?![a-z0-9])", low) is not None


def classify_link(href: str) -> list[str]:
    """Every field this link might serve. Sorted, possibly empty, never one
    'best' guess: a link that matches two fields is reported under both."""
    low = href.lower().split("#")[0]
    path = low.split("?")[0]
    if path.endswith(ASSET_EXT):
        return []
    hits = {f for f, kws in LINK_KEYWORDS.items()
            if any(_token_hit(low, k) for k in kws)}
    if not hits and (path.endswith(DOWNLOADABLE) or ".ajax" in low):
        hits.add("unclassified_download")
    return sorted(hits)


def extract_links(page: str, base: str) -> dict:
    """Candidates, plus the counts of what was thrown away. **Nothing is
    dropped silently**: a link that vanished into `dropped_assets` is a link
    this probe chose not to report, and that number is printed."""
    page = unescape(page)
    raw = {m.strip() for m in HREF_RE.findall(page)}
    raw |= {m.strip() for m in ENDPOINT_RE.findall(page)}

    seen: dict[str, dict] = {}
    dropped_assets = 0
    for href in sorted(raw):
        if not href or href.startswith(("#", "javascript:", "mailto:", "data:")):
            continue
        absolute = urljoin(base, href)
        if absolute.lower().split("?")[0].endswith(ASSET_EXT):
            dropped_assets += 1
            continue
        fields = classify_link(absolute)
        if not fields:
            continue
        seen.setdefault(absolute, {"url": absolute, "fields": fields})
    return {
        "candidates": sorted(seen.values(), key=lambda d: d["url"]),
        "urls_seen": len(raw),
        "dropped_assets": dropped_assets,
    }


# ---------------------------------------------------------------------------
# Asset scan. The step the first two runs proved is necessary.
#
# IVV and EFA both returned 239 URLs and **not one endpoint**. Both pages are
# client-rendered, so the data URL is assembled by script at runtime and is
# absent from the initial HTML. Guessing the endpoint from memory is what §10's
# "write the result down whichever way it comes out" forbids in practice: a
# wrong guess reports a published field as unavailable. So the endpoint is
# **read out of the issuer's own bundle** instead.
#
# This finds nothing on a server-rendered site, which is the correct outcome
# there, and it is recorded as nothing found rather than as a failure.
# ---------------------------------------------------------------------------

ASSET_SCAN_PATTERNS = {
    "ajax_endpoint": re.compile(r"""[\w./%-]*\.ajax\b[^\s"'`<>]*"""),
    "file_name_param": re.compile(r"""fileName=[\w%.+-]*"""),
    "data_type_param": re.compile(r"""dataType=[\w%.+-]*"""),
    "file_type_param": re.compile(r"""fileType=[\w%.+-]*"""),
    "csv_path": re.compile(r"""[\w./%-]{3,}\.csv\b"""),
    "product_path": re.compile(r"""/us/products/[\w./%-]+"""),
}

SCRIPT_SRC_RE = re.compile(r"""<script[^>]+src\s*=\s*["']([^"']+)["']""", re.I)


def same_origin(url: str, base: str) -> bool:
    return urlsplit(url).netloc.lower() == urlsplit(base).netloc.lower()


def extract_assets(page: str, base: str) -> list[str]:
    """Same-origin JavaScript referenced by the page. Sorted, deduplicated."""
    page = unescape(page)
    out: set[str] = set()
    for m in list(HREF_RE.findall(page)) + list(SCRIPT_SRC_RE.findall(page)):
        u = urljoin(base, m.strip())
        if u.lower().split("?")[0].endswith((".js", ".mjs")) and same_origin(u, base):
            out.add(u)
    return sorted(out)


def scan_text_for_endpoints(text: str) -> dict[str, list[str]]:
    """Every endpoint-shaped string, by pattern. Sorted, deduplicated."""
    return {
        name: sorted({m for m in rx.findall(text) if len(m) > 4})
        for name, rx in sorted(ASSET_SCAN_PATTERNS.items())
    }


#: **Four formats, and the window matters as much as the patterns.** The first
#: version read ISO and US dates out of the leading 2,000 bytes and called all
#: four iShares payloads a snapshot. Those files open with a block of fund
#: metadata, so the first dated row is well past 2,000 bytes, and the label was
#: an artefact of the window rather than a property of the file.
DATE_RE = re.compile(
    r"""(\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{4}"""
    r"""|[A-Z][a-z]{2}\s\d{1,2},\s?\d{4}|\d{1,2}-[A-Z][a-z]{2}-\d{2,4})""")


def granularity_of_payload(text: str) -> str:
    """history if two or more distinct dates appear **anywhere**, snapshot if a
    figure but no second date, none if neither. A triage label a human checks,
    reported next to the raw sample."""
    if len(set(DATE_RE.findall(text))) >= 2:
        return "history"
    if re.search(r"\d", text):
        return "snapshot"
    return "none"


def payload_profile(text: str) -> dict:
    """What a downloaded file actually is. Reads the whole of it."""
    lines = text.splitlines()
    dates = DATE_RE.findall(text)
    # The header is the nearest comma-bearing line above the first dated row
    # that is not itself dated. Counting commas to find it fails on a two-column
    # file, which is how the first version missed `Date,Close`.
    header_idx, header = None, ""
    first_dated = next((i for i, ln in enumerate(lines) if DATE_RE.search(ln)), None)
    if first_dated is not None:
        for i in range(first_dated - 1, -1, -1):
            if "," in lines[i] and not DATE_RE.search(lines[i]):
                header_idx, header = i, lines[i]
                break
    return {
        "lines": len(lines),
        "distinct_dates": len(set(dates)),
        "date_first_in_file": dates[0] if dates else "",
        "date_last_in_file": dates[-1] if dates else "",
        "header_line_index": header_idx,
        "header_line": header[:300],
        "granularity": granularity_of_payload(text),
    }


# ---------------------------------------------------------------------------
# Cache with resume and truncation detection. the project's engineering rule 6.
# ---------------------------------------------------------------------------


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def cache_paths(ticker: str, name: str) -> tuple[Path, Path, Path]:
    d = CACHE / ticker
    return d / name, d / f"{name}.partial", d / f"{name}.meta.json"


def cached_ok(final: Path, meta: Path) -> bool:
    """True only if the payload is on disk **and** its hash still matches the
    sidecar. A byte that changed under us is not silently read."""
    if not (final.is_file() and meta.is_file()):
        return False
    try:
        rec = json.loads(meta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    return rec.get("sha256") == _sha256(final.read_bytes())


def fetch(url: str, ticker: str, name: str, force: bool = False) -> dict:
    """One request. Returns the sidecar record whatever happens.

    Never raises on a network condition, never deletes, never promotes a payload
    whose length disagrees with the declared Content-Length.
    """
    final, partial, meta = cache_paths(ticker, name)
    if cached_ok(final, meta) and not force:
        rec = json.loads(meta.read_text(encoding="utf-8"))
        rec["from_cache"] = True
        return rec

    final.parent.mkdir(parents=True, exist_ok=True)
    rec: dict = {
        "url": url, "name": name, "from_cache": False,
        "retrieved_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT_S) as resp:
            declared = resp.headers.get("Content-Length")
            body = resp.read()
            rec["status"] = f"http_{resp.status}"
            rec["content_type"] = resp.headers.get("Content-Type", "")
            rec["final_url"] = resp.geturl()
    except urllib.error.HTTPError as e:
        rec["status"] = f"http_{e.code}"
        rec["error"] = str(e.reason)
        _write_meta(meta, rec)
        return rec
    except (urllib.error.URLError, TimeoutError, OSError) as e:
        rec["status"] = "unreachable"
        rec["error"] = repr(e)
        _write_meta(meta, rec)
        return rec

    partial.write_bytes(body)
    rec["bytes"] = len(body)
    rec["content_length_declared"] = int(declared) if declared and declared.isdigit() else None
    if rec["content_length_declared"] is not None and rec["content_length_declared"] != len(body):
        # Truncated or over-read. The bytes stay in .partial and are never read
        # as data. Nothing is deleted.
        rec["status"] = "truncated"
        rec["kept_at"] = partial.name
        _write_meta(meta, rec)
        return rec

    rec["sha256"] = _sha256(body)
    final.write_bytes(body)
    _write_meta(meta, rec)
    return rec


def _write_meta(meta: Path, rec: dict) -> None:
    meta.parent.mkdir(parents=True, exist_ok=True)
    meta.write_text(
        json.dumps(rec, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n",
    )


# ---------------------------------------------------------------------------
# discover
# ---------------------------------------------------------------------------


def landing_url(fund: dict, overrides: dict) -> str:
    return overrides.get(fund["ticker"], fund["landing"])


def discover(tickers: list[str], third_party: bool, force: bool) -> dict:
    overrides = {}
    if LANDING_OVERRIDES.is_file():
        overrides = json.loads(LANDING_OVERRIDES.read_text(encoding="utf-8"))

    out: dict = {}
    for fund in FUNDS:
        t = fund["ticker"]
        if tickers and t not in tickers:
            continue
        url = landing_url(fund, overrides)
        rec = fetch(url, t, "landing.html", force=force)
        entry = {"landing": rec, "candidates": [], "third_party": None}

        final, _, _ = cache_paths(t, "landing.html")
        if rec.get("status", "").startswith("http_2") and final.is_file():
            page = final.read_bytes().decode("utf-8", errors="replace")
            ex = extract_links(page, rec.get("final_url", url))
            entry["candidates"] = ex["candidates"]
            entry["urls_seen"] = ex["urls_seen"]
            entry["dropped_assets"] = ex["dropped_assets"]
            entry["landing_bytes"] = len(page)

        if third_party:
            entry["third_party"] = fetch(
                THIRD_PARTY_CLOSE.format(t=t.lower()), t, "stooq_close.csv", force=force
            )
            tp_final, _, _ = cache_paths(t, "stooq_close.csv")
            if tp_final.is_file():
                head = tp_final.read_bytes()[:4000].decode("utf-8", errors="replace")
                entry["third_party"]["granularity"] = granularity_of_payload(head)
                entry["third_party"]["head"] = head.splitlines()[:3]

        out[t] = entry
        (CACHE / t).mkdir(parents=True, exist_ok=True)
        (CACHE / t / "discovered_links.json").write_text(
            json.dumps(entry, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n",
        )
        print(f"{t:5s} {rec.get('status','?'):12s} "
              f"urls={entry.get('urls_seen', 0)} "
              f"assets_dropped={entry.get('dropped_assets', 0)} "
              f"candidates={len(entry['candidates'])}"
              + (f" third_party={entry['third_party'].get('status','?')}"
                 if third_party else ""))
        for c in entry["candidates"][:40]:
            print(f"        {','.join(c['fields']):32s} {c['url'][:120]}")
        if not rec.get("from_cache"):
            time.sleep(SLEEP_S)
    return out


def scan_assets(tickers: list[str], max_assets: int, force: bool) -> dict:
    """Fetch the page's own JavaScript and read endpoint templates out of it."""
    out: dict = {}
    for fund in FUNDS:
        t = fund["ticker"]
        if tickers and t not in tickers:
            continue
        landing, _, meta = cache_paths(t, "landing.html")
        if not landing.is_file():
            print(f"{t:5s} no cached landing page. Run --discover first.")
            continue
        base = json.loads(meta.read_text(encoding="utf-8")).get(
            "final_url", fund["landing"])
        page = landing.read_bytes().decode("utf-8", errors="replace")

        assets = extract_assets(page, base)
        taken, skipped = assets[:max_assets], assets[max_assets:]
        print(f"{t:5s} same-origin js={len(assets)} fetching={len(taken)} "
              f"skipped_by_cap={len(skipped)}")

        hits: dict[str, dict[str, list[str]]] = {}
        fetched: list[dict] = []
        for url in taken:
            name = "asset_" + hashlib.sha1(url.encode()).hexdigest()[:10] + ".js"
            rec = fetch(url, t, name, force=force)
            fetched.append({"url": url, "status": rec.get("status"),
                            "bytes": rec.get("bytes")})
            final, _, _ = cache_paths(t, name)
            if not final.is_file():
                continue
            body = final.read_bytes().decode("utf-8", errors="replace")
            for pattern, matches in scan_text_for_endpoints(body).items():
                for m in matches:
                    hits.setdefault(pattern, {}).setdefault(m, [])
                    if name not in hits[pattern][m]:
                        hits[pattern][m].append(name)
            if not rec.get("from_cache"):
                time.sleep(SLEEP_S)

        entry = {
            "assets_total": len(assets),
            "assets_fetched": len(taken),
            "assets_skipped_by_cap": skipped,
            "fetched": fetched,
            "hits": {p: {m: sorted(v) for m, v in sorted(d.items())}
                     for p, d in sorted(hits.items())},
        }
        (CACHE / t).mkdir(parents=True, exist_ok=True)
        (CACHE / t / "asset_scan.json").write_text(
            json.dumps(entry, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        out[t] = entry

        # **Print the fetch tally before the hits.** Without it, twelve 403s and
        # twelve clean bundles containing nothing print the same line, which is
        # this file's own rule against inferring absence from a failed fetch,
        # broken at the display layer.
        tally: dict[str, int] = {}
        for f in fetched:
            tally[f["status"] or "?"] = tally.get(f["status"] or "?", 0) + 1
        got = sum(n for s, n in tally.items() if s.startswith("http_2"))
        print("    fetched: " + ", ".join(f"{s}={n}" for s, n in sorted(tally.items())))

        for pattern in sorted(hits):
            print(f"    {pattern} ({len(hits[pattern])})")
            for m in sorted(hits[pattern])[:25]:
                print(f"        {m[:150]}")
        if not hits and got == 0:
            print("    **no bundle was retrieved**, so this says nothing about "
                  "what the bundles contain.")
        elif not hits:
            print(f"    nothing endpoint-shaped in the {got} bundles retrieved. "
                  "That is a reading, and it covers those bundles only.")
        if skipped:
            print(f"    {len(skipped)} bundles were not looked at (cap). "
                  "Raise --max-assets to cover them.")
    return out


# ---------------------------------------------------------------------------
# Constructed candidates, with a positive control.
#
# The bundle scan found nothing in twelve of IVV's thirty-four scripts. That
# leaves two readings apart, and only a fetch separates them: the endpoint is in
# one of the other twenty-two, or the site does not use one. So the endpoint
# shapes below are **tried**, and each result is the HTTP status that came back.
#
# **The control line is the point of this table.** `latest-holdings.csv` was
# measured at 200 by --discover, so it is fetched here beside the untested ones.
# If the control passes and the rest 404, the failures are about those URLs. If
# everything fails together, the reading is about access, not about what the
# issuer publishes, and §10's rule then forbids writing it down as absence.
# ---------------------------------------------------------------------------

#: **Two controls, and the second one is the informative one.** The first round
#: had one control on a different URL shape (`latest-holdings.csv`), so a soft
#: 404 on every `.ajax` line left two readings open: the ajax shape is wrong, or
#: the `fileName` report token is wrong. `{T}_holdings` on the ajax shape splits
#: them, and it is documented rather than guessed: two independent third-party
#: scrapers use `1467271812596.ajax?fileType=csv&fileName={ticker}_holdings
#: &dataType=fund`, and a third uses the same id with
#: `fileType=json&tab=all&asOfDate=`. Cited as literature per engineering rule 9;
#: nothing here is taken from them as a measurement.
CONSTRUCTED_ENDPOINTS = {
    "iShares": [
        {"id": "control_holdings", "control": True,
         "path": "{prod}/latest-holdings.csv",
         "source": "--discover measured 200; positive control, a different URL shape"},
        {"id": "control_ajax_holdings", "control": True,
         "path": "{prod}/1467271812596.ajax?fileType=csv&fileName={T}_holdings"
                 "&dataType=fund",
         "source": "skiamu/ETF and talsan/ishares, used in two places, a positive control for the ajax shape itself"},
        {"id": "nav_history",
         "path": "{prod}/1467271812596.ajax?fileType=csv&fileName={T}_NAV&dataType=fund",
         "source": "written from memory，measured a soft 404 on the first pass，logged with both reported"},
        {"id": "nav_history_alt",
         "path": "{prod}/1467271812596.ajax?fileType=csv&fileName={T}_fund&dataType=fund",
         "source": "as above; measured a soft 404 on the first pass，logged with both reported"},
        {"id": "premium_discount",
         "path": "{prod}/1467271812596.ajax?fileType=csv&fileName={T}_premium_discount"
                 "&dataType=fund",
         "source": "as above; measured a soft 404 on the first pass，logged with both reported"},
        {"id": "distributions",
         "path": "{prod}/1467271812596.ajax?fileType=csv&fileName={T}_distributions"
                 "&dataType=fund",
         "source": "as above; measured a soft 404 on the first pass，logged with both reported"},
        {"id": "performance",
         "path": "{prod}/1467271812596.ajax?fileType=csv&fileName={T}_performance"
                 "&dataType=fund",
         "source": "a new candidate for the report term，unverified"},
        {"id": "nav_history_token",
         "path": "{prod}/1467271812596.ajax?fileType=csv&fileName={T}_nav_history"
                 "&dataType=fund",
         "source": "a new candidate for the report term，unverified"},
    ],
    # **SSGA publishes static files at predictable paths.** No product id, no
    # ajax, no session. Enumerated off the XLF and SPY pages on 2026-08-16, each
    # link quoted from the page that carries it, so `abs` templates take the
    # ticker alone and need no product path.
    #
    # `pdhist` is premium/discount history and is the one that matters: it is the
    # only free per-day series found so far that carries the loop's inputs.
    # **No issuer found so far publishes a bid-ask spread history**, iShares and
    # SSGA both give the 30-day median as one number, which is §12.1's problem
    # measured on two issuers rather than argued from one.
    "SSGA": [
        {"id": "control_holdings", "control": True, "abs": True,
         "path": "https://www.ssga.com/library-content/products/fund-data/etfs/us/"
                 "holdings-daily-us-en-{t}.xlsx",
         "source": "XLF and SPY, both pages labelled Download All Holdings: Daily, positive control"},
        {"id": "premium_discount_history", "abs": True,
         "path": "https://www.ssga.com/library-content/products/fund-data/etfs/us/"
                 "pdhist-us-en-{t}.xlsx",
         "source": "both pages are labelled Fund Data，not measured"},
        {"id": "nav_history", "abs": True,
         "path": "https://www.ssga.com/library-content/products/fund-data/etfs/us/"
                 "navhist-us-en-{t}.xlsx",
         "source": "both pages are labelled Most Recent NAV / NAV History，not measured"},
        {"id": "product_data_all_funds", "abs": True,
         "path": "https://www.ssga.com/library-content/products/fund-data/etfs/us/"
                 "spdr-product-data-us-en.xlsx",
         "source": "SPY page labelled Download Product Data, one table across funds; "
                   "may carry a 30-day median spread snapshot per fund，not measured"},
    ],
}


def product_base(ticker: str) -> str | None:
    """Derive the product path from the holdings URL that --discover measured.
    Earned from a link that returned 200, not written into this file."""
    p = CACHE / ticker / "discovered_links.json"
    if not p.is_file():
        return None
    entry = json.loads(p.read_text(encoding="utf-8"))
    for c in entry.get("candidates", []):
        u = c["url"]
        if u.lower().endswith("/latest-holdings.csv"):
            return u.rsplit("/", 1)[0]
    return None


def try_endpoints(tickers: list[str], force: bool) -> dict:
    out: dict = {}
    for fund in FUNDS:
        t, issuer = fund["ticker"], fund["issuer"]
        if tickers and t not in tickers:
            continue
        # A retired fund is skipped unless it is named. Re-probing the iShares
        # rows would spend seven requests to re-measure seven soft 404s that are
        # already recorded, and `--only IVV` still reaches them.
        if not tickers and fund["arm"] == "retired_availability":
            continue
        templates = CONSTRUCTED_ENDPOINTS.get(issuer)
        if not templates:
            print(f"{t:5s} no declared template for {issuer}. Nothing tried, "
                  "and nothing concluded.")
            continue
        needs_prod = any(not s.get("abs") for s in templates)
        prod = product_base(t)
        if needs_prod and prod is None:
            print(f"{t:5s} no measured product path. Run --discover first.")
            continue

        print(f"{t:5s} {prod or '(absolute templates, no product path needed)'}")
        rows = []
        for spec in templates:
            url = spec["path"].format(prod=prod, T=t, t=t.lower())
            rec = fetch(url, t, f"constructed_{spec['id']}", force=force)
            row = {"id": spec["id"], "url": url, "source": spec["source"],
                   "status": rec.get("status"), "bytes": rec.get("bytes"),
                   "content_type": rec.get("content_type", "")}
            final, _, _ = cache_paths(t, f"constructed_{spec['id']}")
            if final.is_file():
                raw = final.read_bytes()
                text = raw.decode("utf-8", errors="replace")
                row["kind"] = payload_kind(raw)
                row["granularity"] = granularity_of_payload(text)
                row["head"] = text.splitlines()[:3]
            row["verdict"] = classify_response(row["status"], row.get("kind", "none"))
            rows.append(row)
            print(f"    {spec['id']:20s} {row['status']:10s} "
                  f"{str(row.get('bytes', '')):>9s}  {row.get('kind', '-'):10s} "
                  f"{row['verdict']:16s} {row['content_type'][:32]}")
            if not rec.get("from_cache"):
                time.sleep(SLEEP_S)

        # **The verdict reads `verdict`, not the status.** The first version
        # counted 2xx and announced four working endpoints, all four of which
        # were the product page under a csv content type.
        controls = {s["id"] for s in templates if s.get("control")}
        passed_ctl = [r["id"] for r in rows
                      if r["id"] in controls and r["verdict"] == "ok_tabular"]
        ok = [r for r in rows if r["verdict"] == "ok_tabular"
              and r["id"] not in controls]
        soft = [r for r in rows if r["verdict"] == "soft_404_html"]
        ajax_ctl_ok = "control_ajax_holdings" in passed_ctl

        if not passed_ctl:
            verdict = ("**no control returned a table**, so every line above is "
                       "about access and none of it is about what the issuer "
                       "publishes")
        elif ok:
            verdict = (f"{len(ok)} non-control endpoint(s) returned a table: "
                       f"{', '.join(r['id'] for r in ok)}")
        elif ajax_ctl_ok:
            verdict = (f"**the ajax shape works and every report token failed** "
                       f"({len(soft)} soft 404s). The id and the parameter names "
                       f"are right, so what is wrong is the `fileName` token, and "
                       f"the remaining tokens are unknown rather than absent")
        else:
            verdict = (f"only the non-ajax control returned a table "
                       f"({len(soft)} soft 404s). **The ajax shape itself did not "
                       f"work here**, so nothing is settled about report tokens")
        print(f"    verdict: {verdict}")

        entry = {"product_base": prod, "rows": rows, "verdict": verdict}
        (CACHE / t / "constructed_endpoints.json").write_text(
            json.dumps(entry, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        out[t] = entry
    return out


# ---------------------------------------------------------------------------
# inspect: what did those payloads turn out to be. No network.
#
# The four constructed iShares URLs all returned 200 and about 2.23 MB, and
# their byte counts differ by exactly the length differences of the `fileName`
# values: IVV_NAV is 7 characters against IVV_premium_discount's 20, and the
# payloads differ by 13 bytes. **Four names, one file**, with the name echoed
# inside it. Reporting them as four available fields would be the eighth failure
# mode, membership error, in its download form: the label on the request is not
# the content of the response.
#
# So the identity check is mechanical. Strip every `<TICKER>_<word>` token, hash
# what is left, and group. Payloads that collapse into one group are one file.
# ---------------------------------------------------------------------------


def payload_kind(raw: bytes) -> str:
    """What the bytes are, read from the bytes.

    **The header lied and the status lied.** Four constructed iShares URLs
    returned `200` with `Content-Type: text/csv;charset=UTF-8` and 2.23 MB of
    HTML product page, differing from each other only in a per-response CSP
    nonce at offset 1371. A soft 404. `stooq` returned `200` and an anti-bot
    JavaScript challenge. Neither is visible to a status check, so the shape is
    taken from the payload and nothing else."""
    if raw[:4] == b"PK\x03\x04":
        # xlsx and xls are zip containers. A table, and it must not fall through
        # to "other", or the verdict would read a working endpoint as a failure.
        return "xlsx_zip"
    head = raw[:2000].lstrip().lower()
    if head.startswith(b"<!doctype") or head.startswith(b"<html") or b"<html" in head:
        return "html"
    if head.startswith((b"{", b"[")):
        return "json"
    if any(d in head for d in (b",", b"|", b"\t")):
        return "delimited"
    return "other"


def classify_response(status: str, kind: str) -> str:
    """The verdict a single fetch earns. A 2xx that carries a page is a soft
    404 and **may not be counted as data**."""
    if not (status or "").startswith("http_2"):
        return "non_2xx"
    if kind == "html":
        return "soft_404_html"
    if kind in ("delimited", "xlsx_zip"):
        return "ok_tabular"
    return f"2xx_but_{kind}"


def normalised_digest(text: str, ticker: str) -> str:
    """Hash with the echoed filename removed, so that two responses that differ
    only in the name they were asked for hash the same.

    Case-insensitive: the first version matched `IVV_` only, and an echo written
    `ivv_premium_discount` slipped through, which left four payloads in four
    groups while their byte counts said one file."""
    stripped = re.sub(rf"{re.escape(ticker)}_[A-Za-z_]+", "", text, flags=re.IGNORECASE)
    return hashlib.sha256(stripped.encode("utf-8", "replace")).hexdigest()


def first_difference(a: bytes, b: bytes, window: int = 90) -> dict:
    """Where two payloads stop agreeing, with a window from each.

    **This is the measurement that settles it.** Four byte counts differing by
    exactly the length differences of four names is an argument. The offset of
    the first differing byte, and what sits there, is a reading."""
    n = min(len(a), len(b))
    i = 0
    while i < n and a[i] == b[i]:
        i += 1
    if i == n and len(a) == len(b):
        return {"identical": True}
    lo = max(0, i - window // 3)
    return {
        "identical": False,
        "offset": i,
        "differs_only_in_length": i == n,
        "a": a[lo:i + window].decode("utf-8", "replace").replace("\n", "\\n"),
        "b": b[lo:i + window].decode("utf-8", "replace").replace("\n", "\\n"),
    }


def inspect(tickers: list[str], head_lines: int) -> dict:
    out: dict = {}
    for fund in FUNDS:
        t, issuer = fund["ticker"], fund["issuer"]
        if tickers and t not in tickers:
            continue
        names = [f"constructed_{s['id']}"
                 for s in CONSTRUCTED_ENDPOINTS.get(issuer, [])]
        names.append("stooq_close.csv")
        rows = []
        for name in names:
            final, _, _ = cache_paths(t, name)
            if not final.is_file():
                continue
            raw = final.read_bytes()
            kind = payload_kind(raw)
            if kind == "xlsx_zip":
                # A zip is not text. Profiling its bytes would print a date
                # count read out of compressed noise.
                text = ""
                prof = {"lines": 0, "distinct_dates": 0, "granularity": "binary",
                        "date_first_in_file": "", "date_last_in_file": "",
                        "header_line_index": None, "header_line": ""}
            else:
                text = raw.decode("utf-8", errors="replace")
                prof = payload_profile(text)
            rows.append({
                "name": name, "bytes": len(raw),
                "kind": kind,
                "sha256": _sha256(raw)[:16],
                "normalised": normalised_digest(text, t)[:16],
                **prof,
                "head": text.splitlines()[:head_lines],
            })
        if not rows:
            print(f"{t:5s} nothing cached to inspect.")
            continue

        groups: dict[str, list[str]] = {}
        for r in rows:
            groups.setdefault(r["normalised"], []).append(r["name"])

        print(f"\n{t}")
        print(f"    {'name':30s} {'bytes':>9s} {'kind':10s} {'lines':>7s} "
              f"{'dates':>7s} {'granularity':12s} {'norm':10s}")
        for r in rows:
            print(f"    {r['name']:30s} {r['bytes']:>9d} {r['kind']:10s} "
                  f"{r['lines']:>7d} {r['distinct_dates']:>7d} "
                  f"{r['granularity']:12s} {r['normalised'][:8]:10s}")
        for digest, members in sorted(groups.items()):
            if len(members) > 1:
                print(f"    **same file under {len(members)} names**: "
                      f"{', '.join(sorted(members))}")

        # One head per group, not one head for the fund. The first version
        # printed rows[0] only, which was the holdings control, so the payload
        # the run was actually about never appeared on screen.
        by_digest = {}
        for r in rows:
            by_digest.setdefault(r["normalised"], r)
        for digest, r in sorted(by_digest.items(), key=lambda kv: kv[1]["name"]):
            print(f"\n    --- {r['name']} ({r['bytes']} bytes, "
                  f"{r['distinct_dates']} distinct dates) ---")
            if r.get("header_line_index") is not None:
                print(f"    table header at line {r['header_line_index']}: "
                      f"{r['header_line'][:200]}")
            print(f"    dates in file order: {r['date_first_in_file']} .. "
                  f"{r['date_last_in_file']}")
            for ln in r["head"]:
                print(f"        {ln[:170]}")

        # Where the constructed payloads stop agreeing.
        constructed = [r for r in rows
                       if r["name"].startswith("constructed_")
                       and r["name"] != "constructed_control_holdings"]
        diffs = {}
        if len(constructed) >= 2:
            base_name = constructed[0]["name"]
            base = cache_paths(t, base_name)[0].read_bytes()
            print(f"\n    first byte where each differs from {base_name}:")
            for r in constructed[1:]:
                other = cache_paths(t, r["name"])[0].read_bytes()
                d = first_difference(base, other)
                diffs[r["name"]] = d
                if d.get("identical"):
                    print(f"        {r['name']:32s} identical")
                else:
                    print(f"        {r['name']:32s} offset {d['offset']} "
                          f"of {len(base)}")
                    print(f"            base : {d['a'][:150]}")
                    print(f"            this : {d['b'][:150]}")

        # **A soft 404 is still a page, and a page can carry links.** The 2.23 MB
        # that came back is the legacy server-rendered product template, which is
        # a different document from the Astro landing page --discover fetched.
        # Running the extractor over it costs nothing and may hand over the real
        # endpoint. Free, local, no request.
        requested = {}
        ce = CACHE / t / "constructed_endpoints.json"
        if ce.is_file():
            for row in json.loads(ce.read_text(encoding="utf-8")).get("rows", []):
                requested[f"constructed_{row['id']}"] = row["url"]

        salvage = {}
        for digest, r in sorted(by_digest.items(), key=lambda kv: kv[1]["name"]):
            if r["kind"] != "html" or r["name"] == "landing.html":
                continue
            page = cache_paths(t, r["name"])[0].read_bytes().decode("utf-8", "replace")
            ex = extract_links(page, fund["landing"])
            # **Every ajax URL, not the classified ones.** The first version
            # filtered to the five field ids, which hid `unclassified_download`
            # exactly where an unfamiliar `fileName` or a different numeric id
            # would land, that is precisely the thing being looked for.
            ajax = sorted({unescape(u) for u in
                           scan_text_for_endpoints(page)["ajax_endpoint"]
                           if "fileType" in u or "fileName" in u})
            asked = requested.get(r["name"], "")
            echoes = [u for u in ajax if asked and u.split("?", 1)[-1] in asked]
            fresh = [u for u in ajax if u not in echoes]
            salvage[r["name"]] = {"urls_seen": ex["urls_seen"],
                                  "ajax_all": ajax, "ajax_echoed": echoes,
                                  "ajax_new": fresh,
                                  "candidates": ex["candidates"]}
            print(f"\n    inside {r['name']}: {ex['urls_seen']} urls, "
                  f"{len(ajax)} ajax, {len(echoes)} of them the request echoed "
                  f"back, {len(fresh)} new")
            for u in fresh[:30]:
                print(f"        NEW  {u[:150]}")
            for u in echoes[:3]:
                print(f"        echo {u[:150]}")
            if not fresh:
                print("        nothing this page offers that was not asked for. "
                      "**The echo is not a discovery.**")

        entry = {"rows": rows, "diffs": diffs, "salvage": salvage,
                 "groups": {k: sorted(v) for k, v in sorted(groups.items())}}
        (CACHE / t / "payload_inspection.json").write_text(
            json.dumps(entry, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        out[t] = entry
    return out


# ---------------------------------------------------------------------------
# inspect-xlsx: what is inside the SSGA workbooks. No network.
#
# `--inspect` reports that a payload is a zip and stops there, on purpose:
# profiling compressed bytes as text prints a date count read out of noise.
# This mode opens the workbook instead and prints sheet names, the header row,
# the row count and the date span, which is what decides whether `pdhist`
# carries the loop's inputs.
# ---------------------------------------------------------------------------


#: SSGA writes dates as **text**, `14-Aug-2026`, not as workbook dates. The
#: first version counted `isinstance(c, datetime)` and reported zero date cells
#: and an empty span for a file that is nothing but a dated series.
XLSX_DATE_FORMATS = ("%d-%b-%Y", "%b %d %Y", "%b %d, %Y", "%Y-%m-%d", "%m/%d/%Y")


def as_iso_date(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, str):
        s = v.strip()
        for fmt in XLSX_DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                pass
    return None


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, float):
        return f"{v:.6g}"
    return str(v)


def inspect_xlsx(tickers: list[str], head_rows: int, max_rows: int = 200_000) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed in this venv. Install it, then re-run:")
        print("    python -m pip install openpyxl")
        print("Nothing was read and nothing is concluded.")
        return {}

    out: dict = {}
    for fund in FUNDS:
        t, issuer = fund["ticker"], fund["issuer"]
        if tickers and t not in tickers:
            continue
        entry = {}
        for spec in CONSTRUCTED_ENDPOINTS.get(issuer, []):
            name = f"constructed_{spec['id']}"
            final, _, _ = cache_paths(t, name)
            if not final.is_file() or final.read_bytes()[:4] != b"PK\x03\x04":
                continue
            print(f"\n{t}  {name}  ({final.stat().st_size} bytes)")
            # Hand openpyxl a buffer, not the path: the cache names carry no
            # extension, and openpyxl refuses a path it cannot recognise by
            # suffix even when the bytes are a valid workbook.
            wb = load_workbook(io.BytesIO(final.read_bytes()),
                               read_only=True, data_only=True)
            sheets = {}
            for ws in wb.worksheets:
                rows = []
                capped = False
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_rows:
                        capped = True
                        break
                    rows.append(row)
                # Dates are text here, so parse rather than isinstance-check.
                dates = sorted({d for r in rows for c in r
                                if (d := as_iso_date(c)) is not None})
                nonempty = [r for r in rows if any(x is not None for x in r)]
                # header: the nearest row above the first dated row that is not
                # itself dated and carries at least two labels
                first_dated = next((i for i, r in enumerate(rows)
                                    if any(as_iso_date(c) for c in r)), None)
                header = ""
                if first_dated is not None:
                    for i in range(first_dated - 1, -1, -1):
                        vals = [x for x in rows[i] if x is not None]
                        if len(vals) >= 2 and not any(as_iso_date(x) for x in vals):
                            header = " | ".join(_cell(x) for x in rows[i] if x is not None)
                            break
                sheets[ws.title] = {
                    "rows_scanned": len(rows), "rows_nonempty": len(nonempty),
                    "row_cap_hit": capped,
                    "distinct_dates": len(dates),
                    "date_min": dates[0] if dates else "",
                    "date_max": dates[-1] if dates else "",
                    "header": header[:300],
                }
                print(f"  sheet {ws.title!r}: {len(nonempty)} non-empty rows, "
                      f"{len(dates)} distinct dates, "
                      f"{sheets[ws.title]['date_min']} .. {sheets[ws.title]['date_max']}"
                      + ("  **ROW CAP HIT, the count above is this scanner's "
                         "limit and not the file's length**" if capped else ""))
                if header:
                    print(f"    header: {header[:220]}")
                for r in rows[:head_rows]:
                    line = " | ".join(_cell(x) for x in r)[:200]
                    if line.strip(" |"):
                        print(f"      {line}")
            wb.close()
            entry[name] = sheets
        if entry:
            (CACHE / t / "xlsx_inspection.json").write_text(
                json.dumps(entry, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8", newline="\n")
            out[t] = entry
        else:
            print(f"{t:5s} no cached workbook. Run --try-endpoints first.")
    return out


def grep_cells(tickers: list[str], pattern: str, scan_rows: int) -> dict:
    """Every cell in the first `scan_rows` rows matching `pattern`, with its
    sheet, row and column. Local, no network.

    **Why a whole mode for this.** `--inspect-xlsx` prints one header line, and
    a workbook with a banked header (SSGA's product-data sheet spreads its
    column names over three rows) hides most of its columns behind that one
    line. Deciding that a column is absent by reading one printed row is the
    same defect as reading granularity out of the first 2,000 bytes.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed. python -m pip install openpyxl")
        return {}

    rx = re.compile(pattern, re.IGNORECASE)
    out: dict = {}
    for fund in FUNDS:
        t, issuer = fund["ticker"], fund["issuer"]
        if tickers and t not in tickers:
            continue
        for spec in CONSTRUCTED_ENDPOINTS.get(issuer, []):
            name = f"constructed_{spec['id']}"
            final, _, _ = cache_paths(t, name)
            if not final.is_file() or final.read_bytes()[:4] != b"PK\x03\x04":
                continue
            wb = load_workbook(io.BytesIO(final.read_bytes()),
                               read_only=True, data_only=True)
            hits = []
            for ws in wb.worksheets:
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= scan_rows:
                        break
                    for j, c in enumerate(row):
                        s = _cell(c)
                        if s and rx.search(s):
                            hits.append({"sheet": ws.title, "row": i, "col": j,
                                         "text": s[:120]})
            wb.close()
            if hits:
                out.setdefault(t, {})[name] = hits
                print(f"\n{t}  {name}: {len(hits)} matching cells")
                for h in hits[:40]:
                    print(f"    {h['sheet']}!r{h['row']}c{h['col']}  {h['text']}")
    if not out:
        print(f"no cell matched /{pattern}/ in the first {scan_rows} rows of any "
              f"cached workbook. **That covers those rows only.**")
    return out


# ---------------------------------------------------------------------------
# daily capture, registered by §13.3 item 4
#
# The median bid-ask spread exists only as today's figure, so the series this
# stage lacks can only be built forward. One 93 KB cross-fund workbook carries
# it for about 199 SPDR funds at once, so the capture is one request a day.
#
# **Raw file only, no parsing.** A capture that parses is a capture that can lose
# a column when the publisher moves one. The bytes are the asset; the column is
# read later, from the bytes, as often as needed.
# ---------------------------------------------------------------------------

DAILY = CACHE / "_daily"
DAILY_URL = ("https://www.ssga.com/library-content/products/fund-data/etfs/us/"
             "spdr-product-data-us-en.xlsx")

# **§34.7's correction.** The workbook above carries the spread, which has no
# history and can only be built forward. `pdhist` is the opposite problem: it
# has a history and **loses it from the back**, a rolling 404 days, so a day not
# archived before it falls off is gone. Capturing only the workbook leaves the
# window sliding rather than growing, and §34.6 needs it to grow: the arm wants
# about 1,600 trading days and has 404.
#
# Retention is 404 days so any interval under about four hundred days loses
# nothing, and daily is chosen because it is the simplest thing that cannot
# silently skip. **Raw bytes only, no parsing**, same rule as the workbook.
PDHIST_URL = ("https://www.ssga.com/library-content/products/fund-data/etfs/us/"
              "pdhist-us-en-{t}.xlsx")


def pdhist_url(ticker: str) -> str:
    """**One place, because two places drifted apart on the first run.**
    `try_endpoints` substitutes `{t}` with the **lowercase** ticker, SSGA's
    paths are lowercase, and the first capture formatted the uppercase one and
    fetched sixteen HTML error pages.

    **The selftest that was supposed to catch this compared the two templates
    and passed**, because it checked the string and not the call. The check
    below now goes through this function."""
    return PDHIST_URL.format(t=ticker.lower())


def capture_set() -> list:
    """The funds whose premium history is archived. **Retired availability rows
    are excluded by their own field rather than by a second list**, so a fund
    cannot be in the sample and out of the archive at the same time."""
    return [f["ticker"] for f in FUNDS
            if f["issuer"] == "SSGA" and f.get("arm") != "retired_availability"]


# ---------------------------------------------------------------------------
# §36: B9-A-6. Does one venue's closing BBO midpoint equal the disclosed NBBO
# midpoint? 404 days of ground truth already on disk decide it.
#
# **Nothing here parses a workbook.** `b9_omega.py --dump-days` exports the day
# list, the NAV, the disclosed premium and §6's calm flag, so the sample this
# compares on is the sample the stage measured on, filter for filter.
# ---------------------------------------------------------------------------

NBBO = CACHE / "_nbbo"
DAYS_JSON = ROOT / "results" / "b9_days.json"
DB_BASE = "https://hist.databento.com/v0"
DB_DATASET = "XNAS.ITCH"          # §36.2: one venue, coverage from 2018
DB_SCHEMA = "bbo-1s"              # one-second subsampled top of book
NAV_STRIKE_ET = (16, 0, 0)        # §36.2: when NAV is struck
WINDOW_SECONDS = 10               # 15:59:55 to 16:00:05, both sides of the mark
HALF_CENT_EQUAL = 1e-4            # §36.4: equality on the half-cent grid
# **Read off the first capture, not assumed.** The CSV carries prices as
# fixed-point nanodollars (`25015000000` is `$25.015`) and timestamps as epoch
# nanoseconds, and it carries **no symbol column** unless `map_symbols` is
# asked for. The first parser assumed ISO timestamps, decimal prices and a
# symbol column, and matched nothing, which is how the three were found.
DB_PX_SCALE = 1e-9
DB_MAP_SYMBOLS = "true"


def _days_from_civil(y: int, m: int, d: int) -> int:
    """Days since 1970-01-01, Hinnant's algorithm. Written out because the mark
    has to be an epoch nanosecond and a date library is not imported here."""
    y -= m <= 2
    era = (y if y >= 0 else y - 399) // 400
    yoe = y - era * 400
    doy = (153 * (m + (-3 if m > 2 else 9)) + 2) // 5 + d - 1
    doe = yoe * 365 + yoe // 4 - yoe // 100 + doy
    return era * 146097 + doe - 719468


def mark_ns(day: str) -> int:
    """`16:00:00` America/New_York on `day`, in epoch nanoseconds. **The same
    offset function the fetch window uses**, so the two cannot drift."""
    y, m, d = (int(x) for x in day.split("-"))
    h, mi, s = NAV_STRIKE_ET
    secs = (_days_from_civil(y, m, d) * 86400
            + (h + et_offset_hours(day)) * 3600 + mi * 60 + s)
    return secs * 1_000_000_000


#: §38.4(b): asked of the vendor, not read off a marketing page. `ARCX.PILLAR`
#: is first because these funds are NYSE Arca listings, which §36.2 did not
#: consider and should have.
CANDIDATE_DATASETS = ["ARCX.PILLAR", "XNAS.ITCH", "XNYS.PILLAR", "BATS.PITCH",
                      "IEXG.TOPS", "DBEQ.BASIC", "EQUS.MINI", "EQUS.SUMMARY"]


def db_key() -> str | None:
    """Read from the environment only. **A key never enters this file, the
    repository or the cache**, and its absence is a normal state that prints an
    instruction rather than raising."""
    import os
    return os.environ.get("DATABENTO_API_KEY") or None


def et_offset_hours(day: str) -> int:
    """America/New_York offset from UTC for a date, without a timezone library.
    US DST runs from the second Sunday in March to the first Sunday in November.

    **Written out rather than approximated** because a one-hour error puts the
    sample an hour away from the close and would read as the venue disagreeing
    with the NBBO."""
    y, m, d = (int(x) for x in day.split("-"))

    def _dow(yy, mm, dd):                      # Sakamoto, 0 = Sunday
        t = [0, 3, 2, 5, 0, 3, 5, 1, 4, 6, 2, 4]
        if mm < 3:
            yy -= 1
        return (yy + yy // 4 - yy // 100 + yy // 400 + t[mm - 1] + dd) % 7

    # second Sunday in March
    mar = [dd for dd in range(1, 32) if _dow(y, 3, dd) == 0][1]
    # first Sunday in November
    nov = [dd for dd in range(1, 31) if _dow(y, 11, dd) == 0][0]
    if (m, d) < (3, mar) or (m, d) > (11, nov):
        return 5                                # EST
    if (m, d) == (3, mar) or (m, d) == (11, nov):
        return 5 if (m == 11) else 4            # transition days, close is DST-side
    return 4                                    # EDT


def nbbo_window(day: str) -> tuple[str, str]:
    """The UTC range that brackets `16:00:00` America/New_York on `day`."""
    off = et_offset_hours(day)
    h, mi, s = NAV_STRIKE_ET
    base = (h + off) * 3600 + mi * 60 + s
    lo, hi = base - WINDOW_SECONDS // 2, base + WINDOW_SECONDS // 2
    fmt = lambda n: f"{day}T{n // 3600:02d}:{(n % 3600) // 60:02d}:{n % 60:02d}"
    return fmt(lo), fmt(hi)


def db_post(path: str, params: dict, key: str) -> tuple[int, bytes]:
    """One request against the vendor's HTTP API. Basic auth, key as username,
    empty password, which is what the vendor documents."""
    body = urllib.parse.urlencode(params, doseq=True).encode("utf-8")
    req = urllib.request.Request(f"{DB_BASE}/{path}", data=body,
                                 headers={"User-Agent": USER_AGENT})
    token = base64.b64encode(f"{key}:".encode("utf-8")).decode("ascii")
    req.add_header("Authorization", f"Basic {token}")
    try:
        with urllib.request.urlopen(req, timeout=90) as r:
            return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()
    except Exception as e:                       # never raises on a network state
        return 0, str(e).encode("utf-8")


def db_get(path: str, params: dict, key: str) -> tuple[int, bytes]:
    """A GET against the vendor's metadata endpoints."""
    q = urllib.parse.urlencode(params)
    req = urllib.request.Request(f"{DB_BASE}/{path}?{q}",
                                 headers={"User-Agent": USER_AGENT})
    token = base64.b64encode(f"{key}:".encode("utf-8")).decode("ascii")
    req.add_header("Authorization", f"Basic {token}")
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()
    except Exception as e:
        return 0, str(e).encode("utf-8")


def nbbo_dir(dataset: str) -> Path:
    """Captures are namespaced by dataset so two venues cannot collide.

    **The flat layout predates `--dataset` and is moved, not deleted**, which is
    the standing rule. The move runs once and is a no-op afterwards."""
    d = NBBO / dataset
    d.mkdir(parents=True, exist_ok=True)
    if dataset == "XNAS.ITCH":
        moved = 0
        for p in sorted(NBBO.glob("bbo-*")):
            if p.is_file():
                p.rename(d / p.name)
                moved += 1
        if moved:
            print(f"    moved {moved} flat captures into "
                  f"{d.relative_to(ROOT)} (§38: namespaced by dataset)")
    return d


def nbbo_datasets(enumerate_all: bool = False) -> dict:
    """§38.4(b). What the vendor says its own coverage is.

    With `enumerate_all`, the full catalogue is listed first rather than only
    the hand-written candidates. **B9-B needs off-exchange prints**: the four
    exchange feeds miss the TRFs, which carry a large share of volume and carry
    it in a way that is correlated with trade size, which is §30.3's defect in a
    new position. A hand-written list cannot discover a dataset nobody thought
    of, so the catalogue is asked for."""
    key = db_key()
    if not key:
        print("DATABENTO_API_KEY is not set. Nothing requested.")
        return {}
    out = {}
    if enumerate_all:
        status, raw = db_get("metadata.list_datasets", {}, key)
        try:
            allds = json.loads(raw)
        except Exception:
            allds = []
        print(f"catalogue: http {status}, {len(allds)} datasets")
        for ds in sorted(allds):
            print(f"    {ds}")
        out["_catalogue"] = allds
    for ds in CANDIDATE_DATASETS:
        status, raw = db_get("metadata.get_dataset_range", {"dataset": ds}, key)
        body = raw[:220].decode("utf-8", "replace").replace("\n", " ")
        out[ds] = {"status": status, "body": body}
        print(f"{ds:14s} http {status}  {body}")
    (ROOT / "results" / "b9_datasets.json").write_text(
        json.dumps(out, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print("wrote results/b9_datasets.json")
    print("§38.4(b): whatever this says is what gets recorded. A start date "
          "later than 2018 makes the 2020 and 2022 objective a finding about "
          "the route.")
    return out


def load_days() -> dict:
    if not DAYS_JSON.is_file():
        print(f"{DAYS_JSON.relative_to(ROOT)} is absent. Run "
              f"`b9_omega.py --dump-days` first: §36 takes its sample from the "
              f"stage, not from a second copy of the filter.")
        return {}
    return json.loads(DAYS_JSON.read_text(encoding="utf-8"))


def nbbo_cost() -> dict:
    """§36.6(2). Estimate before spending, and print the extrapolation openly."""
    key = db_key()
    if not key:
        print("DATABENTO_API_KEY is not set. Nothing requested.")
        return {}
    rec = load_days()
    if not rec:
        return {}
    days, syms = rec["days"], sorted(rec["funds"])
    lo, hi = nbbo_window(days[len(days) // 2])
    params = {"dataset": DB_DATASET, "symbols": ",".join(syms),
              "schema": DB_SCHEMA, "start": lo, "end": hi,
              "stype_in": "raw_symbol", "mode": "historical"}
    status, raw = db_post("metadata.get_cost", params, key)
    print(f"one day ({lo} to {hi}, {len(syms)} symbols): http {status} "
          f"{raw[:200].decode('utf-8', 'replace')}")
    try:
        per_day = float(json.loads(raw))
    except Exception:
        print("**cost not parsed.** Nothing is fetched on an unparsed estimate.")
        return {"status": status, "parsed": False}
    total = per_day * len(days)
    print(f"§36.6  per day ${per_day:.6f} x {len(days)} days = **${total:.2f}**")
    print(f"§36.6  this is an extrapolation from one day, stated as such. "
          f"Nothing is fetched without --confirm.")
    return {"per_day": per_day, "days": len(days), "estimate_usd": total}


def _nbbo_one(day: str, syms: list, key: str, ddir: Path,
              dataset: str) -> tuple:
    """One day. **Returns rather than raises**, so a worker cannot take the
    pool down, and writes only a payload that is not a JSON error body."""
    target = ddir / f"bbo-{day}.csv"
    if target.is_file() and target.stat().st_size > 0:
        head = target.read_text(encoding="utf-8", errors="replace").split(
            "\n", 1)[0]
        if "symbol" in head.split(","):
            return day, "present", None
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
        target.rename(ddir / f"bbo-{day}.csv.expired_{stamp}_no_symbol")
    lo, hi = nbbo_window(day)
    params = {"dataset": dataset, "symbols": ",".join(syms),
              "schema": DB_SCHEMA, "start": lo, "end": hi,
              "stype_in": "raw_symbol", "encoding": "csv",
              "map_symbols": DB_MAP_SYMBOLS}
    # **Back off on a rate limit rather than losing the day.** A 429 that is
    # counted as a failure costs a whole further pass to recover; three short
    # waits cost seconds. Anything that is not a 429 is returned as it is.
    status, raw = 0, b""
    for attempt in range(3):
        status, raw = db_post("timeseries.get_range", params, key)
        if status != 429:
            break
        time.sleep(1.0 + 2.0 * attempt)
    if status != 200 or not raw or raw[:1] == b"{":
        return day, "failed", (status, raw[:120].decode("utf-8", "replace"))
    target.write_bytes(raw)
    return day, "fetched", None


def nbbo_fetch(confirm: bool, limit: int = 0, workers: int = 4,
               dataset: str = DB_DATASET) -> dict:
    """§36.6(2): refuses without an explicit confirmation. Idempotent per day,
    and a day already on disk is never refetched."""
    key = db_key()
    if not key:
        print("DATABENTO_API_KEY is not set. Nothing requested.")
        return {}
    if not confirm:
        print("**Refused.** Run --nbbo-cost first, then pass --confirm. "
              "§36.6: cost is printed before anything is fetched.")
        return {}
    rec = load_days()
    if not rec:
        return {}
    days, syms = rec["days"], sorted(rec["funds"])
    ddir = nbbo_dir(dataset)
    todo = days[:limit] if limit else days
    print(f"§38  dataset {dataset}, {len(todo)} days, {len(syms)} symbols")
    done, fetched, failed = 0, 0, []
    # **One request per day is the shape, and it is the shape on purpose**: the
    # alternative is a continuous range, which would pull every second of every
    # session, about 9 GB, to use ten seconds of each. So the fix for a slow
    # run is concurrency, not a wider request.
    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=max(1, workers)) as pool:
        for i, (day, state, err) in enumerate(
                pool.map(lambda d: _nbbo_one(d, syms, key, ddir, dataset),
                         todo), 1):
            if state == "present":
                done += 1
            elif state == "fetched":
                fetched += 1
            else:
                failed.append((day, err[0], err[1]))
            if i % 25 == 0 or i == len(todo):
                print(f"    {i}/{len(todo)}  on disk {done}  fetched {fetched}"
                      f"  failed {len(failed)}", flush=True)
    print(f"§36  days already on disk {done}, fetched {fetched}, failed "
          f"{len(failed)}")
    for f in failed[:5]:
        print(f"    {f[0]}: http {f[1]} {f[2]}")
    return {"present": done, "fetched": fetched, "failed": failed}


def _mid_at_close(rows: list, day: str, offset_s: int = 0,
                  min_size: int = 0) -> dict:
    """Per symbol, the midpoint of the last record at or before the mark, plus
    the crossed-or-locked flag §36.6(1) asks for.

    Timestamps are epoch nanoseconds and prices are fixed-point nanodollars,
    both read off the capture rather than assumed."""
    mark = mark_ns(day) + offset_s * 1_000_000_000
    out = {}
    for r in rows:
        sym = (r.get("symbol") or "").strip()
        if not sym:
            continue
        try:
            ts = int(r.get("ts_event"))
            bid = int(r.get("bid_px_00")) * DB_PX_SCALE
            ask = int(r.get("ask_px_00")) * DB_PX_SCALE
        except (TypeError, ValueError):
            continue
        if ts > mark or bid <= 0 or ask <= 0:
            continue
        # §42.4: the NBBO is a round-lot construct and a venue's level zero is
        # not. **A record failing this is dropped, not corrected** (§42.5):
        # bbo-1s carries no deeper level, so the round-lot quote behind an odd
        # lot is simply absent.
        if min_size:
            try:
                if (int(r.get("bid_sz_00")) < min_size
                        or int(r.get("ask_sz_00")) < min_size):
                    continue
            except (TypeError, ValueError):
                continue
        prev = out.get(sym)
        if prev is None or ts > prev["ts"]:
            out[sym] = {"ts": ts, "bid": bid, "ask": ask,
                        "mid": (bid + ask) / 2.0, "crossed": bid >= ask,
                        "spread_cents": round((ask - bid) * 100.0, 4)}
    return out


def nbbo_compare(dataset: str = DB_DATASET, offset_s: int = 0,
                 min_size: int = 0) -> dict:
    """§36.4 and §36.5. Exact equality on the half-cent grid, with the
    stress-clustering sub-test that decides the middle row."""
    rec = load_days()
    if not rec:
        return {}
    ddir = nbbo_dir(dataset)
    files = sorted(ddir.glob("bbo-*.csv"))
    if not files:
        print(f"no captures under {ddir.relative_to(ROOT)}. Nothing compared.")
        return {}
    print(f"§38  dataset {dataset}, offset {offset_s:+d}s, "
          f"min_size {min_size}, {len(files)} days")
    per_fund, crossed_n, total_n = {}, 0, 0
    diffs_bp, no_symbol, steps_all = [], [], {}
    # **§49.5(2).** The denominator used to be "fund-days we happened to get".
    # Every registered fund-day is now accounted for: matched, or missing a
    # venue record, or on a day with no capture at all.
    registered = {(t, d) for t, days in rec["funds"].items() for d in days}
    seen = set()
    empty_days, header_only = 0, 0
    for p in files:
        day = p.name[4:-4]
        text = p.read_text(encoding="utf-8", errors="replace")
        lines = text.splitlines()
        if len(lines) < 2:
            header_only += 1
            continue
        head = lines[0].split(",")
        if "symbol" not in head:
            no_symbol.append(day)
            continue
        rows = [dict(zip(head, ln.split(","))) for ln in lines[1:] if ln]
        mids = _mid_at_close(rows, day, offset_s, min_size)
        for sym, m in mids.items():
            truth = rec["funds"].get(sym, {}).get(day)
            if not truth:
                continue
            total_n += 1
            seen.add((sym, day))
            crossed_n += 1 if m["crossed"] else 0
            d = m["mid"] - truth["disclosed_price"]
            exact = abs(d) < HALF_CENT_EQUAL
            f = per_fund.setdefault(sym, {"n": 0, "exact": 0, "calm_n": 0,
                                          "calm_exact": 0, "stress_n": 0,
                                          "stress_exact": 0, "spread_1tick": 0,
                                          "steps": {}})
            f["n"] += 1
            f["exact"] += 1 if exact else 0
            k = "calm" if truth["calm"] else "stress"
            f[f"{k}_n"] += 1
            f[f"{k}_exact"] += 1 if exact else 0
            f["spread_1tick"] += 1 if abs(m["spread_cents"] - 1.0) < 1e-6 else 0
            # **§24's lesson, applied to this reading's own units.** A basis
            # point is not the natural quantum here; the half cent is, because
            # §24.1 measured the disclosed price onto that grid. A discrepancy
            # is therefore counted in half-cent steps, and "one step" is the
            # smallest disagreement the grid allows.
            step = int(round(abs(d) / 0.005))
            f["steps"][str(step)] = f["steps"].get(str(step), 0) + 1
            steps_all[str(step)] = steps_all.get(str(step), 0) + 1
            if truth["nav"]:
                diffs_bp.append(abs(d) / truth["nav"] * 1e4)

    if no_symbol:
        print(f"**{len(no_symbol)} captures carry no symbol column** and were "
              f"skipped, first {no_symbol[0]}. Re-run --nbbo-fetch, which now "
              f"archives them and asks for the column.")
    if not total_n:
        print("no fund-day matched a captured day. Nothing concluded.")
        return {}
    ex = sum(f["exact"] for f in per_fund.values())
    rate = ex / total_n
    cn = sum(f["calm_n"] for f in per_fund.values())
    sn = sum(f["stress_n"] for f in per_fund.values())
    cr = (sum(f["calm_exact"] for f in per_fund.values()) / cn) if cn else None
    sr = (sum(f["stress_exact"] for f in per_fund.values()) / sn) if sn else None
    diffs_bp.sort()
    med_bp = diffs_bp[len(diffs_bp) // 2] if diffs_bp else None
    p90_bp = diffs_bp[int(0.9 * len(diffs_bp))] if diffs_bp else None

    for t in sorted(per_fund):
        f = per_fund[t]
        print(f"{t:6s} exact {f['exact']:4d}/{f['n']:4d} = "
              f"{f['exact'] / f['n']:.4f}   1-tick spread "
              f"{f['spread_1tick'] / f['n']:.3f}   steps "
              + " ".join(f"{k}:{v}" for k, v in sorted(f["steps"].items(),
                                                       key=lambda kv: int(kv[0]))))
    print("\n§36  discrepancy in half-cent steps, pooled: "
          + " ".join(f"{k}:{v}" for k, v in sorted(steps_all.items(),
                                                   key=lambda kv: int(kv[0]))))
    print("     (a step is the smallest disagreement the grid allows. "
          "§24.1 put the disclosed price on that grid, so this is the unit.)")
    print(f"\n§36.4  exact-match rate {ex}/{total_n} = **{rate:.4f}**")
    print(f"§36.5  calm {cr if cr is None else round(cr, 4)}   "
          f"stress {sr if sr is None else round(sr, 4)}   "
          f"gap {None if (cr is None or sr is None) else round(cr - sr, 4)}")
    print(f"§36.5  |discrepancy| median {med_bp and round(med_bp, 3)} bp, "
          f"90th {p90_bp and round(p90_bp, 3)} bp, "
          f"**against λ at 1.2 to 1.7 bp**")
    print(f"§36.6  crossed or locked at the mark: {crossed_n}/{total_n}")
    unmatched = len(registered - seen)
    print(f"§49.5  registered fund-days {len(registered)}, compared {total_n}, "
          f"**never compared {unmatched}**"
          + (f", header-only captures {header_only}" if header_only else ""))

    if rate >= 0.99 and (sr is None or cr is None or sr >= 0.99):
        cell = "passes"
    elif rate < 0.90:
        cell = "fails"
    elif sr is not None and cr is not None and sr < cr - 0.02:
        cell = "fails_on_stress_clustering"
    else:
        cell = "passes_on_subtest"
    print(f"§36.5  VERDICT: {cell}")

    out = {"stage": "B9-A-6 (§36)", "diagnostic_only": False,
           "dataset": dataset, "offset_s": offset_s, "min_size": min_size,
           "exact_rate": rate, "n": total_n, "calm_rate": cr, "stress_rate": sr,
           "median_abs_bp": med_bp, "p90_abs_bp": p90_bp,
           "crossed_at_mark": crossed_n, "verdict": cell, "funds": per_fund,
           "registered_fund_days": len(registered), "never_compared": unmatched,
           "header_only_captures": header_only,
           "halfcent_steps": steps_all,
           "one_tick_spread_rate": (sum(f["spread_1tick"] for f in per_fund.values())
                                    / total_n)}
    (ROOT / "results" /
     f"b9_nbbo_overlap_{dataset.replace('.', '_')}"
     f"{('_sz' + str(min_size)) if min_size else ''}.json").write_text(
        json.dumps(out, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print(f"wrote results/b9_nbbo_overlap_{dataset.replace('.', '_')}"
          f"{('_sz' + str(min_size)) if min_size else ''}.json")
    return out


def _read_day(ddir: Path, day: str) -> list:
    p = ddir / f"bbo-{day}.csv"
    if not p.is_file():
        return []
    lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    if len(lines) < 2:
        return []
    head = lines[0].split(",")
    if "symbol" not in head:
        return []
    return [dict(zip(head, ln.split(","))) for ln in lines[1:] if ln]


def nbbo_combine(datasets: list, offset_s: int = 0,
                 min_size: int = 0) -> dict:
    """§39.3 step 2. The highest bid and the lowest ask across the given venues
    at the mark.

    **This is not the NBBO and does not claim to be** (§39.4). The question is
    whether it reproduces the disclosed price on the ground truth, which is a
    question with an answer, unlike a question about its formal status."""
    rec = load_days()
    if not rec:
        return {}
    dirs = {ds: nbbo_dir(ds) for ds in datasets}
    missing = [ds for ds, d in dirs.items() if not list(d.glob("bbo-*.csv"))]
    if missing:
        print(f"**no captures for {', '.join(missing)}.** Fetch them first; a "
              f"combination of the venues that happen to be on disk is not the "
              f"combination that was registered.")
        return {}
    per_fund, total_n, exact_n, steps_all, venue_at_inside = {}, 0, 0, {}, {}
    # **§49.5(3).** The coverage guard above only asks whether each dataset has
    # any file at all. A fund-day present in one venue and absent from three is
    # still a one-venue observation, and the combination can lock or cross a
    # book that no single venue showed. Both are now counted.
    venues_per_obs, crossed_n = {}, 0
    for day in rec["days"]:
        book = {}
        for ds, d in dirs.items():
            for sym, m in _mid_at_close(_read_day(d, day), day, offset_s,
                                        min_size).items():
                b = book.setdefault(sym, {"bid": None, "ask": None,
                                          "bid_src": "", "ask_src": ""})
                if b["bid"] is None or m["bid"] > b["bid"]:
                    b["bid"], b["bid_src"] = m["bid"], ds
                if b["ask"] is None or m["ask"] < b["ask"]:
                    b["ask"], b["ask_src"] = m["ask"], ds
        for sym, b in book.items():
            truth = rec["funds"].get(sym, {}).get(day)
            if not truth or b["bid"] is None or b["ask"] is None:
                continue
            nv = len({b["bid_src"], b["ask_src"]} - {""})
            venues_per_obs[nv] = venues_per_obs.get(nv, 0) + 1
            if b["bid"] >= b["ask"]:
                crossed_n += 1
            mid = (b["bid"] + b["ask"]) / 2.0
            d_ = mid - truth["disclosed_price"]
            exact = abs(d_) < HALF_CENT_EQUAL
            total_n += 1
            exact_n += 1 if exact else 0
            step = str(int(round(abs(d_) / 0.005)))
            steps_all[step] = steps_all.get(step, 0) + 1
            f = per_fund.setdefault(sym, {"n": 0, "exact": 0, "calm_n": 0,
                                          "calm_exact": 0, "stress_n": 0,
                                          "stress_exact": 0})
            f["n"] += 1
            f["exact"] += 1 if exact else 0
            k = "calm" if truth["calm"] else "stress"
            f[f"{k}_n"] += 1
            f[f"{k}_exact"] += 1 if exact else 0
            for side in ("bid_src", "ask_src"):
                venue_at_inside[b[side]] = venue_at_inside.get(b[side], 0) + 1

    if not total_n:
        print("nothing compared.")
        return {}
    rate = exact_n / total_n
    cn = sum(f["calm_n"] for f in per_fund.values())
    sn = sum(f["stress_n"] for f in per_fund.values())
    cr = (sum(f["calm_exact"] for f in per_fund.values()) / cn) if cn else None
    sr = (sum(f["stress_exact"] for f in per_fund.values()) / sn) if sn else None
    for t in sorted(per_fund):
        f = per_fund[t]
        print(f"{t:6s} exact {f['exact']:5d}/{f['n']:5d} = {f['exact'] / f['n']:.4f}")
    print(f"\n§39.3  venues " + ", ".join(datasets)
          + f"   min_size {min_size}")
    print("§39.3  half-cent steps: " + " ".join(
        f"{k}:{v}" for k, v in sorted(steps_all.items(), key=lambda kv: int(kv[0]))[:8]))
    print("§39.3  times each venue held the inside (bid or ask): " + " ".join(
        f"{k}:{v}" for k, v in sorted(venue_at_inside.items(),
                                      key=lambda kv: -kv[1])))
    print("§49.5  distinct venues holding the inside, per observation: "
          + " ".join(f"{k}:{v}" for k, v in sorted(venues_per_obs.items())))
    print(f"§49.5  **the combination locked or crossed the book on "
          f"{crossed_n}/{total_n}**, which no single venue did")
    print(f"§39.3  exact-match rate {exact_n}/{total_n} = **{rate:.4f}**")
    print(f"§39.3  calm {cr and round(cr, 4)}  stress {sr and round(sr, 4)}")
    cell = ("passes" if rate >= 0.99 else
            "fails_and_stops" if rate < 0.99 else "x")
    print(f"§39.3  VERDICT: {cell}"
          + ("" if rate >= 0.99 else
             "  (§39.3's stopping rule: the combining stops at four venues. "
             "A construction tuned until it matches is not a measurement.)"))
    out = {"stage": "B9-A-6 four-venue (§39.3)", "diagnostic_only": False,
           "datasets": datasets, "offset_s": offset_s,
           "min_size": min_size, "exact_rate": rate,
           "n": total_n, "calm_rate": cr, "stress_rate": sr,
           "halfcent_steps": steps_all, "venue_at_inside": venue_at_inside,
           "venues_per_observation": venues_per_obs, "crossed_combined": crossed_n,
           "verdict": cell, "funds": per_fund}
    (ROOT / "results" /
     f"b9_nbbo_combined{('_sz' + str(min_size)) if min_size else ''}.json"
     ).write_text(
        json.dumps(out, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print(f"wrote results/b9_nbbo_combined"
          f"{('_sz' + str(min_size)) if min_size else ''}.json")
    return out


RECON_OUT = ROOT / "results" / "b9_recon_premium.json"


def nbbo_export(dataset: str = DB_DATASET, offset_s: int = 0,
                min_size: int = 0) -> dict:
    """§40.3 step 0. The reconstructed premium per fund-day, written where the
    stage can read it.

    **The NAV comes from the stage's own export**, so the reconstruction differs
    from the disclosed series in the price and in nothing else. A fund-day whose
    venue record is missing is **absent rather than filled**, and the count of
    absences is reported, because a silently filled day would make the
    comparison look better than the data is."""
    rec = load_days()
    if not rec:
        return {}
    ddir = nbbo_dir(dataset)
    out, missing = {}, 0
    for day in rec["days"]:
        mids = _mid_at_close(_read_day(ddir, day), day, offset_s, min_size)
        for sym, truth in rec["funds"].items():
            if day not in truth:
                continue
            m = mids.get(sym)
            if not m or not truth[day]["nav"]:
                missing += 1
                continue
            out.setdefault(sym, {})[day] = m["mid"] / truth[day]["nav"] - 1.0
    have = sum(len(v) for v in out.values())
    print(f"§40.3  reconstructed {have} fund-days from {dataset}, "
          f"{missing} absent (absent, not filled)")
    for t in sorted(out):
        n_t = len(rec["funds"].get(t, {}))
        print(f"  {t:6s} {len(out[t]):4d}/{n_t:4d}")
    payload = {"stage": "B9 §40.3 step 0", "diagnostic_only": True,
               "diagnostic_reason": (
                   "§40.3 step 0. Fund-day premium reconciled at one fixed offset, "
                   "with the days it cannot fill marked absent rather than filled. The "
                   "premium table and its counts only; no verdict."
               ),
               "dataset": dataset, "offset_s": offset_s, "min_size": min_size,
               "n_fund_days": have, "absent": missing, "premium": out}
    RECON_OUT.write_text(json.dumps(payload, sort_keys=True,
                                    ensure_ascii=False, indent=1) + "\n",
                         encoding="utf-8", newline="\n")
    print(f"wrote {RECON_OUT.relative_to(ROOT)}")
    return payload


def nbbo_sweep(dataset: str = DB_DATASET) -> dict:
    """§38.4(a). **Costs nothing**: the captures are ten seconds wide, so the
    mark can be moved without fetching. Separates a wrong instant from a wrong
    venue, which §38.1 cannot do on its own."""
    rec = load_days()
    if not rec:
        return {}
    ddir = nbbo_dir(dataset)
    files = sorted(ddir.glob("bbo-*.csv"))
    if not files:
        print(f"no captures under {ddir.relative_to(ROOT)}.")
        return {}
    parsed = []
    for p in files:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
        if len(lines) < 2:
            continue
        head = lines[0].split(",")
        if "symbol" not in head:
            continue
        parsed.append((p.name[4:-4],
                       [dict(zip(head, ln.split(","))) for ln in lines[1:] if ln]))
    out = {}
    for off in range(-5, 6):
        ex = tot = 0
        for day, rows in parsed:
            for sym, m in _mid_at_close(rows, day, off).items():
                truth = rec["funds"].get(sym, {}).get(day)
                if not truth:
                    continue
                tot += 1
                ex += 1 if abs(m["mid"] - truth["disclosed_price"]) < HALF_CENT_EQUAL else 0
        out[off] = {"exact": ex, "n": tot, "rate": (ex / tot) if tot else None}
        print(f"  offset {off:+d}s  exact {ex:5d}/{tot:5d} = "
              f"{out[off]['rate'] if out[off]['rate'] is None else round(out[off]['rate'], 4)}")
    best = max((v["rate"] or 0, k) for k, v in out.items())
    at0 = out[0]["rate"] or 0
    zero_is_best = (best[0] - at0) < 0.01
    # **"flat" was the wrong word and is not used.** A rate that is equal at
    # every offset and a rate that spikes at zero both leave zero as the best,
    # and they say different things about the instrument. The drop to the
    # neighbours is what separates them, so it is reported.
    nb = max((out[k]["rate"] or 0) for k in (-1, 1))
    sharp = at0 - nb
    print(f"\n§38.4  best offset {best[1]:+d}s at {best[0]:.4f}, "
          f"offset 0 at {at0:.4f}, gap {best[0] - at0:.4f}")
    print(f"§38.4  drop to the better neighbour: {sharp:+.4f} "
          + ("(**a sharp peak at the mark**, so the instant is right to the "
             "second)" if sharp > 0.05 else
             "(no peak; zero is best but the rate barely moves, which is a "
             "weaker statement about the instant)"))
    print(f"§38.4  VERDICT: "
          + ("the instant is right and the venue is wrong"
             if zero_is_best else
             f"**the instant is wrong by {best[1]:+d}s**; §38.1's venue verdict "
             f"is void until it is re-run at the peak"))
    res = {"stage": "B9 §38.4(a) sampling-instant sweep", "diagnostic_only": True,
           "diagnostic_reason": (
               "§38.4(a). A sweep of the sampling instant: the rate at each offset "
               "and where the peak is. That is the input to whether §38.1's venue "
               "verdict has to be re-run at the peak, and not a reading of its "
               "own."
           ),
           "dataset": dataset, "by_offset": {str(k): v for k, v in out.items()},
           "best_offset": best[1], "best_rate": best[0], "rate_at_zero": at0,
           "zero_is_best": zero_is_best, "neighbour_best": nb,
           "peak_sharpness": sharp}
    (ROOT / "results" / "b9_nbbo_sweep.json").write_text(
        json.dumps(res, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    print("wrote results/b9_nbbo_sweep.json")
    return res


def daily_capture(force: bool = False) -> dict:
    """One file a day. Idempotent: today's capture already on disk is a no-op,
    so this is safe to run from a scheduler that fires more than once."""
    day = datetime.now(timezone.utc).date().isoformat()
    name = f"spdr-product-data-{day}.xlsx"
    DAILY.mkdir(parents=True, exist_ok=True)
    target = DAILY / name
    # **No early return.** An earlier version returned here, which meant a day
    # whose workbook succeeded and whose pdhist failed could never be retried:
    # the second run saw the workbook and stopped. Each item is skipped on its
    # own, and the function always reaches every item.
    existing_kind = (payload_kind(target.read_bytes())
                     if target.is_file() else None)
    if target.is_file() and not force and existing_kind == "xlsx_zip":
        print(f"{day}: workbook already captured, {target.stat().st_size} bytes.")
        rec, kind, verdict = ({"status": "already_present",
                               "bytes": target.stat().st_size},
                              "xlsx_zip", "ok_tabular")
    else:
        # **§49.5(1).** The first version asserted `ok_tabular` as a literal and
        # never re-read the bytes, so an HTML error page written under the
        # workbook's name cemented itself and was counted in the index. The
        # pdhist branch below was fixed and this one was not.
        if target.is_file() and existing_kind != "xlsx_zip" and not force:
            stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
            target.rename(DAILY / f"{name}.expired_{stamp}_{existing_kind}")
            for side in (".meta.json", ".partial"):
                sp = DAILY / f"{name}{side}"
                if sp.is_file():
                    sp.rename(DAILY / f"{name}{side}.expired_{stamp}_{existing_kind}")
            print(f"{day}: previous workbook was {existing_kind}, archived and "
                  f"retried")
        rec = fetch(DAILY_URL, "_daily", name, force=force)
        kind = "none"
        final, partial, _ = cache_paths("_daily", name)
        if final.is_file():
            kind = payload_kind(final.read_bytes())
        verdict = classify_response(rec.get("status", ""), kind)
        print(f"{day}: {rec.get('status')} {rec.get('bytes')} bytes, {kind}, {verdict}")
        if verdict != "ok_tabular":
            print("    **not a workbook.** Kept on disk anyway; nothing is "
                  "deleted, and a bad capture is evidence about that day, not "
                  "a reason to overwrite it.")

    # --- §34.7: the premium history, one file per fund ---------------------
    # The workbook above is fetched once for all funds; this is one request per
    # fund because SSGA publishes `pdhist` per ticker. **The filename carries
    # both the ticker and the day**, so neither a second fund nor a second day
    # can land on an existing capture.
    pd_out = {}
    for t in capture_set():
        pname = f"pdhist-{t}-{day}.xlsx"
        ptarget = DAILY / pname
        if ptarget.is_file() and not force:
            existing = payload_kind(ptarget.read_bytes())
            if existing == "xlsx_zip":
                pd_out[t] = {"status": "already_present",
                             "bytes": ptarget.stat().st_size}
                continue
            # **A bad capture must not cement itself.** The guard skipped on
            # existence alone, so sixteen HTML error pages would have blocked
            # this day permanently and the retry would have looked idempotent.
            # The bad bytes are kept under their own name; nothing is deleted.
            stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
            ptarget.rename(DAILY / f"{pname}.expired_{stamp}_{existing}")
            for side in (".meta.json", ".partial"):
                sp = DAILY / f"{pname}{side}"
                if sp.is_file():
                    sp.rename(DAILY / f"{pname}{side}.expired_{stamp}_{existing}")
            print(f"    {t}: previous capture was {existing}, archived and "
                  f"retried")
        prec = fetch(pdhist_url(t), "_daily", pname, force=force)
        pfinal, _, _ = cache_paths("_daily", pname)
        pkind = payload_kind(pfinal.read_bytes()) if pfinal.is_file() else "none"
        pd_out[t] = {"status": prec.get("status"), "bytes": prec.get("bytes"),
                     "kind": pkind,
                     "verdict": classify_response(prec.get("status", ""), pkind)}
    ok = sum(1 for v in pd_out.values()
             if v.get("verdict") == "ok_tabular" or v["status"] == "already_present")
    bad = sorted(t for t, v in pd_out.items()
                 if v.get("verdict") not in ("ok_tabular", None)
                 and v["status"] != "already_present")
    print(f"{day}: pdhist {ok}/{len(pd_out)} ok"
          + (f", **not a workbook**: {', '.join(bad)}" if bad else ""))

    # An index of what exists, so a gap in the series is visible without a glob.
    caps = sorted(p.name for p in DAILY.glob("spdr-product-data-*.xlsx"))
    pdays = sorted({p.name.rsplit("-", 3)[-3] + "-" + p.name.rsplit("-", 3)[-2]
                    + "-" + p.name.rsplit("-", 3)[-1][:-5]
                    for p in DAILY.glob("pdhist-*-*.xlsx")})
    (DAILY / "capture_index.json").write_text(
        json.dumps({"captures": caps, "count": len(caps),
                    "first": caps[0] if caps else "", "last": caps[-1] if caps else "",
                    "pdhist_days": pdays, "pdhist_day_count": len(pdays),
                    "pdhist_funds_per_day": len(pd_out),
                    "note": "Section 13.3 item 4. There is no spread between 2025-01-02 and the "
                            "first capture, and that is a fact about this run rather than a "
                            "gap that can be filled. Section 34.7 correction: pdhist is a "
                            "rolling 404 days and drops off the back if it is not stored, so "
                            "whether the window can be lengthened depends on this and not on "
                            "the table above."},
                   sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8", newline="\n")
    return {"date": day, "status": rec.get("status"), "kind": kind,
            "verdict": verdict, "pdhist": pd_out, "pdhist_ok": ok}


# ---------------------------------------------------------------------------
# report
# ---------------------------------------------------------------------------


def load_discovered() -> dict:
    out = {}
    for fund in FUNDS:
        p = CACHE / fund["ticker"] / "discovered_links.json"
        if p.is_file():
            out[fund["ticker"]] = json.loads(p.read_text(encoding="utf-8"))
    return out


def build_record(discovered: dict) -> dict:
    per_fund = {}
    for fund in FUNDS:
        t = fund["ticker"]
        d = discovered.get(t)
        row = {k: fund[k] for k in ("issuer", "arm", "structure", "underlying")}
        if d is None:
            row["landing_status"] = "not_attempted"
            row["fields"] = {f["id"]: "not_attempted" for f in FIELDS}
        else:
            row["landing_status"] = d["landing"].get("status", "unknown")
            got = {}
            for f in FIELDS:
                hits = [c["url"] for c in d["candidates"] if f["id"] in c["fields"]]
                got[f["id"]] = ("candidate_link" if hits else "no_candidate_link")
            if d.get("third_party") and d["third_party"].get("status", "").startswith("http_2"):
                got["close_history"] = "third_party_" + d["third_party"].get("granularity", "?")
            row["fields"] = got
            row["candidate_count"] = len(d["candidates"])
        per_fund[t] = row

    return {
        "stage": "B9-A",
        "step": "availability, docs/b9_zero_holonomy.md §10 step 1",
        "schema_version": 1,
        "diagnostic_only": True,
        "diagnostic_reason": (
            "This station answers only which fields are obtainable on free terms and "
            "whether what is obtained is history or a snapshot. It constructs no omega, "
            "computes no lambda and no pi, and carries none of the readings B9-0 through "
            "B9-B-2. B9-A-1's zero has not been computed and no quantity in this file "
            "enters the evidence."
        ),
        "fields": FIELDS,
        "excluded_by_section_7": EXCLUDED_BY_SECTION_7,
        "funds": per_fund,
        "open_questions": [
            "Whether median_bid_ask_spread is history or a 30-day snapshot. Section 12.1 "
            "requires the square root of N to land on the fund-days that survive F1, and a "
            "snapshot cannot do that. This is the headline item for this step.",
            "The structure column is entirely unverified: 6c-11 covers open-end funds, and "
            "SPY, QQQ and DIA are UITs predating that rule. It has to be earned from each "
            "fund's own filings, not asserted by a comment in this file.",
            "The landing column holds candidate URLs, several of them written from memory. "
            "A 404 is an input error and not a finding about the issuer.",
        ],
    }


def render_md(rec: dict) -> str:
    L: list[str] = []
    L.append("# B9-A availability: per field, per fund\n")
    L.append("Generated by `experiments/b9a_availability.py`. Registered in "
             "`docs/b9_zero_holonomy.md` §10 step 1. **Diagnostic only**: no `ω`, "
             "no `λ`, no `π`, no prediction read.\n")

    L.append("\n## What each field is for\n")
    L.append("| field | needed by | granularity required |")
    L.append("|---|---|---|")
    for f in rec["fields"]:
        L.append(f"| `{f['id']}` | {f['needed_by']} | **{f['granularity_required']}** |")

    for arm, title in (("contemporaneous", "Main arm, §6.1 and §13.5"),
                       ("comparison", "Comparison arm, §6.1, reported beside"),
                       ("neither", "Neither arm, carried to record the regime question"),
                       ("retired_availability",
                        "Retired by availability, §13.5. In no arm and in no "
                        "reading, kept because nothing is deleted")):
        rows = sorted((t, r) for t, r in rec["funds"].items() if r["arm"] == arm)
        if not rows:
            continue
        L.append(f"\n## {title}\n")
        head = "| ticker | issuer | structure | landing |" + "".join(
            f" `{f['id']}` |" for f in rec["fields"])
        L.append(head)
        L.append("|---|---|---|---|" + "---|" * len(rec["fields"]))
        for t, r in rows:
            cells = "".join(f" {r['fields'][f['id']]} |" for f in rec["fields"])
            L.append(f"| **{t}** | {r['issuer']} | {r['structure']} | "
                     f"{r['landing_status']} |{cells}")

    L.append("\n## Excluded by §7, declared before retrieval\n")
    for e in rec["excluded_by_section_7"]:
        L.append(f"- {e}")

    L.append("\n## Open, and none of them is decided by this file\n")
    for q in rec["open_questions"]:
        L.append(f"- {q}")

    L.append("\n## What this step does not decide\n")
    L.append("- **It does not run B9-0.** That is §10 step 3 and it needs the `ω` "
             "construction of step 2, which is not written.")
    L.append("- **It does not answer B9-B-1.** Trade-level data with size is a "
             "separate gate at §10 step 5, and a failure there is not a failure "
             "of B9 (§8).")
    L.append("- **A `candidate_link` is not data.** It is a URL whose keyword "
             "matched. Whether it carries a usable history is decided by "
             "fetching and reading it, which this version does not do.")
    return "\n".join(L) + "\n"


# ---------------------------------------------------------------------------
# selftest, no network
# ---------------------------------------------------------------------------

#: Both traps in this fixture came off the first real IVV run: a stylesheet
#: called `navigation-*.css` classified as NAV, and the actual data endpoint
#: absent from every anchor tag.
FIXTURE_HTML = """
<html><head>
<link rel="stylesheet" href="/_astro/navigation-CX4YlZ4h-DDKyWds4.css">
<link rel="stylesheet" href="/_astro/holdings-v3-3kRmXYZm.css">
</head><body>
<a href="/us/products/1/nav-history.csv">NAV history</a>
<a href="https://cdn.example.com/x/median-bid-ask-spread.csv">spread</a>
<a href="/docs/statement-of-additional-information.pdf">SAI</a>
<a href="/premium-discount.xlsx">premium and discount</a>
<a href="#top">top</a>
<a href="/about">about</a>
<a href="/us/products/1/nav-history.csv">NAV history again</a>
<img src="/img/chart-navigator.png">
<script>window.__D={"fundNav":"/us/products/239726/x/1467271812596.ajax?fileType=csv&amp;fileName=IVV_NAV&amp;dataType=fund"};</script>
</body></html>
"""

FIXTURE_CSV_HISTORY = "Date,Close\n2024-01-02,10.5\n2024-01-03,10.6\n"
FIXTURE_CSV_SNAPSHOT = "Median bid-ask spread (30 days): 0.02%\n"


def selftest() -> int:
    fails: list[str] = []

    def check(name: str, cond: bool, detail: str = "") -> None:
        print(f"  {'ok  ' if cond else 'FAIL'} {name}{(' :: ' + detail) if detail else ''}")
        if not cond:
            fails.append(name)

    print("classifier")
    ex = extract_links(FIXTURE_HTML, "https://example.com/us/products/1")
    links = ex["candidates"]
    urls = [c["url"] for c in links]
    check("dedupes repeated href", len(urls) == len(set(urls)), str(len(urls)))
    check("drops #anchor and untagged /about",
          not any(u.endswith("#top") or u.endswith("/about") for u in urls))
    by_field = {f: [c["url"] for c in links if f in c["fields"]] for f in LINK_KEYWORDS}
    check("spread link found", len(by_field["median_bid_ask_spread"]) == 1)
    check("SAI reaches creation_fee", len(by_field["creation_fee"]) == 1)
    check("relative href absolutised",
          all(u.startswith("http") for u in urls), urls[0] if urls else "")

    print("classifier, the two traps from the first IVV run")
    check("navigation-*.css does not reach nav_history",
          not any("navigation" in u.lower() for u in by_field["nav_history"]),
          str(by_field["nav_history"]))
    check("holdings-v3-*.css does not reach holdings",
          not any(u.lower().endswith(".css") for u in by_field["holdings"]))
    check("assets dropped and counted", ex["dropped_assets"] >= 3,
          str(ex["dropped_assets"]))
    ajax = [u for u in urls if ".ajax?" in u]
    check("ajax endpoint pulled out of a script blob", len(ajax) == 1,
          ajax[0] if ajax else "none")
    check("&amp; unescaped in the extracted url",
          bool(ajax) and "&amp;" not in ajax[0] and "fileName=IVV_NAV" in ajax[0])
    check("the ajax endpoint classifies as nav_history",
          bool(ajax) and "nav_history" in classify_link(ajax[0]),
          str(classify_link(ajax[0])) if ajax else "")
    check("the two real csv anchors survive",
          len([u for u in urls if u.endswith("nav-history.csv")]) == 1)
    check("nothing is silently gone",
          ex["urls_seen"] >= len(urls) + ex["dropped_assets"])

    print("asset scan")
    fixture_js = (
        'const P="/us/products/239726/ishares-core-sp-500-etf";'
        'const U=P+"/1467271812596.ajax?fileType=csv&fileName=IVV_NAV&dataType=fund";'
        'fetch(U);const H="latest-holdings.csv";'
    )
    sc = scan_text_for_endpoints(fixture_js)
    check("ajax endpoint read out of a bundle", len(sc["ajax_endpoint"]) == 1,
          sc["ajax_endpoint"][0] if sc["ajax_endpoint"] else "none")
    check("fileName parameter captured", sc["file_name_param"] == ["fileName=IVV_NAV"])
    check("dataType parameter captured", sc["data_type_param"] == ["dataType=fund"])
    check("product path captured", len(sc["product_path"]) >= 1)
    check("csv path captured", "latest-holdings.csv" in sc["csv_path"])
    check("a bundle with no endpoint yields empty lists, not an error",
          all(v == [] for v in scan_text_for_endpoints("var a=1;").values()))
    assets = extract_assets(
        '<script src="/_astro/x.js"></script><script src="https://cdn.other/y.js">'
        '</script><a href="/_astro/z.mjs">z</a>',
        "https://www.ishares.com/us/products/1")
    check("same-origin js only", assets == ["https://www.ishares.com/_astro/x.js",
                                            "https://www.ishares.com/_astro/z.mjs"],
          str(assets))

    print("constructed candidates")
    fake = CACHE / "ZZTEST"
    fake.mkdir(parents=True, exist_ok=True)
    (fake / "discovered_links.json").write_text(json.dumps({"candidates": [
        {"url": "https://www.ishares.com/us/products/239726/x/latest-holdings.csv",
         "fields": ["holdings"]},
        {"url": "https://www.ishares.com/us/literature/y.pdf", "fields": ["creation_fee"]},
    ]}, sort_keys=True) + "\n", encoding="utf-8", newline="\n")
    check("product path derived from the measured holdings link",
          product_base("ZZTEST") == "https://www.ishares.com/us/products/239726/x",
          str(product_base("ZZTEST")))
    check("no holdings link means no derivation, not a guess",
          product_base("ZZNOTHERE") is None)
    specs = CONSTRUCTED_ENDPOINTS["iShares"]
    ctl = [s for s in specs if s.get("control")]
    check("two controls declared", len(ctl) == 2, str([s["id"] for s in ctl]))
    check("one control is on the ajax shape itself",
          any(".ajax?" in s["path"] for s in ctl))
    check("one control is on a different shape, so they can disagree",
          any(".ajax?" not in s["path"] for s in ctl))
    check("every non-control line records why it is there",
          all(("unverified" in s["source"] or "logged with both reported" in s["source"])
              for s in specs if not s.get("control")))
    check("ids are unique", len({s["id"] for s in specs}) == len(specs))
    built = next(s for s in specs if s["id"] == "control_ajax_holdings"
                 )["path"].format(prod="P", T="IVV")
    check("the documented control builds the documented url",
          built == "P/1467271812596.ajax?fileType=csv&fileName=IVV_holdings"
                   "&dataType=fund", built)

    print("granularity")
    check("two dated rows read as history",
          granularity_of_payload(FIXTURE_CSV_HISTORY) == "history")
    check("single figure reads as snapshot",
          granularity_of_payload(FIXTURE_CSV_SNAPSHOT) == "snapshot")
    check("empty reads as none", granularity_of_payload("") == "none")
    check("Aug 15, 2026 style dates read as history",
          granularity_of_payload("as of\nAug 15, 2026,1\nAug 14, 2026,2\n") == "history")
    # The regression that produced four wrong 'snapshot' labels: a metadata
    # block long enough that no date falls in the first 2,000 bytes.
    buried = "Fund Name,iShares\n" + ("filler,x\n" * 300) + FIXTURE_CSV_HISTORY
    check("a date past the first 2000 bytes still reads as history",
          len(buried) > 2000 and granularity_of_payload(buried) == "history",
          f"{len(buried)} bytes")
    prof = payload_profile(buried)
    check("header line located above the first dated row",
          prof["header_line"].startswith("Date,Close"), prof["header_line"][:40])
    check("distinct dates counted over the whole file", prof["distinct_dates"] == 2)

    print("xlsx dates are text, which the first version missed")
    check("SSGA's 14-Aug-2026 parses", as_iso_date("14-Aug-2026") == "2026-08-14")
    check("the product-data sheet's Aug 13 2026 parses",
          as_iso_date("Aug 13 2026") == "2026-08-13")
    check("a real workbook date still parses",
          as_iso_date(datetime(2026, 8, 14)) == "2026-08-14")
    check("a fund name is not a date", as_iso_date("State Street® XLF") is None)
    check("a number is not a date", as_iso_date(58.2388) is None)
    check("iso strings sort, so min/max are the span",
          sorted(["2026-08-14", "2010-01-04"])[0] == "2010-01-04")

    print("payload kind, the trap that ate the last run")
    check("a doctype page is html whatever the content type says",
          payload_kind(b'<!DOCTYPE html>\n<html><head>') == "html")
    check("an anti-bot challenge is html too",
          payload_kind(b'<!DOCTYPE html><html><head><meta charset="utf-8">'
                       b'<noscript>This site requires JavaScript') == "html")
    check("a csv is delimited", payload_kind(FIXTURE_CSV_HISTORY.encode()) == "delimited")
    check("json is json", payload_kind(b'{"a":1}') == "json")
    check("200 plus html is a soft 404, not data",
          classify_response("http_200", "html") == "soft_404_html")
    check("200 plus a table is the only thing that counts",
          classify_response("http_200", "delimited") == "ok_tabular")
    check("404 stays non_2xx", classify_response("http_404", "html") == "non_2xx")
    check("an xlsx is a zip container and reads as one",
          payload_kind(b"PK\x03\x04\x14\x00\x06\x00") == "xlsx_zip")
    check("an xlsx counts as a table, not as other",
          classify_response("http_200", "xlsx_zip") == "ok_tabular")
    ssga = CONSTRUCTED_ENDPOINTS["SSGA"]
    check("SSGA templates are absolute and need no product path",
          all(s.get("abs") for s in ssga))
    check("SSGA has its own positive control",
          any(s.get("control") for s in ssga))
    built_ssga = next(s for s in ssga if s["id"] == "premium_discount_history"
                      )["path"].format(prod=None, T="XLF", t="xlf")
    check("lowercase ticker substituted for SSGA",
          built_ssga.endswith("pdhist-us-en-xlf.xlsx"), built_ssga[-40:])

    print("echo versus discovery")
    asked_url = ("https://x/us/products/239726/y/1467271812596.ajax"
                 "?fileType=csv&fileName=IVV_NAV&dataType=fund")
    page = (f'<html><head><link rel="canonical" href="{asked_url}"/></head>'
            '<body><a href="/us/products/239726/y/1521942788811.ajax'
            '?fileType=csv&amp;fileName=IVV_other&amp;dataType=fund">other</a>'
            '</body></html>')
    found = sorted({unescape(u) for u in
                    scan_text_for_endpoints(page)["ajax_endpoint"]
                    if "fileType" in u or "fileName" in u})
    ech = [u for u in found if u.split("?", 1)[-1] in asked_url]
    new = [u for u in found if u not in ech]
    check("the echoed request is recognised as an echo", len(ech) == 1, str(ech))
    check("a genuinely different ajax id survives as new", len(new) == 1, str(new))
    check("the new one carries the other numeric id",
          bool(new) and "1521942788811" in new[0])
    check("an unfamiliar fileName is no longer filtered out",
          "unclassified_download" in classify_link(new[0]) or
          any(f in classify_link(new[0]) for f in LINK_KEYWORDS),
          str(classify_link(new[0])))

    print("payload identity")
    a = "Fund,IVV_NAV\nDate,Close\n2024-01-02,10.5\n2024-01-03,10.6\n"
    b = "Fund,IVV_premium_discount\nDate,Close\n2024-01-02,10.5\n2024-01-03,10.6\n"
    c = "Fund,IVV_NAV\nDate,Close\n2024-01-02,99.9\n2024-01-03,10.6\n"
    check("two names, one file, hashes the same after normalising",
          normalised_digest(a, "IVV") == normalised_digest(b, "IVV"))
    check("a lowercase echo is normalised too",
          normalised_digest(a, "IVV")
          == normalised_digest(b.replace("IVV_", "ivv_"), "IVV"))
    d = first_difference(a.encode(), b.encode())
    check("first difference located at the echoed name",
          not d["identical"] and d["offset"] == len("Fund,IVV_"), str(d.get("offset")))
    check("identical payloads report identical",
          first_difference(a.encode(), a.encode())["identical"])
    short = first_difference(b"abc", b"abcdef")
    check("a pure length difference is labelled as one",
          short["differs_only_in_length"] and short["offset"] == 3)
    check("a real content difference still separates them",
          normalised_digest(a, "IVV") != normalised_digest(c, "IVV"))
    check("raw hashes do differ, which is why the raw one is not the test",
          _sha256(a.encode()) != _sha256(b.encode()))

    print("cache")
    sandbox = CACHE / "_selftest"
    sandbox.mkdir(parents=True, exist_ok=True)
    final, partial, meta = sandbox / "x.csv", sandbox / "x.csv.partial", sandbox / "x.csv.meta.json"
    body = FIXTURE_CSV_HISTORY.encode()
    final.write_bytes(body)
    _write_meta(meta, {"sha256": _sha256(body), "status": "http_200"})
    check("sidecar round-trip", cached_ok(final, meta))
    final.write_bytes(body + b"tampered")
    check("hash mismatch is not read as cached", not cached_ok(final, meta))
    final.write_bytes(body)
    check("restored payload reads as cached", cached_ok(final, meta))
    partial.write_bytes(b"half")
    check("a .partial next to a good payload changes nothing",
          cached_ok(final, meta) and partial.is_file())

    print("record")
    rec = build_record({})
    check("diagnostic_only present from the first version",
          rec.get("diagnostic_only") is True)
    check("diagnostic_reason is a sentence, not a placeholder",
          len(rec.get("diagnostic_reason", "")) > 40)
    check("every fund starts as not_attempted",
          all(r["landing_status"] == "not_attempted" for r in rec["funds"].values()))
    a = json.dumps(rec, sort_keys=True, ensure_ascii=False, indent=2)
    b = json.dumps(build_record({}), sort_keys=True, ensure_ascii=False, indent=2)
    check("record renders byte-identical twice", a == b)
    md1, md2 = render_md(rec), render_md(build_record({}))
    check("markdown renders byte-identical twice", md1 == md2)
    check("markdown carries no wall-clock",
          not re.search(r"\b20\d\d-\d\d-\d\dT", md1))

    print("sample, after §13.5")
    arms = {f["arm"] for f in FUNDS}
    check("four arms declared, retirement among them",
          arms == {"contemporaneous", "comparison", "neither", "retired_availability"},
          str(sorted(arms)))
    main = [f for f in FUNDS if f["arm"] == "contemporaneous"]
    check("main arm is the eleven sector SPDRs", len(main) == 11, str(len(main)))
    check("main arm is one issuer, which §13.3 item 2 relies on",
          {f["issuer"] for f in main} == {"SSGA"})
    check("comparison arm is non-empty and also SSGA",
          {f["issuer"] for f in FUNDS if f["arm"] == "comparison"} == {"SSGA"})
    retired = [f for f in FUNDS if f["arm"] == "retired_availability"]
    check("the retired rows are still declared, not deleted", len(retired) >= 10,
          str(len(retired)))
    check("no retired fund sits in an arm",
          all(f["arm"] == "retired_availability" for f in retired))
    check("every fund has an endpoint table for its issuer",
          all(f["issuer"] in CONSTRUCTED_ENDPOINTS for f in FUNDS
              if f["arm"] != "retired_availability"))
    check("tickers unique", len({f["ticker"] for f in FUNDS}) == len(FUNDS))
    check("the daily capture url is the cross-fund workbook",
          DAILY_URL.endswith("spdr-product-data-us-en.xlsx"))

    print("§36: the closing mark, where a one-hour error would read as the "
          "venue disagreeing with the NBBO")
    check("January is EST", et_offset_hours("2025-01-15") == 5)
    check("July is EDT", et_offset_hours("2025-07-15") == 4)
    check("the Saturday before the second Sunday in March is still EST",
          et_offset_hours("2025-03-08") == 5)
    check("the second Sunday in March is EDT by the close",
          et_offset_hours("2025-03-09") == 4)
    check("the Saturday before the first Sunday in November is still EDT",
          et_offset_hours("2025-11-01") == 4)
    check("the first Sunday in November is EST by the close",
          et_offset_hours("2025-11-02") == 5)
    check("2020 and 2022, the years §36 exists to reach, land in EDT in "
          "March-after and June",
          et_offset_hours("2020-03-16") == 4 and et_offset_hours("2022-06-15") == 4)
    check("winter brackets 21:00 UTC", nbbo_window("2025-01-15")
          == ("2025-01-15T20:59:55", "2025-01-15T21:00:05"))
    check("summer brackets 20:00 UTC", nbbo_window("2025-07-15")
          == ("2025-07-15T19:59:55", "2025-07-15T20:00:05"))
    # **The check with teeth**: the fetch asks for a window and the comparison
    # picks a mark, computed in two places. If they drift the sample is an hour
    # from the close and the venue takes the blame.
    for _d, _m in (("2025-01-15", "2025-01-15T21:00:00"),
                   ("2025-07-15", "2025-07-15T20:00:00"),
                   ("2026-08-13", "2026-08-13T20:00:00")):
        _lo, _hi = nbbo_window(_d)
        check(f"the mark sits inside the fetched window on {_d}",
              _lo <= _m <= _hi)

    check("the epoch-day algorithm agrees with known dates",
          _days_from_civil(1970, 1, 1) == 0
          and _days_from_civil(2000, 3, 1) == 11017
          and _days_from_civil(2025, 1, 2) == 20090)
    # The first capture's own bytes: `ts_recv` on the first row of
    # bbo-2025-01-02.csv reads 1735851595000000000, which is 20:59:55 UTC, and
    # the mark must be the 21:00:00 five seconds after it.
    check("the mark matches the first capture's own window, to the second",
          mark_ns("2025-01-02") == 1735851600 * 10**9)
    check("and it moves an hour in summer, not zero and not two",
          mark_ns("2025-07-15") - mark_ns("2025-01-15")
          == (_days_from_civil(2025, 7, 15) - _days_from_civil(2025, 1, 15))
          * 86400 * 10**9 - 3600 * 10**9)

    _base = 1735851600 * 10**9
    _rows = [
        {"ts_event": str(_base - 2 * 10**9), "symbol": "XLF",
         "bid_px_00": "40000000000", "ask_px_00": "40010000000"},
        {"ts_event": str(_base - 10**8), "symbol": "XLF",
         "bid_px_00": "40010000000", "ask_px_00": "40020000000"},
        {"ts_event": str(_base + 10**9), "symbol": "XLF",
         "bid_px_00": "99000000000", "ask_px_00": "100000000000"},
        {"ts_event": str(_base - 10**8), "symbol": "XLE",
         "bid_px_00": "88020000000", "ask_px_00": "88010000000"},
    ]
    _m = _mid_at_close(_rows, "2025-01-02")
    check("fixed-point nanodollars are scaled, so a price reads as dollars",
          abs(_m["XLF"]["bid"] - 40.01) < 1e-9)
    check("the last record at or before the mark wins",
          abs(_m["XLF"]["mid"] - 40.015) < 1e-12)
    check("a record after the mark is ignored, not merely ranked lower",
          _m["XLF"]["ts"] == _base - 10**8)
    check("a crossed book is flagged rather than silently averaged",
          _m["XLE"]["crossed"] is True)
    check("a row with no symbol is dropped, which is what the first three "
          "captures were made of",
          not _mid_at_close([{"ts_event": str(_base - 10**9),
                              "bid_px_00": "1", "ask_px_00": "2"}],
                            "2025-01-02"))
    check("a one-cent spread gives a mid on the half-cent grid, which is the "
          "grid §24.1 measured the disclosed price on",
          abs((_m["XLF"]["mid"] / 0.005) - round(_m["XLF"]["mid"] / 0.005))
          < 1e-9)
    check("half a cent apart is not equal on the half-cent grid",
          not abs(40.015 - 40.010) < HALF_CENT_EQUAL)
    check("and the same price is",
          abs(40.015 - 40.0150000001) < HALF_CENT_EQUAL)
    # **§49.5(7).** The first version read `db_key() is None or len(...) > 0`,
    # which is true for every possible return value. What is worth asserting is
    # that no key literal sits in the source, which is a fact about the file.
    _src = Path(__file__).read_text(encoding="utf-8", errors="replace")
    _keyish = re.findall(r"db-[A-Za-z0-9]{20,}", _src)
    check("no API key literal is in this file",
          not _keyish, f"{len(_keyish)} key-shaped literals found")
    check("db_key reads the environment and nothing else",
          "os.environ.get(\"DATABENTO_API_KEY\")" in _src
          or "os.environ.get('DATABENTO_API_KEY')" in _src)

    # **§49.5(7).** The two `.css` traps are dropped by extension before
    # `_token_hit` runs, so the word-boundary regex the file header names as a
    # fixed defect had no coverage at all. It is now tested directly.
    check("_token_hit requires a word boundary, not a substring",
          not _token_hit("navigation-CX4YlZ4h.css", "nav")
          and not _token_hit("spreadsheet.js", "spread")
          and not _token_hit("recreation.pdf", "creation"))
    check("and it still fires on a real token",
          _token_hit("nav_history-us-en-xlf.xlsx", "nav")
          and _token_hit("pdhist-us-en-xlf.xlsx", "pdhist"))

    print("§34.7: the premium history archive")
    _reg = [e for e in CONSTRUCTED_ENDPOINTS["SSGA"]
            if e["id"] == "premium_discount_history"]
    check("the capture URL is the endpoint that was measured, not a second "
          "copy of it that can drift",
          len(_reg) == 1 and _reg[0]["path"].replace("{t}", "X")
          == PDHIST_URL.replace("{t}", "X"))
    # **The check above passed while the capture fetched sixteen error pages.**
    # It compared templates. The substitution is where the two paths differed,
    # so the check has to go through the builder that the capture calls.
    check("the built URL matches what try_endpoints would build, substitution "
          "included, which the template comparison did not test",
          pdhist_url("XLRE")
          == _reg[0]["path"].format(prod=None, T="XLRE", t="xlre"))
    check("the ticker goes in lowercase, and the uppercase form is absent",
          "XLRE" not in pdhist_url("XLRE") and "xlre" in pdhist_url("XLRE"))
    check("a bad payload is not mistaken for a good one by the retry guard",
          payload_kind(b"<!doctype html><html>") == "html"
          and payload_kind(b"PK\x03\x04rest") == "xlsx_zip")
    _cs = capture_set()
    check("the archive covers every SSGA fund that is not retired",
          set(_cs) == {f["ticker"] for f in FUNDS if f["issuer"] == "SSGA"
                       and f.get("arm") != "retired_availability"})
    check("and every main-arm fund is in it",
          all(f["ticker"] in _cs for f in FUNDS
              if f.get("arm") == "contemporaneous"))
    check("no retired row sneaks in",
          not any(f["ticker"] in _cs for f in FUNDS
                  if f.get("arm") == "retired_availability"))
    check("the capture set is not empty and is stated as a count",
          len(_cs) >= 11, f"{len(_cs)} funds")
    _n1 = f"pdhist-{'XLF'}-2026-08-17.xlsx"
    _n2 = f"pdhist-{'XLE'}-2026-08-17.xlsx"
    _n3 = f"pdhist-{'XLF'}-2026-08-18.xlsx"
    check("a name carries both ticker and day, so neither another fund nor "
          "another day can land on it", _n1 != _n2 and _n1 != _n3)
    check("the day parses back out of the name",
          _n1.rsplit("-", 3)[-3] + "-" + _n1.rsplit("-", 3)[-2] + "-"
          + _n1.rsplit("-", 3)[-1][:-5] == "2026-08-17")
    check("and it parses for a four-letter ticker too, which is why the split "
          "counts from the right",
          (lambda n: n.rsplit("-", 3)[-3] + "-" + n.rsplit("-", 3)[-2] + "-"
           + n.rsplit("-", 3)[-1][:-5])("pdhist-XLRE-2026-08-17.xlsx")
          == "2026-08-17")

    print(f"\n{len(fails)} failed" if fails else "\nall passed")
    return 1 if fails else 0


# ---------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--selftest", action="store_true", help="no network at all")
    ap.add_argument("--discover", action="store_true",
                    help="fetch each fund's landing page, enumerate candidate links")
    ap.add_argument("--scan-assets", action="store_true",
                    help="read endpoint templates out of the page's own JavaScript")
    ap.add_argument("--try-endpoints", action="store_true",
                    help="fetch the declared constructed candidates, with the "
                         "measured holdings URL as a positive control")
    ap.add_argument("--inspect", action="store_true",
                    help="what the cached payloads turned out to be; no network")
    ap.add_argument("--inspect-xlsx", action="store_true",
                    help="open the cached SSGA workbooks; sheets, header, span")
    ap.add_argument("--nbbo-cost", action="store_true",
                    help="§36.6: estimate the job's cost and fetch nothing")
    ap.add_argument("--nbbo-fetch", action="store_true",
                    help="§36: fetch the closing BBO over the overlap. "
                         "Refuses without --confirm.")
    ap.add_argument("--nbbo-compare", action="store_true",
                    help="§36.4/§36.5: exact-match rate against the disclosed "
                         "price, with the stress-clustering sub-test")
    ap.add_argument("--confirm", action="store_true",
                    help="required by --nbbo-fetch, which spends money")
    ap.add_argument("--limit", type=int, default=0, metavar="N",
                    help="fetch only the first N days, for a trial")
    ap.add_argument("--workers", type=int, default=4, metavar="N",
                    help="concurrent day requests for --nbbo-fetch (default 4)")
    ap.add_argument("--nbbo-sweep", action="store_true",
                    help="§38.4(a): move the sampling instant across the "
                         "captured window. Fetches nothing.")
    ap.add_argument("--nbbo-combine", action="store_true",
                    help="§39.3 step 2: highest bid, lowest ask across "
                         "--datasets, compared to the disclosed price")
    ap.add_argument("--datasets", default="ARCX.PILLAR,XNAS.ITCH,XNYS.PILLAR,"
                                          "BATS.PITCH", metavar="A,B,C",
                    help="venues for --nbbo-combine (§39.3 stops at four)")
    ap.add_argument("--nbbo-export", action="store_true",
                    help="§40.3 step 0: write the reconstructed premium per "
                         "fund-day for b9_omega.py --recon")
    ap.add_argument("--nbbo-datasets", action="store_true",
                    help="§38.4(b): ask the vendor its own coverage ranges")
    ap.add_argument("--list-all", action="store_true",
                    help="with --nbbo-datasets: enumerate the whole catalogue "
                         "first, because a hand-written list cannot discover a "
                         "dataset nobody thought of")
    ap.add_argument("--dataset", default=DB_DATASET, metavar="CODE",
                    help=f"venue dataset for --nbbo-* (default {DB_DATASET})")
    ap.add_argument("--min-size", type=int, default=0, metavar="N",
                    help="§42.4: require N shares on both sides (100 = round "
                         "lot). 0 keeps every quote, which is the default and "
                         "is what §38 to §41 were run with.")
    ap.add_argument("--offset", type=int, default=0, metavar="S",
                    help="seconds to move the mark for --nbbo-compare")
    ap.add_argument("--daily-capture", action="store_true",
                    help="§13.3 item 4: one cross-fund workbook per day, raw, "
                         "idempotent, safe to schedule")
    ap.add_argument("--grep-cells", metavar="REGEX",
                    help="print every cell matching REGEX in the cached workbooks")
    ap.add_argument("--scan-rows", type=int, default=12,
                    help="rows per sheet scanned by --grep-cells")
    ap.add_argument("--max-rows", type=int, default=200_000,
                    help="row cap per sheet; hitting it is printed, never silent")
    ap.add_argument("--head", type=int, default=12,
                    help="lines of the first payload to print under --inspect")
    ap.add_argument("--max-assets", type=int, default=12,
                    help="cap on JS files fetched per fund; what the cap skipped "
                         "is printed and recorded, never silently dropped")
    ap.add_argument("--report", action="store_true",
                    help="build results/b9a_availability.{json,md} from the cache")
    ap.add_argument("--only", action="append", default=[], metavar="TICKER")
    ap.add_argument("--third-party", action="store_true",
                    help="also probe the declared stooq fallback for close prices")
    ap.add_argument("--force", action="store_true",
                    help="re-fetch even when the cached hash still matches")
    args = ap.parse_args()

    if args.selftest:
        return selftest()

    tickers = [t.upper() for t in args.only]
    unknown = sorted(set(tickers) - {f["ticker"] for f in FUNDS})
    if unknown:
        print(f"not in the registered sample: {unknown}", file=sys.stderr)
        return 2

    if args.discover:
        discover(tickers, args.third_party, args.force)
        print("\nnothing selected, nothing parsed. Read the candidates above, "
              f"then fill {ENDPOINTS.relative_to(ROOT)} by hand.")
        return 0

    if args.scan_assets:
        scan_assets(tickers, args.max_assets, args.force)
        print("\nstill nothing selected. A string that looks like an endpoint "
              "is a candidate; whether it serves a usable history is decided by "
              "fetching it, which this mode does not do.")
        return 0

    if args.nbbo_cost:
        nbbo_cost()
        return 0
    if args.nbbo_fetch:
        nbbo_fetch(args.confirm, args.limit, args.workers, args.dataset)
        return 0
    if args.nbbo_compare:
        nbbo_compare(args.dataset, args.offset, args.min_size)
        return 0
    if args.nbbo_combine:
        nbbo_combine([d.strip() for d in args.datasets.split(",") if d.strip()],
                     args.offset, args.min_size)
        return 0
    if args.nbbo_export:
        nbbo_export(args.dataset, args.offset, args.min_size)
        return 0
    if args.nbbo_sweep:
        nbbo_sweep(args.dataset)
        return 0
    if args.nbbo_datasets:
        nbbo_datasets(args.list_all)
        return 0
    if args.daily_capture:
        daily_capture(args.force)
        return 0

    if args.grep_cells:
        grep_cells(tickers, args.grep_cells, args.scan_rows)
        return 0

    if args.inspect_xlsx:
        inspect_xlsx(tickers, args.head, args.max_rows)
        print("\nnothing computed. This mode prints what the workbook is, and "
              "whether it carries the loop's inputs is read off the header.")
        return 0

    if args.inspect:
        inspect(tickers, args.head)
        print("\n**a payload group with more than one name is one file.** The "
              "field it serves is decided by its columns, not by the fileName "
              "that was asked for.")
        return 0

    if args.try_endpoints:
        try_endpoints(tickers, args.force)
        print("\na non-200 above is a constructed URL that did not work. "
              "**It is not evidence that the issuer does not publish the field**, "
              "and it may not be written down as one.")
        return 0

    if args.report:
        rec = build_record(load_discovered())
        OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
        OUT_JSON.write_text(
            json.dumps(rec, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8", newline="\n")
        OUT_MD.write_text(render_md(rec), encoding="utf-8", newline="\n")
        print(f"wrote {OUT_JSON.relative_to(ROOT)} and {OUT_MD.relative_to(ROOT)}")
        return 0

    ap.print_help()
    print("\n--probe is not implemented in this version, on purpose: the "
          "endpoints are not known until --discover has returned and a human "
          "has picked from the candidates.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
